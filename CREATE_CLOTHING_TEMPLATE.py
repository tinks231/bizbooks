"""
🎯 BRIDGE SOLUTION: Clothing Retail Import Template (FINAL VERSION)
====================================================================

Purpose: Temporary Excel template for clothing retail inventory import
         with structured naming that's easy to parse into variants later

Item Name Formula: Brand + Category + Product + Size + Color + Style
Example: "Levi's Jeans 501 32 Blue Slim Fit"

Why This Works:
✅ Clear on invoices (customers know what they're buying)
✅ Searchable (search "jeans" finds all jeans!)
✅ Drag-fill friendly (fast Excel data entry)
✅ Future-proof (easy to parse Brand/Category/Size/Color)
✅ Works with current system (no code changes needed)

Author: Smart solution by Rish! 🎯
Version: 2.0 (FINAL - Category included in Item Name)
Date: December 19, 2025
"""

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.comments import Comment
except ImportError:
    print("❌ Error: openpyxl not installed!")
    print("📦 Installing openpyxl...")
    import subprocess
    subprocess.check_call(['pip3', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.comments import Comment

def create_clothing_retail_template():
    """
    Create Excel template for clothing retail with:
    - Structured input columns (Brand, Category, Product, Size, Color, Style)
    - Auto-generated Item Name (Excel formula)
    - All standard columns from current system
    - Dropdowns for easy data entry
    - Example data
    - Instructions
    """
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Import"
    
    print("🎨 Creating Excel template...")
    
    # ========================================
    # DEFINE COLUMNS
    # ========================================
    
    # Column format: (Name, Description, Required)
    column_structure = [
        # STRUCTURED INPUT COLUMNS (for future parsing)
        ("Brand*", "Brand name: Levi's, Nike, Adidas, H&M, Zara", True, "input"),
        ("Category*", "Product type: Jeans, T-Shirts, Shirts, Shoes, Jackets", True, "input"),
        ("Product Name*", "Model/Style name: 501, Dri-FIT, Air Max, Superstar", True, "input"),
        ("Size*", "Select from dropdown: 28, 30, 32, S, M, L, XL, XXL", True, "input"),
        ("Color*", "Select from dropdown: Blue, Black, Red, White, Grey", True, "input"),
        ("Style/Fit", "Optional: Slim Fit, Regular Fit, Loose Fit, Casual, Formal", False, "input"),
        
        # AUTO-GENERATED COLUMN
        ("🔶 Item Name (Auto)", "Auto-generated: Brand + Category + Product + Size + Color + Style", False, "auto"),
        
        # STANDARD COLUMNS (current system format)
        ("SKU", "Optional: Leave blank to auto-generate unique SKU", False, "optional"),
        ("Barcode", "Optional: Leave blank to auto-generate unique barcode", False, "optional"),
        ("Group*", "Group/Department: Men's Wear, Women's Wear, Kids Wear", True, "input"),
        ("Unit*", "Unit of measurement: pc (pieces), pair, set, kg, ltr", True, "input"),
        ("Stock Quantity*", "Current stock available in your inventory", True, "input"),
        ("Reorder Point", "Optional: Low stock alert threshold (e.g., 10)", False, "optional"),
        ("Cost Price*", "Your purchase/cost price (for profit calculations)", True, "input"),
        ("MRP", "Optional: Maximum Retail Price (if blank, system uses Selling Price)", False, "optional"),
        ("Discount %", "Auto-calculated: ((MRP - Selling Price) / MRP) × 100", False, "auto"),
        ("Selling Price*", "Final price charged to customer", True, "input"),
        ("Tax Rate (%)", "GST rate: 0, 5, 12, 18, or 28", False, "optional"),
        ("HSN/SAC Code", "Optional: HSN code for GST compliance", False, "optional"),
        ("Description", "Optional: Additional product details", False, "optional"),
    ]
    
    # ========================================
    # CREATE HEADER ROW
    # ========================================
    
    for col_idx, (name, description, required, col_type) in enumerate(column_structure, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = name
        
        # Color coding
        if required and col_type == "input":
            # Required columns - Blue
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
        elif col_type == "auto":
            # Auto-generated columns - Orange
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.font = Font(color="000000", bold=True, size=11)
        else:
            # Optional columns - Green
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
        
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.comment = Comment(description, "BizBooks")
    
    # Set column widths
    widths = [20, 15, 20, 10, 12, 15, 45, 15, 15, 15, 10, 15, 12, 12, 12, 12, 15, 10, 15, 25]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    print("✅ Headers created")
    
    # ========================================
    # ADD DATA VALIDATION (Dropdowns)
    # ========================================
    
    # Brand dropdown (common brands)
    brand_values = '"Levi\'s,Nike,Adidas,Puma,Reebok,H&M,Zara,Mango,Pepe Jeans,Wrangler,Lee,Allen Solly,Peter England,Van Heusen,Arrow,Other"'
    brand_dv = DataValidation(type="list", formula1=brand_values, allow_blank=False)
    ws.add_data_validation(brand_dv)
    brand_dv.add('A2:A10000')
    
    # Category dropdown
    category_values = '"Jeans,T-Shirts,Shirts,Trousers,Shorts,Jackets,Sweaters,Hoodies,Shoes,Sandals,Sneakers,Formal Shoes,Accessories,Belts,Wallets,Caps"'
    category_dv = DataValidation(type="list", formula1=category_values, allow_blank=False)
    ws.add_data_validation(category_dv)
    category_dv.add('B2:B10000')
    
    # Size dropdown
    size_values = '"28,30,32,34,36,38,40,42,44,46,XS,S,M,L,XL,XXL,XXXL,6,7,8,9,10,11,12,Free Size"'
    size_dv = DataValidation(type="list", formula1=size_values, allow_blank=False)
    ws.add_data_validation(size_dv)
    size_dv.add('D2:D10000')
    
    # Color dropdown
    color_values = '"Blue,Black,White,Red,Green,Yellow,Grey,Brown,Pink,Purple,Orange,Navy,Maroon,Beige,Cream,Olive,Khaki,Indigo"'
    color_dv = DataValidation(type="list", formula1=color_values, allow_blank=False)
    ws.add_data_validation(color_dv)
    color_dv.add('E2:E10000')
    
    # Style dropdown
    style_values = '"Slim Fit,Regular Fit,Loose Fit,Relaxed Fit,Athletic Fit,Straight Fit,Bootcut,Skinny,Casual,Formal,Party,Sports"'
    style_dv = DataValidation(type="list", formula1=style_values, allow_blank=True)
    ws.add_data_validation(style_dv)
    style_dv.add('F2:F10000')
    
    # Group dropdown
    group_values = '"Men\'s Wear,Women\'s Wear,Boys Wear,Girls Wear,Kids Wear,Unisex,Accessories"'
    group_dv = DataValidation(type="list", formula1=group_values, allow_blank=False)
    ws.add_data_validation(group_dv)
    group_dv.add('J2:J10000')
    
    # Unit dropdown
    unit_values = '"pc,pair,set,kg,gm,ltr,ml,dozen"'
    unit_dv = DataValidation(type="list", formula1=unit_values, allow_blank=False)
    ws.add_data_validation(unit_dv)
    unit_dv.add('K2:K10000')
    
    # Tax Rate dropdown
    tax_values = '"0,5,12,18,28"'
    tax_dv = DataValidation(type="list", formula1=tax_values, allow_blank=True)
    ws.add_data_validation(tax_dv)
    tax_dv.add('R2:R10000')
    
    print("✅ Dropdowns added")
    
    # ========================================
    # ADD EXCEL FORMULAS
    # ========================================
    
    for row in range(2, 10001):  # Support up to 10,000 items
        # Column G: Item Name = Brand + Category + Product + Size + Color + Style
        ws[f'G{row}'].value = f'=TRIM(A{row}&" "&B{row}&" "&C{row}&" "&D{row}&" "&E{row}&IF(F{row}<>"", " "&F{row}, ""))'
        ws[f'G{row}'].font = Font(color="FF6B00", bold=True)  # Orange to indicate auto-generated
        
        # Column P: Discount % = ((MRP - Selling) / MRP) * 100
        ws[f'P{row}'].value = f'=IF(AND(O{row}>0,Q{row}>0,Q{row}<=O{row}),ROUND((O{row}-Q{row})/O{row}*100,2),0)'
        ws[f'P{row}'].number_format = '0.00'
    
    print("✅ Formulas added")
    
    # ========================================
    # ADD EXAMPLE DATA
    # ========================================
    
    examples = [
        # Brand, Category, Product, Size, Color, Style, SKU, Barcode, Group, Unit, Stock, Reorder, Cost, MRP, Selling, Tax, HSN, Desc
        ["Levi's", "Jeans", "501", "32", "Blue", "Slim Fit", "", "", "Men's Wear", "pc", 15, 5, 1200, 3999, 2499, 12, "62034090", "Classic blue jeans"],
        ["Levi's", "Jeans", "501", "34", "Blue", "Slim Fit", "", "", "Men's Wear", "pc", 12, 5, 1200, 3999, 2499, 12, "62034090", ""],
        ["Levi's", "Jeans", "501", "32", "Black", "Regular Fit", "", "", "Men's Wear", "pc", 8, 5, 1200, 3999, 2499, 12, "62034090", ""],
        ["Nike", "T-Shirts", "Dri-FIT", "M", "Red", "", "", "", "Unisex", "pc", 50, 10, 400, 1299, 799, 12, "61091000", ""],
        ["Nike", "T-Shirts", "Dri-FIT", "L", "Red", "", "", "", "Unisex", "pc", 30, 10, 400, 1299, 799, 12, "61091000", ""],
        ["Adidas", "Shoes", "Superstar", "10", "White", "Casual", "", "", "Men's Wear", "pair", 8, 3, 2500, 5999, 3999, 18, "64039900", ""],
    ]
    
    for row_idx, example in enumerate(examples, start=2):
        for col_idx, value in enumerate(example, start=1):
            if col_idx not in [7, 16]:  # Skip Item Name (col 7) and Discount % (col 16) - they're auto-generated
                cell = ws.cell(row=row_idx, column=col_idx if col_idx < 7 else col_idx + 1 if col_idx < 16 else col_idx + 2)
                if col_idx <= 6:  # Structured columns
                    cell.value = value
                elif col_idx == 8:  # SKU
                    cell.value = example[6]
                elif col_idx == 9:  # Barcode
                    cell.value = example[7]
                elif col_idx >= 10:  # Rest of columns
                    actual_col = col_idx - 6 + 7 + 2  # Adjust for Item Name and Discount %
                    ws.cell(row=row_idx, column=actual_col).value = value
        
        # Grey background for example rows
        for col in range(1, 21):
            ws.cell(row=row_idx, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    print("✅ Example data added")
    
    # ========================================
    # ADD INSTRUCTIONS SHEET
    # ========================================
    
    ws_inst = wb.create_sheet("📖 Instructions", 0)
    
    instructions = [
        ["🎯 CLOTHING RETAIL IMPORT TEMPLATE", ""],
        ["Bridge Solution for Easy Inventory Setup", ""],
        ["", ""],
        ["📋 WHAT IS THIS?", ""],
        ["", ""],
        ["This is a temporary Excel template designed specifically for clothing retail businesses.", ""],
        ["It helps you import inventory with structured naming that:", ""],
        ["  ✅ Works with the current system (no code changes needed!)", ""],
        ["  ✅ Makes data entry super fast (drag-fill feature)", ""],
        ["  ✅ Keeps data organized for future migration to variants system", ""],
        ["  ✅ Creates clear, searchable item names", ""],
        ["", ""],
        ["📝 HOW TO USE (Step by Step)", ""],
        ["", ""],
        ["STEP 1: Fill the BLUE columns (Required)", ""],
        ["  These are mandatory fields - you must fill them:", ""],
        ["", ""],
        ["  A. Brand* → Select or type brand name (Levi's, Nike, Adidas...)", ""],
        ["  B. Category* → Select product type (Jeans, T-Shirts, Shirts...)", ""],
        ["  C. Product Name* → Type model/style (501, Dri-FIT, Air Max...)", ""],
        ["  D. Size* → Select from dropdown (28, 30, 32, S, M, L...)", ""],
        ["  E. Color* → Select from dropdown (Blue, Black, Red...)", ""],
        ["  J. Group* → Select department (Men's Wear, Women's Wear...)", ""],
        ["  K. Unit* → Select unit (pc for pieces, pair for shoes)", ""],
        ["  L. Stock Quantity* → Enter current stock count", ""],
        ["  N. Cost Price* → Your purchase/cost price", ""],
        ["  Q. Selling Price* → Price you charge to customers", ""],
        ["", ""],
        ["STEP 2: Fill the GREEN columns (Optional)", ""],
        ["  These are optional - fill if you have the information:", ""],
        ["", ""],
        ["  F. Style/Fit → Slim Fit, Regular Fit, etc. (optional)", ""],
        ["  H. SKU → Leave blank to auto-generate", ""],
        ["  I. Barcode → Leave blank to auto-generate", ""],
        ["  M. Reorder Point → Low stock alert threshold", ""],
        ["  O. MRP → Maximum Retail Price (optional)", ""],
        ["  R. Tax Rate → GST percentage (5, 12, 18, 28)", ""],
        ["  S. HSN Code → For GST compliance", ""],
        ["  T. Description → Additional details", ""],
        ["", ""],
        ["STEP 3: Item Name Auto-Generates! (ORANGE column)", ""],
        ["  Column G creates the full item name automatically:", ""],
        ["", ""],
        ["  🔶 DO NOT EDIT this column - it's automatic!", ""],
        ["  🔶 Formula: Brand + Category + Product + Size + Color + Style", ""],
        ["", ""],
        ["  Examples:", ""],
        ["    Input: Levi's | Jeans | 501 | 32 | Blue | Slim Fit", ""],
        ["    Output: \"Levi's Jeans 501 32 Blue Slim Fit\"", ""],
        ["", ""],
        ["    Input: Nike | T-Shirts | Dri-FIT | M | Red | (blank)", ""],
        ["    Output: \"Nike T-Shirts Dri-FIT M Red\"", ""],
        ["", ""],
        ["⚡ DRAG-FILL TRICK (Super Fast Data Entry!)", ""],
        ["", ""],
        ["For 80 Levi's jeans items, do this:", ""],
        ["", ""],
        ["  1. Row 2: Type \"Levi's\" in Brand column (2 seconds)", ""],
        ["  2. Select cell A2, grab bottom-right corner", ""],
        ["  3. Drag down to row 81", ""],
        ["  4. Done! All 80 rows now have \"Levi's\" ✨", ""],
        ["", ""],
        ["  Repeat for:", ""],
        ["    - Category (B column): Drag \"Jeans\" across all rows", ""],
        ["    - Product (C column): Drag \"501\" across all rows", ""],
        ["    - Group (J column): Drag \"Men's Wear\" across all rows", ""],
        ["    - Unit (K column): Drag \"pc\" across all rows", ""],
        ["    - Cost Price (N column): Drag same price if applicable", ""],
        ["    - MRP (O column): Drag same MRP if applicable", ""],
        ["    - Selling Price (Q column): Drag same price if applicable", ""],
        ["", ""],
        ["  Only Size and Color vary per item - use dropdowns for these!", ""],
        ["", ""],
        ["  💡 Result: 80 items in 5-10 minutes instead of hours!", ""],
        ["", ""],
        ["💾 STEP 4: Save and Upload", ""],
        ["", ""],
        ["  1. Delete example rows 2-7 (grey rows)", ""],
        ["  2. Fill your actual data starting from row 2", ""],
        ["  3. Save the file as .xlsx format", ""],
        ["  4. Go to BizBooks → Admin → Items → Import", ""],
        ["  5. Upload this file", ""],
        ["  6. System processes and imports! ✅", ""],
        ["", ""],
        ["✅ WHY THIS IS BRILLIANT", ""],
        ["", ""],
        ["1. Works Today", ""],
        ["   No waiting for new features - use immediately!", ""],
        ["", ""],
        ["2. Clear Item Names", ""],
        ["   Invoices show: \"Levi's Jeans 501 32 Blue Slim Fit\"", ""],
        ["   Customers know exactly what they're buying!", ""],
        ["", ""],
        ["3. Searchable", ""],
        ["   Search \"jeans\" → Finds all jeans ✅", ""],
        ["   Search \"nike red\" → Finds Nike red items ✅", ""],
        ["", ""],
        ["4. Fast Data Entry", ""],
        ["   Drag-fill common values (Brand, Category, Prices)", ""],
        ["   10,000 items in hours instead of days!", ""],
        ["", ""],
        ["5. Future-Proof", ""],
        ["   When variants system is ready:", ""],
        ["   System will parse: \"Levi's Jeans 501 32 Blue Slim Fit\"", ""],
        ["   And extract: Brand=Levi's, Category=Jeans, Size=32, Color=Blue", ""],
        ["   Automatically convert to proper variants! ✨", ""],
        ["", ""],
        ["🔮 FUTURE MIGRATION (Automatic!)", ""],
        ["", ""],
        ["When variant system is implemented:", ""],
        ["", ""],
        ["  Current: 3 separate items", ""],
        ["    - Levi's Jeans 501 32 Blue Slim Fit", ""],
        ["    - Levi's Jeans 501 34 Blue Slim Fit", ""],
        ["    - Levi's Jeans 501 32 Black Regular Fit", ""],
        ["", ""],
        ["  Future: 1 parent product + 3 variants", ""],
        ["    Parent: Levi's Jeans 501", ""],
        ["    ├─ Variant 1: Size=32, Color=Blue, Style=Slim Fit", ""],
        ["    ├─ Variant 2: Size=34, Color=Blue, Style=Slim Fit", ""],
        ["    └─ Variant 3: Size=32, Color=Black, Style=Regular Fit", ""],
        ["", ""],
        ["  Migration is AUTOMATIC because naming is structured! 🎉", ""],
        ["", ""],
        ["📞 SUPPORT", ""],
        ["", ""],
        ["Questions? Need help?", ""],
        ["  📱 Call/WhatsApp: +91 8983121201", ""],
        ["  📧 Email: bizbooks.notifications@gmail.com", ""],
        ["", ""],
        ["💡 This bridge solution was designed by Rish - Smart thinking!", ""],
    ]
    
    for row_idx, (text, value) in enumerate(instructions, start=1):
        ws_inst[f'A{row_idx}'].value = text
        
        # Styling
        if any(emoji in text for emoji in ["🎯", "📋", "📝", "⚡", "💾", "✅", "🔮", "📞"]):
            ws_inst[f'A{row_idx}'].font = Font(bold=True, size=14, color="0000FF")
        elif "STEP" in text:
            ws_inst[f'A{row_idx}'].font = Font(bold=True, size=12, color="C00000")
        
        ws_inst[f'A{row_idx}'].alignment = Alignment(wrap_text=True, vertical='top')
    
    ws_inst.column_dimensions['A'].width = 90
    
    print("✅ Instructions sheet created")
    
    # ========================================
    # SAVE FILE
    # ========================================
    
    filename = "BizBooks_Clothing_Retail_Import_Template.xlsx"
    wb.save(filename)
    
    return filename


if __name__ == "__main__":
    print("=" * 70)
    print("🚀 CREATING CLOTHING RETAIL IMPORT TEMPLATE")
    print("=" * 70)
    print()
    
    try:
        output_file = create_clothing_retail_template()
        
        print()
        print("=" * 70)
        print("✅ SUCCESS! Template created!")
        print("=" * 70)
        print()
        print(f"📁 File: {output_file}")
        print()
        print("📋 Template Features:")
        print("  ✅ Structured columns: Brand, Category, Product, Size, Color, Style")
        print("  ✅ Auto-generates Item Name: 'Levi's Jeans 501 32 Blue Slim Fit'")
        print("  ✅ Dropdowns for easy selection (Brand, Category, Size, Color, Style)")
        print("  ✅ Excel formulas (Item Name and Discount % auto-calculate)")
        print("  ✅ All standard columns included (SKU, Barcode, Group, Unit, Stock, etc.)")
        print("  ✅ Example data (delete before using)")
        print("  ✅ Complete instructions sheet")
        print()
        print("🎯 Key Benefits:")
        print("  ✅ Works with current system (no code changes!)")
        print("  ✅ Super-fast data entry (drag-fill feature)")
        print("  ✅ Clear, searchable item names")
        print("  ✅ Future-proof (easy to parse into variants)")
        print()
        print("📖 Next Steps:")
        print("  1. Open the Excel file")
        print("  2. Read the Instructions sheet")
        print("  3. Delete example rows (grey rows 2-7)")
        print("  4. Fill your inventory data")
        print("  5. Use drag-fill for common values (Brand, Category, Prices)")
        print("  6. Save and upload to BizBooks!")
        print()
        print("💡 Pro Tip:")
        print("  For 10,000 items: ~2-3 hours with drag-fill vs 40+ hours manually!")
        print()
        print("=" * 70)
        
    except Exception as e:
        print()
        print("❌ ERROR creating template:")
        print(f"   {str(e)}")
        print()
        import traceback
        traceback.print_exc()

