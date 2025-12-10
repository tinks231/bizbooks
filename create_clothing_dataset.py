#!/usr/bin/env python3
"""
Generate Clothing Retail Dataset Excel File
Temporary script to create downloadable Excel with 89 products
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

def create_clothing_dataset_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Clothing Inventory"
    
    # Headers
    headers = ['Item Name*', 'SKU', 'Barcode', 'Category*', 'Group*', 'Unit*', 'Stock Quantity*', 
               'Cost Price*', 'Selling Price*', 'MRP', 'Tax Rate (%)', 'HSN Code', 'Description']
    
    # Style headers
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data - All 89 products
    products = [
        # Men's T-Shirts
        ["Men's Cotton T-Shirt - White - S", "TSH-MEN-WHT-S", "8901234567001", "Men's Wear", "T-Shirts", "Pcs", 30, 450, 599, 699, 12, "6109", "Premium cotton t-shirt, size S"],
        ["Men's Cotton T-Shirt - White - M", "TSH-MEN-WHT-M", "8901234567002", "Men's Wear", "T-Shirts", "Pcs", 50, 450, 599, 699, 12, "6109", "Premium cotton t-shirt, size M"],
        ["Men's Cotton T-Shirt - White - L", "TSH-MEN-WHT-L", "8901234567003", "Men's Wear", "T-Shirts", "Pcs", 45, 450, 599, 699, 12, "6109", "Premium cotton t-shirt, size L"],
        ["Men's Cotton T-Shirt - White - XL", "TSH-MEN-WHT-XL", "8901234567004", "Men's Wear", "T-Shirts", "Pcs", 40, 450, 599, 699, 12, "6109", "Premium cotton t-shirt, size XL"],
        ["Men's Cotton T-Shirt - Black - S", "TSH-MEN-BLK-S", "8901234567005", "Men's Wear", "T-Shirts", "Pcs", 28, 450, 599, 699, 12, "6109", "Premium cotton t-shirt, size S"],
        ["Men's Cotton T-Shirt - Black - M", "TSH-MEN-BLK-M", "8901234567006", "Men's Wear", "T-Shirts", "Pcs", 45, 450, 599, 699, 12, "6109", "Premium cotton t-shirt, size M"],
        ["Men's Cotton T-Shirt - Black - L", "TSH-MEN-BLK-L", "8901234567007", "Men's Wear", "T-Shirts", "Pcs", 42, 450, 599, 699, 12, "6109", "Premium cotton t-shirt, size L"],
        ["Men's Cotton T-Shirt - Black - XL", "TSH-MEN-BLK-XL", "8901234567008", "Men's Wear", "T-Shirts", "Pcs", 38, 450, 599, 699, 12, "6109", "Premium cotton t-shirt, size XL"],
        ["Men's Cotton T-Shirt - Navy - M", "TSH-MEN-NVY-M", "8901234567009", "Men's Wear", "T-Shirts", "Pcs", 40, 450, 599, 699, 12, "6109", "Premium cotton t-shirt, size M"],
        ["Men's Cotton T-Shirt - Navy - L", "TSH-MEN-NVY-L", "8901234567010", "Men's Wear", "T-Shirts", "Pcs", 35, 450, 599, 699, 12, "6109", "Premium cotton t-shirt, size L"],
        
        # Men's Formal Shirts
        ["Men's Formal Shirt - Blue - 38", "SH-MEN-BLU-38", "8901234567011", "Men's Wear", "Shirts", "Pcs", 20, 950, 1299, 1499, 12, "6109", "Formal shirt, wrinkle-free, size 38"],
        ["Men's Formal Shirt - Blue - 40", "SH-MEN-BLU-40", "8901234567012", "Men's Wear", "Shirts", "Pcs", 30, 950, 1299, 1499, 12, "6109", "Formal shirt, wrinkle-free, size 40"],
        ["Men's Formal Shirt - Blue - 42", "SH-MEN-BLU-42", "8901234567013", "Men's Wear", "Shirts", "Pcs", 25, 950, 1299, 1499, 12, "6109", "Formal shirt, wrinkle-free, size 42"],
        ["Men's Formal Shirt - White - 38", "SH-MEN-WHT-38", "8901234567014", "Men's Wear", "Shirts", "Pcs", 25, 950, 1299, 1499, 12, "6109", "Formal shirt, wrinkle-free, size 38"],
        ["Men's Formal Shirt - White - 40", "SH-MEN-WHT-40", "8901234567015", "Men's Wear", "Shirts", "Pcs", 35, 950, 1299, 1499, 12, "6109", "Formal shirt, wrinkle-free, size 40"],
        ["Men's Formal Shirt - White - 42", "SH-MEN-WHT-42", "8901234567016", "Men's Wear", "Shirts", "Pcs", 28, 950, 1299, 1499, 12, "6109", "Formal shirt, wrinkle-free, size 42"],
        
        # Men's Jeans
        ["Men's Jeans - Blue - 30", "JNS-MEN-BLU-30", "8901234567017", "Men's Wear", "Jeans", "Pcs", 20, 1400, 1899, 2199, 12, "6109", "Slim fit denim jeans, waist 30"],
        ["Men's Jeans - Blue - 32", "JNS-MEN-BLU-32", "8901234567018", "Men's Wear", "Jeans", "Pcs", 40, 1400, 1899, 2199, 12, "6109", "Slim fit denim jeans, waist 32"],
        ["Men's Jeans - Blue - 34", "JNS-MEN-BLU-34", "8901234567019", "Men's Wear", "Jeans", "Pcs", 35, 1400, 1899, 2199, 12, "6109", "Slim fit denim jeans, waist 34"],
        ["Men's Jeans - Blue - 36", "JNS-MEN-BLU-36", "8901234567020", "Men's Wear", "Jeans", "Pcs", 30, 1400, 1899, 2199, 12, "6109", "Slim fit denim jeans, waist 36"],
        ["Men's Jeans - Black - 30", "JNS-MEN-BLK-30", "8901234567021", "Men's Wear", "Jeans", "Pcs", 15, 1400, 1899, 2199, 12, "6109", "Slim fit denim jeans, waist 30"],
        ["Men's Jeans - Black - 32", "JNS-MEN-BLK-32", "8901234567022", "Men's Wear", "Jeans", "Pcs", 25, 1400, 1899, 2199, 12, "6109", "Slim fit denim jeans, waist 32"],
        ["Men's Jeans - Black - 34", "JNS-MEN-BLK-34", "8901234567023", "Men's Wear", "Jeans", "Pcs", 22, 1400, 1899, 2199, 12, "6109", "Slim fit denim jeans, waist 34"],
        ["Men's Jeans - Black - 36", "JNS-MEN-BLK-36", "8901234567024", "Men's Wear", "Jeans", "Pcs", 18, 1400, 1899, 2199, 12, "6109", "Slim fit denim jeans, waist 36"],
        
        # Men's Trousers
        ["Men's Chinos - Beige - 30", "CHN-MEN-BGE-30", "8901234567025", "Men's Wear", "Trousers", "Pcs", 15, 1100, 1499, 1799, 12, "6109", "Casual chinos, perfect fit, waist 30"],
        ["Men's Chinos - Beige - 32", "CHN-MEN-BGE-32", "8901234567026", "Men's Wear", "Trousers", "Pcs", 20, 1100, 1499, 1799, 12, "6109", "Casual chinos, perfect fit, waist 32"],
        ["Men's Chinos - Beige - 34", "CHN-MEN-BGE-34", "8901234567027", "Men's Wear", "Trousers", "Pcs", 18, 1100, 1499, 1799, 12, "6109", "Casual chinos, perfect fit, waist 34"],
        ["Men's Formal Trousers - Black - 32", "TRS-MEN-BLK-32", "8901234567028", "Men's Wear", "Trousers", "Pcs", 20, 1350, 1799, 1999, 12, "6109", "Formal trousers, office wear, waist 32"],
        ["Men's Formal Trousers - Black - 34", "TRS-MEN-BLK-34", "8901234567029", "Men's Wear", "Trousers", "Pcs", 28, 1350, 1799, 1999, 12, "6109", "Formal trousers, office wear, waist 34"],
        ["Men's Formal Trousers - Black - 36", "TRS-MEN-BLK-36", "8901234567030", "Men's Wear", "Trousers", "Pcs", 25, 1350, 1799, 1999, 12, "6109", "Formal trousers, office wear, waist 36"],
        
        # Women's Kurtas
        ["Women's Cotton Kurta - Blue - S", "KRT-WOM-BLU-S", "8901234567031", "Women's Wear", "Kurtas", "Pcs", 35, 650, 899, 1099, 12, "6109", "Comfortable cotton kurta, size S"],
        ["Women's Cotton Kurta - Blue - M", "KRT-WOM-BLU-M", "8901234567032", "Women's Wear", "Kurtas", "Pcs", 60, 650, 899, 1099, 12, "6109", "Comfortable cotton kurta, size M"],
        ["Women's Cotton Kurta - Blue - L", "KRT-WOM-BLU-L", "8901234567033", "Women's Wear", "Kurtas", "Pcs", 45, 650, 899, 1099, 12, "6109", "Comfortable cotton kurta, size L"],
        ["Women's Cotton Kurta - Blue - XL", "KRT-WOM-BLU-XL", "8901234567034", "Women's Wear", "Kurtas", "Pcs", 30, 650, 899, 1099, 12, "6109", "Comfortable cotton kurta, size XL"],
        ["Women's Cotton Kurta - Pink - S", "KRT-WOM-PNK-S", "8901234567035", "Women's Wear", "Kurtas", "Pcs", 32, 650, 899, 1099, 12, "6109", "Comfortable cotton kurta, size S"],
        ["Women's Cotton Kurta - Pink - M", "KRT-WOM-PNK-M", "8901234567036", "Women's Wear", "Kurtas", "Pcs", 55, 650, 899, 1099, 12, "6109", "Comfortable cotton kurta, size M"],
        ["Women's Cotton Kurta - Pink - L", "KRT-WOM-PNK-L", "8901234567037", "Women's Wear", "Kurtas", "Pcs", 42, 650, 899, 1099, 12, "6109", "Comfortable cotton kurta, size L"],
        ["Women's Cotton Kurta - Pink - XL", "KRT-WOM-PNK-XL", "8901234567038", "Women's Wear", "Kurtas", "Pcs", 28, 650, 899, 1099, 12, "6109", "Comfortable cotton kurta, size XL"],
        
        # Women's Leggings
        ["Women's Leggings - Black - S", "LGG-WOM-BLK-S", "8901234567039", "Women's Wear", "Leggings", "Pcs", 40, 280, 399, 499, 12, "6109", "Stretchable leggings, size S"],
        ["Women's Leggings - Black - M", "LGG-WOM-BLK-M", "8901234567040", "Women's Wear", "Leggings", "Pcs", 80, 280, 399, 499, 12, "6109", "Stretchable leggings, size M"],
        ["Women's Leggings - Black - L", "LGG-WOM-BLK-L", "8901234567041", "Women's Wear", "Leggings", "Pcs", 60, 280, 399, 499, 12, "6109", "Stretchable leggings, size L"],
        ["Women's Leggings - Black - XL", "LGG-WOM-BLK-XL", "8901234567042", "Women's Wear", "Leggings", "Pcs", 40, 280, 399, 499, 12, "6109", "Stretchable leggings, size XL"],
        ["Women's Leggings - Navy - M", "LGG-WOM-NVY-M", "8901234567043", "Women's Wear", "Leggings", "Pcs", 75, 280, 399, 499, 12, "6109", "Stretchable leggings, size M"],
        ["Women's Leggings - Navy - L", "LGG-WOM-NVY-L", "8901234567044", "Women's Wear", "Leggings", "Pcs", 55, 280, 399, 499, 12, "6109", "Stretchable leggings, size L"],
        
        # Women's Dresses
        ["Women's Dress - Floral - M", "DRS-WOM-FLR-M", "8901234567045", "Women's Wear", "Dresses", "Pcs", 30, 1100, 1499, 1799, 12, "6109", "Floral print summer dress, size M"],
        ["Women's Dress - Floral - L", "DRS-WOM-FLR-L", "8901234567046", "Women's Wear", "Dresses", "Pcs", 25, 1100, 1499, 1799, 12, "6109", "Floral print summer dress, size L"],
        ["Women's Dress - Solid Blue - S", "DRS-WOM-BLU-S", "8901234567047", "Women's Wear", "Dresses", "Pcs", 15, 1100, 1499, 1799, 12, "6109", "Solid color summer dress, size S"],
        ["Women's Dress - Solid Blue - M", "DRS-WOM-BLU-M", "8901234567048", "Women's Wear", "Dresses", "Pcs", 25, 1100, 1499, 1799, 12, "6109", "Solid color summer dress, size M"],
        ["Women's Dress - Solid Blue - L", "DRS-WOM-BLU-L", "8901234567049", "Women's Wear", "Dresses", "Pcs", 20, 1100, 1499, 1799, 12, "6109", "Solid color summer dress, size L"],
        
        # Women's Jeans
        ["Women's Jeans - Blue - 28", "JNS-WOM-BLU-28", "8901234567050", "Women's Wear", "Jeans", "Pcs", 25, 1250, 1699, 1999, 12, "6109", "Skinny fit jeans, waist 28"],
        ["Women's Jeans - Blue - 30", "JNS-WOM-BLU-30", "8901234567051", "Women's Wear", "Jeans", "Pcs", 35, 1250, 1699, 1999, 12, "6109", "Skinny fit jeans, waist 30"],
        ["Women's Jeans - Blue - 32", "JNS-WOM-BLU-32", "8901234567052", "Women's Wear", "Jeans", "Pcs", 30, 1250, 1699, 1999, 12, "6109", "Skinny fit jeans, waist 32"],
        ["Women's Jeans - Blue - 34", "JNS-WOM-BLU-34", "8901234567053", "Women's Wear", "Jeans", "Pcs", 20, 1250, 1699, 1999, 12, "6109", "Skinny fit jeans, waist 34"],
        
        # Women's Tops
        ["Women's Top - White - S", "TOP-WOM-WHT-S", "8901234567054", "Women's Wear", "Tops", "Pcs", 30, 420, 599, 699, 12, "6109", "Casual top, comfortable, size S"],
        ["Women's Top - White - M", "TOP-WOM-WHT-M", "8901234567055", "Women's Wear", "Tops", "Pcs", 50, 420, 599, 699, 12, "6109", "Casual top, comfortable, size M"],
        ["Women's Top - White - L", "TOP-WOM-WHT-L", "8901234567056", "Women's Wear", "Tops", "Pcs", 40, 420, 599, 699, 12, "6109", "Casual top, comfortable, size L"],
        
        # Kids T-Shirts
        ["Kids T-Shirt - Blue - 6-7Y", "TSH-KID-BLU-6-7Y", "8901234567057", "Kids Wear", "T-Shirts", "Pcs", 25, 210, 299, 349, 5, "6109", "Kids cotton t-shirt, age 6-7 years"],
        ["Kids T-Shirt - Blue - 8-9Y", "TSH-KID-BLU-8-9Y", "8901234567058", "Kids Wear", "T-Shirts", "Pcs", 40, 210, 299, 349, 5, "6109", "Kids cotton t-shirt, age 8-9 years"],
        ["Kids T-Shirt - Blue - 10-11Y", "TSH-KID-BLU-10-11Y", "8901234567059", "Kids Wear", "T-Shirts", "Pcs", 30, 210, 299, 349, 5, "6109", "Kids cotton t-shirt, age 10-11 years"],
        ["Kids T-Shirt - Red - 6-7Y", "TSH-KID-RED-6-7Y", "8901234567060", "Kids Wear", "T-Shirts", "Pcs", 22, 210, 299, 349, 5, "6109", "Kids cotton t-shirt, age 6-7 years"],
        ["Kids T-Shirt - Red - 8-9Y", "TSH-KID-RED-8-9Y", "8901234567061", "Kids Wear", "T-Shirts", "Pcs", 35, 210, 299, 349, 5, "6109", "Kids cotton t-shirt, age 8-9 years"],
        ["Kids T-Shirt - Red - 10-11Y", "TSH-KID-RED-10-11Y", "8901234567062", "Kids Wear", "T-Shirts", "Pcs", 28, 210, 299, 349, 5, "6109", "Kids cotton t-shirt, age 10-11 years"],
        
        # Kids Shorts & Jeans
        ["Kids Shorts - Blue - 8-9Y", "SHT-KID-BLU-8-9Y", "8901234567063", "Kids Wear", "Shorts", "Pcs", 30, 280, 399, 449, 5, "6109", "Comfortable kids shorts, age 8-9 years"],
        ["Kids Shorts - Blue - 10-11Y", "SHT-KID-BLU-10-11Y", "8901234567064", "Kids Wear", "Shorts", "Pcs", 25, 280, 399, 449, 5, "6109", "Comfortable kids shorts, age 10-11 years"],
        ["Kids Jeans - Blue - 8-9Y", "JNS-KID-BLU-8-9Y", "8901234567065", "Kids Wear", "Jeans", "Pcs", 25, 560, 799, 899, 5, "6109", "Durable kids jeans, age 8-9 years"],
        ["Kids Jeans - Blue - 10-11Y", "JNS-KID-BLU-10-11Y", "8901234567066", "Kids Wear", "Jeans", "Pcs", 22, 560, 799, 899, 5, "6109", "Durable kids jeans, age 10-11 years"],
        ["Kids Jeans - Blue - 12-13Y", "JNS-KID-BLU-12-13Y", "8901234567067", "Kids Wear", "Jeans", "Pcs", 18, 560, 799, 899, 5, "6109", "Durable kids jeans, age 12-13 years"],
        
        # Accessories - Belts
        ["Men's Belt - Black - M", "BLT-MEN-BLK-M", "8901234567068", "Accessories", "Belts", "Pcs", 20, 350, 499, 599, 12, "4203", "Genuine leather belt, size M (32-34)"],
        ["Men's Belt - Black - L", "BLT-MEN-BLK-L", "8901234567069", "Accessories", "Belts", "Pcs", 25, 350, 499, 599, 12, "4203", "Genuine leather belt, size L (36-38)"],
        ["Men's Belt - Brown - M", "BLT-MEN-BRN-M", "8901234567070", "Accessories", "Belts", "Pcs", 18, 350, 499, 599, 12, "4203", "Genuine leather belt, size M (32-34)"],
        ["Men's Belt - Brown - L", "BLT-MEN-BRN-L", "8901234567071", "Accessories", "Belts", "Pcs", 22, 350, 499, 599, 12, "4203", "Genuine leather belt, size L (36-38)"],
        
        # Accessories - Bags
        ["Women's Handbag - Brown", "BAG-WOM-BRN-001", "8901234567072", "Accessories", "Bags", "Pcs", 15, 900, 1299, 1599, 12, "4203", "Stylish handbag, faux leather"],
        ["Women's Handbag - Black", "BAG-WOM-BLK-001", "8901234567073", "Accessories", "Bags", "Pcs", 18, 900, 1299, 1599, 12, "4203", "Stylish handbag, faux leather"],
        ["Women's Purse - Red", "PURSE-WOM-RED-001", "8901234567074", "Accessories", "Bags", "Pcs", 12, 450, 699, 799, 12, "4203", "Compact purse, multiple compartments"],
        ["Women's Purse - Black", "PURSE-WOM-BLK-001", "8901234567075", "Accessories", "Bags", "Pcs", 15, 450, 699, 799, 12, "4203", "Compact purse, multiple compartments"],
        
        # Accessories - Caps
        ["Unisex Cap - Black", "CAP-UNI-BLK-001", "8901234567076", "Accessories", "Caps", "Pcs", 30, 210, 299, 349, 12, "4203", "Adjustable cap, one size fits all"],
        ["Unisex Cap - Navy", "CAP-UNI-NVY-001", "8901234567077", "Accessories", "Caps", "Pcs", 25, 210, 299, 349, 12, "4203", "Adjustable cap, one size fits all"],
        ["Unisex Cap - White", "CAP-UNI-WHT-001", "8901234567078", "Accessories", "Caps", "Pcs", 20, 210, 299, 349, 12, "4203", "Adjustable cap, one size fits all"],
        
        # Accessories - Socks & Scarves
        ["Men's Socks - Black - Pack of 3", "SOCK-MEN-BLK-3PK", "8901234567079", "Accessories", "Socks", "Pack", 50, 150, 249, 299, 5, "6115", "Cotton blend socks, 3 pairs pack"],
        ["Men's Socks - White - Pack of 3", "SOCK-MEN-WHT-3PK", "8901234567080", "Accessories", "Socks", "Pack", 45, 150, 249, 299, 5, "6115", "Cotton blend socks, 3 pairs pack"],
        ["Women's Socks - Multicolor - Pack of 5", "SOCK-WOM-MUL-5PK", "8901234567081", "Accessories", "Socks", "Pack", 35, 200, 349, 399, 5, "6115", "Ankle socks, 5 pairs pack"],
        ["Women's Scarf - Floral", "SCARF-WOM-FLR-001", "8901234567082", "Accessories", "Scarves", "Pcs", 22, 280, 399, 499, 5, "6214", "Lightweight scarf, floral print"],
        ["Women's Scarf - Solid Black", "SCARF-WOM-BLK-001", "8901234567083", "Accessories", "Scarves", "Pcs", 18, 280, 399, 499, 5, "6214", "Lightweight scarf, solid color"],
    ]
    
    # Add data rows
    for row_num, product in enumerate(products, start=2):
        for col_num, value in enumerate(product, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            # Format number columns
            if col_num in [7, 8, 9, 10, 11]:  # Stock, Cost, Selling, MRP, Tax
                cell.number_format = '0.00' if col_num in [8, 9, 10] else '0'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 35  # Item Name
    ws.column_dimensions['B'].width = 18  # SKU
    ws.column_dimensions['C'].width = 16  # Barcode
    ws.column_dimensions['D'].width = 18  # Category
    ws.column_dimensions['E'].width = 15  # Group
    ws.column_dimensions['F'].width = 8   # Unit
    ws.column_dimensions['G'].width = 14  # Stock
    ws.column_dimensions['H'].width = 12  # Cost
    ws.column_dimensions['I'].width = 14  # Selling
    ws.column_dimensions['J'].width = 10  # MRP
    ws.column_dimensions['K'].width = 12  # Tax
    ws.column_dimensions['L'].width = 12  # HSN
    ws.column_dimensions['M'].width = 40  # Description
    
    # Save file
    output_path = "BizBooks_Clothing_Retail_89_Products.xlsx"
    wb.save(output_path)
    print(f"✅ Excel file created: {output_path}")
    print(f"📊 Total products: {len(products)}")
    print(f"📁 Location: {output_path}")
    print("\n🚀 Ready to import into BizBooks!")
    
    return output_path

if __name__ == "__main__":
    create_clothing_dataset_excel()

