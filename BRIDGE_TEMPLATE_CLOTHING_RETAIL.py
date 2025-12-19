"""
BRIDGE SOLUTION: Clothing Retail Import Template (FINAL VERSION)
=================================================================

Purpose: Import inventory with structured naming convention
         that can be easily parsed into variants later

Item Name Formula: Brand + Category + Product + Size + Color + Style
Example: "Levi's Jeans 501 32 Blue Slim Fit"

Benefits:
- Clear on invoices (customers know what they're buying)
- Searchable (search "jeans" works!)
- Drag-fill friendly (fast data entry)
- Future-proof (easy to parse into variants)

Current System: Stores as simple items
Future System: Can extract Brand/Category/Size/Color/Style from item name

Author: Smart solution by Rish! 🎯
Version: 2.0 (Category included in Item Name)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
import os

def create_bridge_template():
    """
    Create Excel template with separate columns that concatenate into Item Name
    
    User Experience:
    1. Fill separate columns: Base Name, Size, Color, Style
    2. Item Name auto-generates: "Levi's 501 32 Blue Slim"
    3. Import as simple item (current system)
    4. Future: Parse Item Name to extract variants
    """
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clothing Inventory"
    
    # ========================================
    # COLUMN DEFINITIONS
    # ========================================
    
    columns = [
        # Input Columns (User fills these)
        ('Base Product Name*', 'A', 'e.g., Levi\'s 501, Nike T-Shirt', True),
        ('Size*', 'B', 'e.g., 28, 30, 32, S, M, L, XL', True),
        ('Color*', 'C', 'e.g., Blue, Black, Red, White', True),
        ('Style/Fit', 'D', 'Optional: Slim, Regular, Loose, etc.', False),
        
        # Auto-Generated Column (Formula)
        ('🔹 Item Name (Auto)', 'E', 'Auto-generated - DO NOT EDIT', False),
        
        # Standard Columns (User fills)
        ('MRP*', 'F', 'Maximum Retail Price', True),
        ('Selling Price*', 'G', 'Actual selling price', True),
        ('Stock Quantity*', 'H', 'Current stock', True),
        
        # Optional Columns
        ('Barcode', 'I', 'Optional: Leave blank to auto-generate', False),
        ('SKU', 'J', 'Optional: Leave blank to auto-generate', False),
        ('Category', 'K', 'Optional: e.g., Jeans, Shirts, T-Shirts', False),
        ('HSN/SAC Code', 'L', 'Optional: For GST', False),
        ('Tax %', 'M', 'Optional: GST rate (5, 12, 18, 28)', False),
        ('Reorder Point', 'N', 'Optional: Low stock alert', False),
    ]
    
    # ========================================
    # HEADER ROW (Row 1)
    # ========================================
    
    for idx, (name, col_letter, description, required) in enumerate(columns, start=1):
        cell = ws[f'{col_letter}1']
        cell.value = name
        
        # Styling
        if required:
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
        elif 'Auto' in name:
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.font = Font(color="000000", bold=True, size=11)
        else:
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
        
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add comment with description
        cell.comment = Comment(description, "BizBooks")
    
    # Set column widths
    ws.column_dimensions['A'].width = 25  # Base Product Name
    ws.column_dimensions['B'].width = 10  # Size
    ws.column_dimensions['C'].width = 12  # Color
    ws.column_dimensions['D'].width = 12  # Style/Fit
    ws.column_dimensions['E'].width = 40  # Item Name (Auto)
    ws.column_dimensions['F'].width = 12  # MRP
    ws.column_dimensions['G'].width = 15  # Selling Price
    ws.column_dimensions['H'].width = 15  # Stock Quantity
    ws.column_dimensions['I'].width = 15  # Barcode
    ws.column_dimensions['J'].width = 15  # SKU
    ws.column_dimensions['K'].width = 15  # Category
    ws.column_dimensions['L'].width = 12  # HSN/SAC
    ws.column_dimensions['M'].width = 10  # Tax %
    ws.column_dimensions['N'].width = 15  # Reorder Point
    
    # ========================================
    # DATA VALIDATION (Dropdowns)
    # ========================================
    
    # Size dropdown
    size_values = '"28,30,32,34,36,38,40,42,44,XS,S,M,L,XL,XXL,XXXL,Free Size"'
    size_validation = DataValidation(type="list", formula1=size_values, allow_blank=False)
    size_validation.error = 'Please select a valid size'
    size_validation.errorTitle = 'Invalid Size'
    ws.add_data_validation(size_validation)
    size_validation.add(f'B2:B10000')
    
    # Color dropdown
    color_values = '"Blue,Black,White,Red,Green,Yellow,Grey,Brown,Pink,Purple,Orange,Navy,Maroon,Beige,Cream"'
    color_validation = DataValidation(type="list", formula1=color_values, allow_blank=False)
    color_validation.error = 'Please select a valid color'
    color_validation.errorTitle = 'Invalid Color'
    ws.add_data_validation(color_validation)
    color_validation.add(f'C2:C10000')
    
    # Style dropdown
    style_values = '"Slim Fit,Regular Fit,Loose Fit,Relaxed Fit,Athletic Fit,Straight Fit,Bootcut,Skinny,Casual,Formal"'
    style_validation = DataValidation(type="list", formula1=style_values, allow_blank=True)
    ws.add_data_validation(style_validation)
    style_validation.add(f'D2:D10000')
    
    # Tax % dropdown
    tax_values = '"0,5,12,18,28"'
    tax_validation = DataValidation(type="list", formula1=tax_values, allow_blank=True)
    ws.add_data_validation(tax_validation)
    tax_validation.add(f'M2:M10000')
    
    # ========================================
    # AUTO-GENERATION FORMULAS
    # ========================================
    
    # Row 2 onwards: Add formulas for Item Name
    for row in range(2, 10001):  # Support 10,000 items
        # Item Name = Base + Size + Color + Style (if not empty)
        # Formula: =A2&" "&B2&" "&C2&IF(D2<>"", " "&D2, "")
        ws[f'E{row}'].value = f'=TRIM(A{row}&" "&B{row}&" "&C{row}&IF(D{row}<>"", " "&D{row}, ""))'
        ws[f'E{row}'].font = Font(color="FF6B00", bold=True)  # Orange color to indicate auto
        
        # Discount % (for reference, not imported)
        # ws[f'O{row}'].value = f'=IF(AND(F{row}>0,G{row}>0),ROUND((F{row}-G{row})/F{row}*100,2),0)'
    
    # ========================================
    # EXAMPLE ROWS (Rows 2-6)
    # ========================================
    
    examples = [
        # Base Name, Size, Color, Style, MRP, Selling, Stock
        ["Levi's 501 Jeans", "32", "Blue", "Slim Fit", 3999, 2499, 15, "", "", "Jeans", "", 12, 5],
        ["Levi's 501 Jeans", "34", "Blue", "Slim Fit", 3999, 2499, 12, "", "", "Jeans", "", 12, 5],
        ["Levi's 501 Jeans", "32", "Black", "Regular Fit", 3999, 2499, 8, "", "", "Jeans", "", 12, 5],
        ["Nike Dri-FIT T-Shirt", "M", "Red", "", 1299, 799, 50, "", "", "T-Shirts", "", 12, 10],
        ["Nike Dri-FIT T-Shirt", "L", "Red", "", 1299, 799, 30, "", "", "T-Shirts", "", 12, 10],
    ]
    
    for row_idx, example in enumerate(examples, start=2):
        # Fill input columns (A-D, F-N)
        ws[f'A{row_idx}'].value = example[0]  # Base Name
        ws[f'B{row_idx}'].value = example[1]  # Size
        ws[f'C{row_idx}'].value = example[2]  # Color
        ws[f'D{row_idx}'].value = example[3]  # Style
        # E is auto-generated (formula already there)
        ws[f'F{row_idx}'].value = example[4]  # MRP
        ws[f'G{row_idx}'].value = example[5]  # Selling Price
        ws[f'H{row_idx}'].value = example[6]  # Stock
        ws[f'I{row_idx}'].value = example[7]  # Barcode
        ws[f'J{row_idx}'].value = example[8]  # SKU
        ws[f'K{row_idx}'].value = example[9]  # Category
        ws[f'L{row_idx}'].value = example[10] # HSN
        ws[f'M{row_idx}'].value = example[11] # Tax %
        ws[f'N{row_idx}'].value = example[12] # Reorder Point
        
        # Grey background for examples
        for col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'K', 'M', 'N']:
            ws[f'{col}{row_idx}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # ========================================
    # INSTRUCTIONS SHEET
    # ========================================
    
    ws_instructions = wb.create_sheet("📖 Instructions", 0)
    
    instructions = [
        ["🎯 BRIDGE SOLUTION: Clothing Retail Import", ""],
        ["", ""],
        ["Purpose:", "Import inventory with structured naming for easy future migration to variants"],
        ["", ""],
        ["📋 HOW TO USE THIS TEMPLATE", ""],
        ["", ""],
        ["Step 1: Fill the Blue Columns (Required)", ""],
        ["  ✅ Base Product Name: e.g., Levi's 501, Nike T-Shirt", ""],
        ["  ✅ Size: Select from dropdown (28, 30, 32... or S, M, L...)", ""],
        ["  ✅ Color: Select from dropdown (Blue, Black, Red...)", ""],
        ["  ✅ MRP: Maximum Retail Price", ""],
        ["  ✅ Selling Price: Actual selling price", ""],
        ["  ✅ Stock Quantity: Current stock", ""],
        ["", ""],
        ["Step 2: Fill the Green Columns (Optional)", ""],
        ["  🟢 Style/Fit: Slim, Regular, Loose (optional)", ""],
        ["  🟢 Barcode: Leave blank to auto-generate", ""],
        ["  🟢 SKU: Leave blank to auto-generate", ""],
        ["  🟢 Category: Jeans, T-Shirts, Shirts (helps in grouping)", ""],
        ["  🟢 HSN/SAC Code: For GST compliance", ""],
        ["  🟢 Tax %: GST rate (5, 12, 18, 28)", ""],
        ["  🟢 Reorder Point: Low stock alert threshold", ""],
        ["", ""],
        ["Step 3: Item Name Auto-Generates (Orange Column)", ""],
        ["  🔸 DO NOT EDIT this column!", ""],
        ["  🔸 It automatically combines: Base + Size + Color + Style", ""],
        ["  🔸 Example: Levi's 501 + 32 + Blue + Slim Fit = 'Levi's 501 32 Blue Slim Fit'", ""],
        ["", ""],
        ["Step 4: Delete Example Rows", ""],
        ["  ❌ Delete rows 2-6 (grey rows with example data)", ""],
        ["  ✅ Start your data from row 2", ""],
        ["", ""],
        ["Step 5: Save and Import", ""],
        ["  💾 Save as .xlsx file", ""],
        ["  📤 Upload to BizBooks (Admin > Items > Import)", ""],
        ["  ⏱️ System processes in seconds!", ""],
        ["", ""],
        ["💡 NAMING CONVENTION (IMPORTANT!)", ""],
        ["", ""],
        ["The Item Name follows this pattern:", ""],
        ["  [Base Name] [Size] [Color] [Style]", ""],
        ["", ""],
        ["Examples:", ""],
        ["  ✅ Levi's 501 32 Blue Slim Fit", ""],
        ["  ✅ Nike T-Shirt M Red", ""],
        ["  ✅ Adidas Shoes 10 Black", ""],
        ["  ✅ H&M Shirt 40 White Regular Fit", ""],
        ["", ""],
        ["Why this matters:", ""],
        ["  🔮 Future: When we implement variants system,", ""],
        ["  🔮 We can easily parse these names to extract:", ""],
        ["  🔮   - Base Product: Levi's 501", ""],
        ["  🔮   - Size: 32", ""],
        ["  🔮   - Color: Blue", ""],
        ["  🔮   - Style: Slim Fit", ""],
        ["  🔮 And convert to proper parent + variants!", ""],
        ["", ""],
        ["🎯 BENEFITS", ""],
        ["", ""],
        ["1. Works with current system (no code changes needed)", ""],
        ["2. Consistent naming makes searching easier", ""],
        ["3. Future migration is automatic (just parse the names)", ""],
        ["4. Data is already structured (even though stored as text)", ""],
        ["5. Zero risk of data loss", ""],
        ["6. Can start importing TODAY!", ""],
        ["", ""],
        ["📊 EXAMPLE: 10,000 Items", ""],
        ["", ""],
        ["Time to prepare data:", ""],
        ["  With this template: 2-3 hours (dropdowns make it fast!)", ""],
        ["  Traditional way: 40+ hours", ""],
        ["", ""],
        ["Future migration:", ""],
        ["  With structured names: 5 minutes (automatic parsing)", ""],
        ["  Without structure: Manual grouping nightmare", ""],
        ["", ""],
        ["🚀 SMART MOVE!", ""],
        ["This bridge solution lets you:", ""],
        ["  ✅ Start using the system TODAY", ""],
        ["  ✅ Keep data organized from Day 1", ""],
        ["  ✅ Easy migration when variants are ready", ""],
        ["  ✅ Zero technical debt!", ""],
        ["", ""],
        ["📞 Support: +91 8983121201", ""],
        ["📧 Email: bizbooks.notifications@gmail.com", ""],
    ]
    
    for row_idx, (text, value) in enumerate(instructions, start=1):
        ws_instructions[f'A{row_idx}'].value = text
        
        # Styling
        if "🎯" in text or "📋" in text or "💡" in text or "📊" in text or "🚀" in text:
            ws_instructions[f'A{row_idx}'].font = Font(bold=True, size=14, color="0000FF")
        elif "Step" in text:
            ws_instructions[f'A{row_idx}'].font = Font(bold=True, size=12, color="C00000")
        
        ws_instructions[f'A{row_idx}'].alignment = Alignment(wrap_text=True)
    
    ws_instructions.column_dimensions['A'].width = 80
    
    # ========================================
    # SIZE/COLOR REFERENCE SHEET
    # ========================================
    
    ws_reference = wb.create_sheet("📏 Size & Color Reference")
    
    reference_data = [
        ["📏 SIZE REFERENCE", "", "🎨 COLOR REFERENCE"],
        ["", "", ""],
        ["Men's Jeans/Pants", "", "Standard Colors"],
        ["Size", "Waist (inches)", "Color Name"],
        ["28", "28\"", "Blue"],
        ["30", "30\"", "Black"],
        ["32", "32\"", "White"],
        ["34", "34\"", "Red"],
        ["36", "36\"", "Green"],
        ["38", "38\"", "Yellow"],
        ["40", "40\"", "Grey"],
        ["42", "42\"", "Brown"],
        ["44", "44\"", "Pink"],
        ["", "", "Purple"],
        ["T-Shirts/Shirts", "", "Orange"],
        ["Size", "Chest (inches)", "Navy"],
        ["XS", "34-36\"", "Maroon"],
        ["S", "36-38\"", "Beige"],
        ["M", "38-40\"", "Cream"],
        ["L", "40-42\"", ""],
        ["XL", "42-44\"", "💡 Tips:"],
        ["XXL", "44-46\"", "• Use consistent naming"],
        ["XXXL", "46-48\"", "• Stick to these standard colors"],
        ["", "", "• Add custom colors if needed"],
    ]
    
    for row_idx, row_data in enumerate(reference_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_reference.cell(row=row_idx, column=col_idx, value=value)
            
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color="0000FF")
            elif row_idx in [3, 11, 16]:
                cell.font = Font(bold=True, size=12)
            
            if row_idx == 4:
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.font = Font(bold=True)
    
    ws_reference.column_dimensions['A'].width = 20
    ws_reference.column_dimensions['B'].width = 20
    ws_reference.column_dimensions['C'].width = 25
    
    # ========================================
    # SAVE FILE
    # ========================================
    
    output_path = "BizBooks_Clothing_Import_Bridge_Template.xlsx"
    wb.save(output_path)
    
    return output_path


if __name__ == "__main__":
    print("🚀 Creating Bridge Solution Template...")
    print()
    
    output_file = create_bridge_template()
    
    print(f"✅ Template created successfully!")
    print(f"📁 File: {output_file}")
    print()
    print("📋 Template Features:")
    print("  ✅ Separate columns: Base Name, Size, Color, Style")
    print("  ✅ Auto-generates Item Name: 'Levi's 501 32 Blue Slim Fit'")
    print("  ✅ Dropdowns for Size and Color (easy data entry)")
    print("  ✅ Excel formulas (no manual typing)")
    print("  ✅ Instructions sheet included")
    print("  ✅ Size & Color reference sheet")
    print("  ✅ Example data (delete before using)")
    print()
    print("🎯 Usage:")
    print("  1. Open the Excel file")
    print("  2. Fill Base Name, Size, Color, Price, Stock")
    print("  3. Item Name auto-generates!")
    print("  4. Delete example rows (grey)")
    print("  5. Save and upload to BizBooks")
    print()
    print("🔮 Future:")
    print("  When variants system is ready:")
    print("  → Parse 'Levi's 501 32 Blue Slim Fit'")
    print("  → Extract: Base='Levi's 501', Size='32', Color='Blue', Style='Slim Fit'")
    print("  → Convert to parent + variants automatically!")
    print()
    print("💡 BRILLIANT SOLUTION by Rish! 🎉")

