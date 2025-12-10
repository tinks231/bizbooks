#!/usr/bin/env python3
"""
Generate Branded Clothing Retail Dataset Excel File
With popular Indian brands - Option A (brand in product name)
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

def create_branded_clothing_dataset():
    wb = Workbook()
    ws = wb.active
    ws.title = "Branded Clothing"
    
    # Headers
    headers = ['Item Name*', 'SKU', 'Barcode', 'Category*', 'Group*', 'Unit*', 'Stock Quantity*', 
               'Cost Price*', 'Selling Price*', 'MRP', 'Tax Rate (%)', 'HSN Code', 'Description']
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Branded Products - Mix of popular Indian brands
    products = [
        # LEVI'S - Men's Jeans (Premium)
        ["Levi's Men's Jeans - Blue - 30", "LEVI-JNS-BLU-30", "8901234001001", "Men's Wear", "Jeans", "Pcs", 20, 1800, 2499, 2999, 12, "6109", "Levi's 511 slim fit jeans, authentic denim"],
        ["Levi's Men's Jeans - Blue - 32", "LEVI-JNS-BLU-32", "8901234001002", "Men's Wear", "Jeans", "Pcs", 35, 1800, 2499, 2999, 12, "6109", "Levi's 511 slim fit jeans, authentic denim"],
        ["Levi's Men's Jeans - Blue - 34", "LEVI-JNS-BLU-34", "8901234001003", "Men's Wear", "Jeans", "Pcs", 30, 1800, 2499, 2999, 12, "6109", "Levi's 511 slim fit jeans, authentic denim"],
        ["Levi's Men's Jeans - Black - 30", "LEVI-JNS-BLK-30", "8901234001004", "Men's Wear", "Jeans", "Pcs", 18, 1800, 2499, 2999, 12, "6109", "Levi's 511 slim fit jeans, black denim"],
        ["Levi's Men's Jeans - Black - 32", "LEVI-JNS-BLK-32", "8901234001005", "Men's Wear", "Jeans", "Pcs", 28, 1800, 2499, 2999, 12, "6109", "Levi's 511 slim fit jeans, black denim"],
        
        # WRANGLER - Men's Jeans (Mid-range)
        ["Wrangler Men's Jeans - Blue - 30", "WRAN-JNS-BLU-30", "8901234002001", "Men's Wear", "Jeans", "Pcs", 25, 1400, 1899, 2299, 12, "6109", "Wrangler regular fit jeans, comfortable"],
        ["Wrangler Men's Jeans - Blue - 32", "WRAN-JNS-BLU-32", "8901234002002", "Men's Wear", "Jeans", "Pcs", 40, 1400, 1899, 2299, 12, "6109", "Wrangler regular fit jeans, comfortable"],
        ["Wrangler Men's Jeans - Blue - 34", "WRAN-JNS-BLU-34", "8901234002003", "Men's Wear", "Jeans", "Pcs", 32, 1400, 1899, 2299, 12, "6109", "Wrangler regular fit jeans, comfortable"],
        ["Wrangler Men's Jeans - Black - 32", "WRAN-JNS-BLK-32", "8901234002004", "Men's Wear", "Jeans", "Pcs", 30, 1400, 1899, 2299, 12, "6109", "Wrangler regular fit jeans, black"],
        
        # ALLEN SOLLY - Men's Formal Shirts
        ["Allen Solly Men's Shirt - White - 38", "AS-SH-WHT-38", "8901234003001", "Men's Wear", "Shirts", "Pcs", 22, 1050, 1499, 1799, 12, "6109", "Allen Solly formal shirt, slim fit"],
        ["Allen Solly Men's Shirt - White - 40", "AS-SH-WHT-40", "8901234003002", "Men's Wear", "Shirts", "Pcs", 35, 1050, 1499, 1799, 12, "6109", "Allen Solly formal shirt, slim fit"],
        ["Allen Solly Men's Shirt - White - 42", "AS-SH-WHT-42", "8901234003003", "Men's Wear", "Shirts", "Pcs", 28, 1050, 1499, 1799, 12, "6109", "Allen Solly formal shirt, slim fit"],
        ["Allen Solly Men's Shirt - Blue - 38", "AS-SH-BLU-38", "8901234003004", "Men's Wear", "Shirts", "Pcs", 20, 1050, 1499, 1799, 12, "6109", "Allen Solly formal shirt, blue, slim fit"],
        ["Allen Solly Men's Shirt - Blue - 40", "AS-SH-BLU-40", "8901234003005", "Men's Wear", "Shirts", "Pcs", 30, 1050, 1499, 1799, 12, "6109", "Allen Solly formal shirt, blue, slim fit"],
        ["Allen Solly Men's Shirt - Pink - 40", "AS-SH-PNK-40", "8901234003006", "Men's Wear", "Shirts", "Pcs", 25, 1050, 1499, 1799, 12, "6109", "Allen Solly formal shirt, pink, slim fit"],
        
        # PETER ENGLAND - Men's Formal Shirts (Budget)
        ["Peter England Men's Shirt - White - 38", "PE-SH-WHT-38", "8901234004001", "Men's Wear", "Shirts", "Pcs", 28, 750, 999, 1199, 12, "6109", "Peter England formal shirt, regular fit"],
        ["Peter England Men's Shirt - White - 40", "PE-SH-WHT-40", "8901234004002", "Men's Wear", "Shirts", "Pcs", 40, 750, 999, 1199, 12, "6109", "Peter England formal shirt, regular fit"],
        ["Peter England Men's Shirt - Blue - 40", "PE-SH-BLU-40", "8901234004003", "Men's Wear", "Shirts", "Pcs", 35, 750, 999, 1199, 12, "6109", "Peter England formal shirt, blue"],
        ["Peter England Men's Shirt - Blue - 42", "PE-SH-BLU-42", "8901234004004", "Men's Wear", "Shirts", "Pcs", 30, 750, 999, 1199, 12, "6109", "Peter England formal shirt, blue"],
        
        # VAN HEUSEN - Men's Trousers
        ["Van Heusen Men's Trousers - Black - 32", "VH-TRS-BLK-32", "8901234005001", "Men's Wear", "Trousers", "Pcs", 22, 1350, 1899, 2299, 12, "6109", "Van Heusen formal trousers, flat front"],
        ["Van Heusen Men's Trousers - Black - 34", "VH-TRS-BLK-34", "8901234005002", "Men's Wear", "Trousers", "Pcs", 30, 1350, 1899, 2299, 12, "6109", "Van Heusen formal trousers, flat front"],
        ["Van Heusen Men's Trousers - Black - 36", "VH-TRS-BLK-36", "8901234005003", "Men's Wear", "Trousers", "Pcs", 25, 1350, 1899, 2299, 12, "6109", "Van Heusen formal trousers, flat front"],
        ["Van Heusen Men's Trousers - Grey - 32", "VH-TRS-GRY-32", "8901234005004", "Men's Wear", "Trousers", "Pcs", 20, 1350, 1899, 2299, 12, "6109", "Van Heusen formal trousers, grey"],
        
        # NIKE - Men's T-Shirts & Sportswear
        ["Nike Men's T-Shirt - White - M", "NIKE-TSH-WHT-M", "8901234006001", "Men's Wear", "T-Shirts", "Pcs", 40, 650, 899, 1099, 12, "6109", "Nike Dri-FIT t-shirt, breathable fabric"],
        ["Nike Men's T-Shirt - White - L", "NIKE-TSH-WHT-L", "8901234006002", "Men's Wear", "T-Shirts", "Pcs", 35, 650, 899, 1099, 12, "6109", "Nike Dri-FIT t-shirt, breathable fabric"],
        ["Nike Men's T-Shirt - Black - M", "NIKE-TSH-BLK-M", "8901234006003", "Men's Wear", "T-Shirts", "Pcs", 38, 650, 899, 1099, 12, "6109", "Nike Dri-FIT t-shirt, black"],
        ["Nike Men's T-Shirt - Black - L", "NIKE-TSH-BLK-L", "8901234006004", "Men's Wear", "T-Shirts", "Pcs", 32, 650, 899, 1099, 12, "6109", "Nike Dri-FIT t-shirt, black"],
        ["Nike Men's T-Shirt - Grey - M", "NIKE-TSH-GRY-M", "8901234006005", "Men's Wear", "T-Shirts", "Pcs", 30, 650, 899, 1099, 12, "6109", "Nike Dri-FIT t-shirt, grey heather"],
        
        # PUMA - Men's T-Shirts & Sportswear
        ["Puma Men's T-Shirt - White - M", "PUMA-TSH-WHT-M", "8901234007001", "Men's Wear", "T-Shirts", "Pcs", 35, 550, 799, 999, 12, "6109", "Puma essential t-shirt, cotton blend"],
        ["Puma Men's T-Shirt - White - L", "PUMA-TSH-WHT-L", "8901234007002", "Men's Wear", "T-Shirts", "Pcs", 30, 550, 799, 999, 12, "6109", "Puma essential t-shirt, cotton blend"],
        ["Puma Men's T-Shirt - Black - M", "PUMA-TSH-BLK-M", "8901234007003", "Men's Wear", "T-Shirts", "Pcs", 32, 550, 799, 999, 12, "6109", "Puma essential t-shirt, black"],
        ["Puma Men's T-Shirt - Navy - M", "PUMA-TSH-NVY-M", "8901234007004", "Men's Wear", "T-Shirts", "Pcs", 28, 550, 799, 999, 12, "6109", "Puma essential t-shirt, navy"],
        
        # BIBA - Women's Kurtas (Premium)
        ["Biba Women's Kurta - Pink - S", "BIBA-KRT-PNK-S", "8901234008001", "Women's Wear", "Kurtas", "Pcs", 25, 950, 1299, 1599, 12, "6109", "Biba printed kurta, cotton blend"],
        ["Biba Women's Kurta - Pink - M", "BIBA-KRT-PNK-M", "8901234008002", "Women's Wear", "Kurtas", "Pcs", 45, 950, 1299, 1599, 12, "6109", "Biba printed kurta, cotton blend"],
        ["Biba Women's Kurta - Pink - L", "BIBA-KRT-PNK-L", "8901234008003", "Women's Wear", "Kurtas", "Pcs", 35, 950, 1299, 1599, 12, "6109", "Biba printed kurta, cotton blend"],
        ["Biba Women's Kurta - Blue - M", "BIBA-KRT-BLU-M", "8901234008004", "Women's Wear", "Kurtas", "Pcs", 40, 950, 1299, 1599, 12, "6109", "Biba printed kurta, blue, cotton blend"],
        ["Biba Women's Kurta - Blue - L", "BIBA-KRT-BLU-L", "8901234008005", "Women's Wear", "Kurtas", "Pcs", 32, 950, 1299, 1599, 12, "6109", "Biba printed kurta, blue, cotton blend"],
        ["Biba Women's Kurta - Green - M", "BIBA-KRT-GRN-M", "8901234008006", "Women's Wear", "Kurtas", "Pcs", 30, 950, 1299, 1599, 12, "6109", "Biba printed kurta, green, cotton blend"],
        
        # W (Westside) - Women's Kurtas (Mid-range)
        ["W Women's Kurta - Blue - M", "W-KRT-BLU-M", "8901234009001", "Women's Wear", "Kurtas", "Pcs", 35, 650, 899, 1099, 12, "6109", "W by Westside kurta, casual wear"],
        ["W Women's Kurta - Blue - L", "W-KRT-BLU-L", "8901234009002", "Women's Wear", "Kurtas", "Pcs", 30, 650, 899, 1099, 12, "6109", "W by Westside kurta, casual wear"],
        ["W Women's Kurta - Pink - M", "W-KRT-PNK-M", "8901234009003", "Women's Wear", "Kurtas", "Pcs", 38, 650, 899, 1099, 12, "6109", "W by Westside kurta, pink, casual wear"],
        ["W Women's Kurta - White - M", "W-KRT-WHT-M", "8901234009004", "Women's Wear", "Kurtas", "Pcs", 32, 650, 899, 1099, 12, "6109", "W by Westside kurta, white, casual wear"],
        
        # FABINDIA - Women's Kurtas (Ethnic Premium)
        ["FabIndia Women's Kurta - Beige - M", "FBI-KRT-BGE-M", "8901234010001", "Women's Wear", "Kurtas", "Pcs", 20, 1200, 1699, 2099, 12, "6109", "FabIndia handloom kurta, pure cotton"],
        ["FabIndia Women's Kurta - Beige - L", "FBI-KRT-BGE-L", "8901234010002", "Women's Wear", "Kurtas", "Pcs", 18, 1200, 1699, 2099, 12, "6109", "FabIndia handloom kurta, pure cotton"],
        ["FabIndia Women's Kurta - Maroon - M", "FBI-KRT-MAR-M", "8901234010003", "Women's Wear", "Kurtas", "Pcs", 15, 1200, 1699, 2099, 12, "6109", "FabIndia handloom kurta, maroon"],
        
        # ZARA - Women's Tops & Dresses
        ["Zara Women's Top - White - S", "ZARA-TOP-WHT-S", "8901234011001", "Women's Wear", "Tops", "Pcs", 25, 650, 999, 1299, 12, "6109", "Zara casual top, trendy design"],
        ["Zara Women's Top - White - M", "ZARA-TOP-WHT-M", "8901234011002", "Women's Wear", "Tops", "Pcs", 40, 650, 999, 1299, 12, "6109", "Zara casual top, trendy design"],
        ["Zara Women's Top - Black - M", "ZARA-TOP-BLK-M", "8901234011003", "Women's Wear", "Tops", "Pcs", 35, 650, 999, 1299, 12, "6109", "Zara casual top, black"],
        ["Zara Women's Dress - Floral - M", "ZARA-DRS-FLR-M", "8901234011004", "Women's Wear", "Dresses", "Pcs", 20, 1400, 1999, 2499, 12, "6109", "Zara floral dress, summer collection"],
        ["Zara Women's Dress - Floral - L", "ZARA-DRS-FLR-L", "8901234011005", "Women's Wear", "Dresses", "Pcs", 18, 1400, 1999, 2499, 12, "6109", "Zara floral dress, summer collection"],
        
        # ONLY - Women's Jeans
        ["Only Women's Jeans - Blue - 28", "ONLY-JNS-BLU-28", "8901234012001", "Women's Wear", "Jeans", "Pcs", 22, 1350, 1799, 2199, 12, "6109", "Only skinny jeans, stretch denim"],
        ["Only Women's Jeans - Blue - 30", "ONLY-JNS-BLU-30", "8901234012002", "Women's Wear", "Jeans", "Pcs", 35, 1350, 1799, 2199, 12, "6109", "Only skinny jeans, stretch denim"],
        ["Only Women's Jeans - Blue - 32", "ONLY-JNS-BLU-32", "8901234012003", "Women's Wear", "Jeans", "Pcs", 28, 1350, 1799, 2199, 12, "6109", "Only skinny jeans, stretch denim"],
        ["Only Women's Jeans - Black - 30", "ONLY-JNS-BLK-30", "8901234012004", "Women's Wear", "Jeans", "Pcs", 25, 1350, 1799, 2199, 12, "6109", "Only skinny jeans, black"],
        
        # LEVI'S - Women's Jeans
        ["Levi's Women's Jeans - Blue - 28", "LEVI-JNS-BLU-W28", "8901234013001", "Women's Wear", "Jeans", "Pcs", 20, 1700, 2299, 2799, 12, "6109", "Levi's 721 high rise skinny jeans"],
        ["Levi's Women's Jeans - Blue - 30", "LEVI-JNS-BLU-W30", "8901234013002", "Women's Wear", "Jeans", "Pcs", 30, 1700, 2299, 2799, 12, "6109", "Levi's 721 high rise skinny jeans"],
        ["Levi's Women's Jeans - Black - 28", "LEVI-JNS-BLK-W28", "8901234013003", "Women's Wear", "Jeans", "Pcs", 18, 1700, 2299, 2799, 12, "6109", "Levi's 721 high rise skinny, black"],
        
        # VERO MODA - Women's Leggings
        ["Vero Moda Leggings - Black - S", "VM-LGG-BLK-S", "8901234014001", "Women's Wear", "Leggings", "Pcs", 30, 350, 499, 599, 12, "6109", "Vero Moda ankle length leggings"],
        ["Vero Moda Leggings - Black - M", "VM-LGG-BLK-M", "8901234014002", "Women's Wear", "Leggings", "Pcs", 60, 350, 499, 599, 12, "6109", "Vero Moda ankle length leggings"],
        ["Vero Moda Leggings - Black - L", "VM-LGG-BLK-L", "8901234014003", "Women's Wear", "Leggings", "Pcs", 45, 350, 499, 599, 12, "6109", "Vero Moda ankle length leggings"],
        ["Vero Moda Leggings - Navy - M", "VM-LGG-NVY-M", "8901234014004", "Women's Wear", "Leggings", "Pcs", 40, 350, 499, 599, 12, "6109", "Vero Moda ankle length leggings, navy"],
        
        # US POLO KIDS - Kids T-Shirts
        ["US Polo Kids T-Shirt - Blue - 6-7Y", "USP-TSH-BLU-6-7Y", "8901234015001", "Kids Wear", "T-Shirts", "Pcs", 20, 350, 499, 599, 5, "6109", "US Polo Kids cotton t-shirt"],
        ["US Polo Kids T-Shirt - Blue - 8-9Y", "USP-TSH-BLU-8-9Y", "8901234015002", "Kids Wear", "T-Shirts", "Pcs", 30, 350, 499, 599, 5, "6109", "US Polo Kids cotton t-shirt"],
        ["US Polo Kids T-Shirt - White - 8-9Y", "USP-TSH-WHT-8-9Y", "8901234015003", "Kids Wear", "T-Shirts", "Pcs", 28, 350, 499, 599, 5, "6109", "US Polo Kids cotton t-shirt, white"],
        ["US Polo Kids T-Shirt - White - 10-11Y", "USP-TSH-WHT-10-11Y", "8901234015004", "Kids Wear", "T-Shirts", "Pcs", 25, 350, 499, 599, 5, "6109", "US Polo Kids cotton t-shirt, white"],
        
        # GINI & JONY - Kids Wear
        ["Gini & Jony Kids Jeans - Blue - 8-9Y", "GJ-JNS-BLU-8-9Y", "8901234016001", "Kids Wear", "Jeans", "Pcs", 22, 650, 899, 1099, 5, "6109", "Gini & Jony kids denim jeans"],
        ["Gini & Jony Kids Jeans - Blue - 10-11Y", "GJ-JNS-BLU-10-11Y", "8901234016002", "Kids Wear", "Jeans", "Pcs", 20, 650, 899, 1099, 5, "6109", "Gini & Jony kids denim jeans"],
        ["Gini & Jony Kids Shorts - Blue - 8-9Y", "GJ-SHT-BLU-8-9Y", "8901234016003", "Kids Wear", "Shorts", "Pcs", 25, 350, 499, 599, 5, "6109", "Gini & Jony kids shorts, comfortable"],
        
        # BAGGIT - Women's Handbags
        ["Baggit Handbag - Brown", "BGT-BAG-BRN-001", "8901234017001", "Accessories", "Bags", "Pcs", 12, 1200, 1699, 1999, 12, "4203", "Baggit stylish handbag, vegan leather"],
        ["Baggit Handbag - Black", "BGT-BAG-BLK-001", "8901234017002", "Accessories", "Bags", "Pcs", 15, 1200, 1699, 1999, 12, "4203", "Baggit stylish handbag, black, vegan leather"],
        ["Baggit Purse - Red", "BGT-PURSE-RED-001", "8901234017003", "Accessories", "Bags", "Pcs", 10, 550, 799, 999, 12, "4203", "Baggit compact purse, red"],
        ["Baggit Purse - Black", "BGT-PURSE-BLK-001", "8901234017004", "Accessories", "Bags", "Pcs", 12, 550, 799, 999, 12, "4203", "Baggit compact purse, black"],
        
        # FASTRACK - Accessories
        ["Fastrack Belt - Black - M", "FT-BLT-BLK-M", "8901234018001", "Accessories", "Belts", "Pcs", 20, 450, 649, 799, 12, "4203", "Fastrack leather belt, size M (32-34)"],
        ["Fastrack Belt - Black - L", "FT-BLT-BLK-L", "8901234018002", "Accessories", "Belts", "Pcs", 25, 450, 649, 799, 12, "4203", "Fastrack leather belt, size L (36-38)"],
        ["Fastrack Belt - Brown - M", "FT-BLT-BRN-M", "8901234018003", "Accessories", "Belts", "Pcs", 18, 450, 649, 799, 12, "4203", "Fastrack leather belt, brown, size M"],
        
        # ADIDAS - Men's T-Shirts
        ["Adidas Men's T-Shirt - White - M", "ADI-TSH-WHT-M", "8901234019001", "Men's Wear", "T-Shirts", "Pcs", 30, 700, 999, 1299, 12, "6109", "Adidas essential t-shirt, 3-stripes"],
        ["Adidas Men's T-Shirt - White - L", "ADI-TSH-WHT-L", "8901234019002", "Men's Wear", "T-Shirts", "Pcs", 28, 700, 999, 1299, 12, "6109", "Adidas essential t-shirt, 3-stripes"],
        ["Adidas Men's T-Shirt - Black - M", "ADI-TSH-BLK-M", "8901234019003", "Men's Wear", "T-Shirts", "Pcs", 25, 700, 999, 1299, 12, "6109", "Adidas essential t-shirt, black"],
    ]
    
    # Add data rows
    for row_num, product in enumerate(products, start=2):
        for col_num, value in enumerate(product, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            # Format number columns
            if col_num in [7, 8, 9, 10, 11]:
                cell.number_format = '0.00' if col_num in [8, 9, 10] else '0'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 45
    
    # Save file
    output_path = "BizBooks_BRANDED_Clothing_90_Products.xlsx"
    wb.save(output_path)
    
    # Print summary
    brands = {}
    for product in products:
        brand = product[0].split()[0]  # Get first word (brand name)
        brands[brand] = brands.get(brand, 0) + 1
    
    print(f"✅ BRANDED Excel file created: {output_path}")
    print(f"📊 Total products: {len(products)}")
    print(f"\n🏷️ Brands included:")
    for brand, count in sorted(brands.items()):
        print(f"   {brand}: {count} products")
    print(f"\n📁 Location: {output_path}")
    print("\n🚀 Ready to import into BizBooks!")
    print("\n💡 All products have Brand + Product + Color + Size format")
    
    return output_path

if __name__ == "__main__":
    create_branded_clothing_dataset()

