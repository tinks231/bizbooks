#!/usr/bin/env python3
"""
Generate Sample Customer Dataset with DOB & Anniversary
15 customers for testing loyalty program
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

def create_customer_dataset():
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Import"
    
    # Headers
    headers = ['Customer Name*', 'Phone*', 'Email', 'GSTIN', 'Address', 'State', 'Credit Limit', 
               'Payment Terms (Days)', 'Opening Balance', 'Date of Birth', 'Anniversary Date', 'Notes']
    
    # Style headers
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000", size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Sample customers with realistic data
    customers = [
        # Retail customers
        ["Rishi Samaiya", "9876543210", "rishi.samaiya@gmail.com", "", "Shop 12, Market Street, Surat", "Gujarat", 10000, 15, 500, "1988-03-15", "2015-11-20", "Regular customer, prefers online payment"],
        ["Priya Sharma", "9876543211", "priya.sharma@yahoo.com", "", "A-201, Sunshine Apartments, Mumbai", "Maharashtra", 5000, 30, 0, "1992-07-22", "2018-02-14", "VIP customer"],
        ["Amit Patel", "9876543212", "amit.patel@gmail.com", "24AABCP1234F1Z5", "15, Green Plaza, Ahmedabad", "Gujarat", 15000, 30, 1200, "1985-12-10", "2010-05-25", "Wholesale buyer"],
        ["Sneha Verma", "9876543213", "sneha.v@outlook.com", "", "B-405, Royal Heights, Pune", "Maharashtra", 8000, 15, 0, "1995-01-08", "2020-12-01", "Birthday in January"],
        ["Rajesh Kumar", "9876543214", "rajesh.kumar@rediffmail.com", "27AABCR5678M1Z3", "Plot 7, Industrial Area, Delhi", "Delhi", 25000, 45, 3500, "1980-06-18", "2005-04-10", "B2B client, high volume"],
        
        # More customers with varied birthdays
        ["Neha Joshi", "9876543215", "neha.joshi@gmail.com", "", "C-102, Lake View, Bangalore", "Karnataka", 6000, 30, 0, "1990-09-25", "2016-07-15", ""],
        ["Vikram Singh", "9876543216", "vikram.s@yahoo.in", "", "56, MG Road, Jaipur", "Rajasthan", 12000, 30, 800, "1987-04-30", "2012-09-08", "Prefers cash payment"],
        ["Anjali Mehta", "9876543217", "anjali.mehta@hotmail.com", "24AABCM9012G1Z8", "Shop 3, City Center, Surat", "Gujarat", 20000, 60, 2500, "1983-11-05", "2008-03-22", "Corporate client"],
        ["Karan Thakur", "9876543218", "karan.thakur@gmail.com", "", "D-301, Skyline Towers, Hyderabad", "Telangana", 7000, 15, 0, "1993-08-17", "2019-11-30", ""],
        ["Divya Reddy", "9876543219", "divya.reddy@outlook.in", "", "12-A, Park Avenue, Chennai", "Tamil Nadu", 9000, 30, 350, "1991-02-28", "2017-06-10", "Birthday end of Feb"],
        
        # Customers with upcoming birthdays/anniversaries
        ["Rohit Agarwal", "9876543220", "rohit.ag@gmail.com", "", "45, Station Road, Indore", "Madhya Pradesh", 11000, 30, 0, "1989-12-15", "2014-08-25", "Birthday in December"],
        ["Pooja Gupta", "9876543221", "pooja.gupta@yahoo.com", "27AABCG3456H1Z9", "E-501, Ocean View, Mumbai", "Maharashtra", 18000, 45, 1800, "1986-05-20", "2011-01-17", "Wholesale dealer"],
        ["Suresh Rao", "9876543222", "suresh.rao@rediffmail.com", "", "23, Temple Street, Mysore", "Karnataka", 5000, 15, 0, "1994-10-12", "2021-02-28", ""],
        ["Kavita Desai", "9876543223", "kavita.desai@gmail.com", "", "F-102, City Heights, Nagpur", "Maharashtra", 8500, 30, 600, "1990-06-05", "2015-09-14", "Regular buyer"],
        ["Manish Jain", "9876543224", "manish.jain@outlook.com", "24AABCJ7890K1Z2", "78, Business Hub, Rajkot", "Gujarat", 22000, 60, 4200, "1982-03-28", "2007-12-20", "B2B premium client"],
    ]
    
    # Add data rows
    for row_num, customer in enumerate(customers, start=2):
        for col_num, value in enumerate(customer, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            # Format number columns
            if col_num in [7, 8, 9]:  # Credit Limit, Payment Terms, Opening Balance
                cell.number_format = '0.00' if col_num in [7, 9] else '0'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20  # Name
    ws.column_dimensions['B'].width = 15  # Phone
    ws.column_dimensions['C'].width = 28  # Email
    ws.column_dimensions['D'].width = 20  # GSTIN
    ws.column_dimensions['E'].width = 40  # Address
    ws.column_dimensions['F'].width = 15  # State
    ws.column_dimensions['G'].width = 14  # Credit Limit
    ws.column_dimensions['H'].width = 20  # Payment Terms
    ws.column_dimensions['I'].width = 18  # Opening Balance
    ws.column_dimensions['J'].width = 16  # DOB
    ws.column_dimensions['K'].width = 16  # Anniversary
    ws.column_dimensions['L'].width = 35  # Notes
    
    # Save file
    output_path = "BizBooks_Customer_Sample_15.xlsx"
    wb.save(output_path)
    
    print(f"✅ Customer dataset created: {output_path}")
    print(f"📊 Total customers: {len(customers)}")
    print(f"\n💼 Customer Summary:")
    print(f"   - With GSTIN: {sum(1 for c in customers if c[3])}")
    print(f"   - With Opening Balance: {sum(1 for c in customers if c[8] and float(c[8]) > 0)}")
    print(f"   - With Date of Birth: {len(customers)} (all)")
    print(f"   - With Anniversary: {len(customers)} (all)")
    print(f"\n🎂 Birthday Distribution:")
    months = {}
    for c in customers:
        if c[9]:  # DOB
            month = c[9].split('-')[1]
            month_name = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"][int(month)]
            months[month_name] = months.get(month_name, 0) + 1
    for m, count in sorted(months.items(), key=lambda x: ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"].index(x[0])):
        print(f"   {m}: {count} customer(s)")
    print(f"\n📁 Location: {output_path}")
    print("🚀 Ready to import into BizBooks!")
    
    return output_path

if __name__ == "__main__":
    create_customer_dataset()

