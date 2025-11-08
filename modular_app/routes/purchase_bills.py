from flask import Blueprint, render_template, request, redirect, url_for, flash, g, jsonify, session
from models import db, PurchaseBill, PurchaseBillItem, Vendor, Item, ItemStock, Site, Tenant, VendorPayment, PaymentAllocation
from utils.tenant_middleware import get_current_tenant_id
from utils.license_check import check_license
from datetime import datetime, date
from decimal import Decimal
from werkzeug.utils import secure_filename
from PIL import Image
from io import BytesIO
import pytz
import os

purchase_bills_bp = Blueprint('purchase_bills', __name__, url_prefix='/admin/purchase-bills')

@purchase_bills_bp.before_request
def check_auth():
    """Ensure user is logged in for all purchase bill routes"""
    if 'tenant_admin_id' not in session:
        flash('Please login to access purchase bills', 'error')
        return redirect(url_for('admin.login'))
    
    # Load tenant into g for easy access
    g.tenant = Tenant.query.get(session['tenant_admin_id'])
    if not g.tenant:
        session.clear()
        flash('Session expired. Please login again.', 'error')
        return redirect(url_for('admin.login'))

@purchase_bills_bp.route('/')
@check_license
def list_bills():
    """List all purchase bills"""
    tenant_id = get_current_tenant_id()
    
    # Get filter parameters
    status_filter = request.args.get('status', 'all')
    payment_filter = request.args.get('payment', 'all')
    search_query = request.args.get('search', '').strip()
    
    # Base query
    query = PurchaseBill.query.filter_by(tenant_id=tenant_id)
    
    # Apply filters
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    
    if payment_filter != 'all':
        query = query.filter_by(payment_status=payment_filter)
    
    if search_query:
        query = query.filter(
            db.or_(
                PurchaseBill.bill_number.like(f'%{search_query}%'),
                PurchaseBill.vendor_name.like(f'%{search_query}%'),
                PurchaseBill.reference_number.like(f'%{search_query}%')
            )
        )
    
    # Get bills sorted by date (newest first)
    bills = query.order_by(PurchaseBill.bill_date.desc()).all()
    
    # Calculate summary stats
    total_bills = len(bills)
    total_amount = sum([bill.total_amount for bill in bills])
    total_paid = sum([bill.paid_amount for bill in bills])
    total_due = sum([bill.balance_due for bill in bills])
    
    return render_template('admin/purchase_bills/list.html',
                         tenant=g.tenant,
                         bills=bills,
                         status_filter=status_filter,
                         payment_filter=payment_filter,
                         search_query=search_query,
                         total_bills=total_bills,
                         total_amount=total_amount,
                         total_paid=total_paid,
                         total_due=total_due)

@purchase_bills_bp.route('/create', methods=['GET', 'POST'])
@check_license
def create_bill():
    """Create new purchase bill"""
    tenant_id = get_current_tenant_id()
    
    if request.method == 'POST':
        try:
            # Create bill
            bill = PurchaseBill()
            bill.tenant_id = tenant_id
            bill.bill_number = bill.generate_bill_number()
            
            # Bill details
            bill.bill_date = datetime.strptime(request.form.get('bill_date'), '%Y-%m-%d').date()
            if request.form.get('due_date'):
                bill.due_date = datetime.strptime(request.form.get('due_date'), '%Y-%m-%d').date()
            
            # Vendor details
            vendor_id = request.form.get('vendor_id')
            if vendor_id:
                vendor = Vendor.query.get(vendor_id)
                if vendor:
                    bill.vendor_id = vendor.id
                    bill.vendor_name = vendor.name
                    bill.vendor_phone = vendor.phone
                    bill.vendor_email = vendor.email
                    bill.vendor_gstin = vendor.gstin
                    bill.vendor_address = vendor.address
                    bill.vendor_state = vendor.state or 'Maharashtra'
            else:
                # Manual vendor entry
                bill.vendor_name = request.form.get('vendor_name')
                bill.vendor_phone = request.form.get('vendor_phone', '')
                bill.vendor_email = request.form.get('vendor_email', '')
                bill.vendor_gstin = request.form.get('vendor_gstin', '')
                bill.vendor_address = request.form.get('vendor_address', '')
                bill.vendor_state = request.form.get('vendor_state', 'Maharashtra')
            
            # Additional fields
            bill.reference_number = request.form.get('reference_number', '')
            bill.payment_terms = request.form.get('payment_terms', '')
            bill.notes = request.form.get('notes', '')
            bill.terms_conditions = request.form.get('terms_conditions', '')
            
            # Handle file upload
            if 'bill_document' in request.files:
                file = request.files['bill_document']
                if file and file.filename:
                    try:
                        filename = secure_filename(file.filename)
                        file_ext = os.path.splitext(filename)[1].lower()
                        
                        # Check if running on Vercel (serverless)
                        if os.environ.get('VERCEL'):
                            # VERCEL: Upload to Vercel Blob Storage with compression
                            from utils.vercel_blob import upload_to_vercel_blob
                            
                            # Compress image before uploading (save blob storage space!)
                            if file_ext in ['.jpg', '.jpeg', '.png']:
                                img = Image.open(file)
                                
                                # Convert RGBA to RGB
                                if img.mode == 'RGBA':
                                    img = img.convert('RGB')
                                
                                # Resize if too large (max 1920px width)
                                max_width = 1920
                                if img.width > max_width:
                                    ratio = max_width / img.width
                                    new_height = int(img.height * ratio)
                                    img = img.resize((max_width, new_height), Image.LANCZOS)
                                
                                # Save compressed image to BytesIO
                                img_io = BytesIO()
                                img.save(img_io, format='JPEG', quality=70, optimize=True)
                                img_io.seek(0)
                                
                                # Generate filename
                                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                                blob_filename = f"purchase_bills/PB_{timestamp}.jpg"
                                
                                # Upload to Vercel Blob
                                blob_url = upload_to_vercel_blob(img_io, blob_filename, 'image/jpeg')
                                
                                if blob_url:
                                    bill.document_url = blob_url
                                    print(f"✅ Compressed & uploaded to Vercel Blob: {blob_url}")
                                else:
                                    print("⚠️ Blob upload failed, continuing without document")
                            else:
                                # For PDFs, upload directly (no compression)
                                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                                blob_filename = f"purchase_bills/PB_{timestamp}{file_ext}"
                                
                                blob_url = upload_to_vercel_blob(file, blob_filename, 'application/pdf')
                                
                                if blob_url:
                                    bill.document_url = blob_url
                                    print(f"✅ Uploaded PDF to Vercel Blob: {blob_url}")
                                else:
                                    print("⚠️ Blob upload failed, continuing without document")
                        else:
                            # LOCAL: Save to filesystem
                            upload_base = os.path.join('modular_app', 'uploads', 'purchase_bills', str(tenant_id))
                            os.makedirs(upload_base, exist_ok=True)
                            
                            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                            new_filename = f"PB_{timestamp}{file_ext}"
                            file_path = os.path.join(upload_base, new_filename)
                            
                            # Compress images for local storage too
                            if file_ext in ['.jpg', '.jpeg', '.png']:
                                img = Image.open(file)
                                if img.mode == 'RGBA':
                                    img = img.convert('RGB')
                                
                                max_width = 1920
                                if img.width > max_width:
                                    ratio = max_width / img.width
                                    new_height = int(img.height * ratio)
                                    img = img.resize((max_width, new_height), Image.LANCZOS)
                                
                                img.save(file_path, quality=75, optimize=True)
                                print(f"✅ Image compressed and saved locally: {file_path}")
                            else:
                                file.save(file_path)
                                print(f"✅ PDF saved locally: {file_path}")
                            
                            bill.document_url = f"/uploads/purchase_bills/{tenant_id}/{new_filename}"
                        
                    except Exception as file_error:
                        print(f"⚠️ Error uploading file: {str(file_error)}")
                        import traceback
                        traceback.print_exc()
                        # Don't fail the entire bill creation if file upload fails
            
            # Get line items
            item_ids = request.form.getlist('item_id[]')
            item_names = request.form.getlist('item_name[]')
            descriptions = request.form.getlist('description[]')
            hsn_codes = request.form.getlist('hsn_code[]')
            quantities = request.form.getlist('quantity[]')
            units = request.form.getlist('unit[]')
            rates = request.form.getlist('rate[]')
            gst_rates = request.form.getlist('gst_rate[]')
            
            # Calculate totals
            subtotal = Decimal('0')
            total_cgst = Decimal('0')
            total_sgst = Decimal('0')
            total_igst = Decimal('0')
            
            # Determine if IGST or CGST/SGST
            use_igst = (bill.vendor_state and bill.vendor_state.lower() != 'maharashtra')
            
            for i in range(len(item_names)):
                if not item_names[i].strip():
                    continue
                
                qty = Decimal(quantities[i]) if quantities[i] else Decimal('0')
                rate = Decimal(rates[i]) if rates[i] else Decimal('0')
                gst_rate = Decimal(gst_rates[i]) if gst_rates[i] else Decimal('0')
                
                # Calculate amounts
                line_total = qty * rate
                taxable_value = line_total
                
                if use_igst:
                    igst = (taxable_value * gst_rate / Decimal('100')).quantize(Decimal('0.01'))
                    cgst = Decimal('0')
                    sgst = Decimal('0')
                else:
                    cgst = (taxable_value * gst_rate / Decimal('200')).quantize(Decimal('0.01'))
                    sgst = cgst
                    igst = Decimal('0')
                
                item_total = taxable_value + cgst + sgst + igst
                
                # Create line item
                line_item = PurchaseBillItem()
                line_item.tenant_id = tenant_id
                
                # Link to inventory item if selected
                if item_ids[i] and item_ids[i].isdigit():
                    line_item.item_id = int(item_ids[i])
                
                line_item.item_name = item_names[i].strip()
                line_item.description = descriptions[i].strip() if i < len(descriptions) else ''
                line_item.hsn_code = hsn_codes[i].strip() if i < len(hsn_codes) else ''
                line_item.quantity = qty
                line_item.unit = units[i] if i < len(units) else 'pcs'
                line_item.rate = rate
                line_item.taxable_value = taxable_value
                line_item.gst_rate = gst_rate
                line_item.cgst_amount = cgst
                line_item.sgst_amount = sgst
                line_item.igst_amount = igst
                line_item.total_amount = item_total
                
                bill.items.append(line_item)
                
                # Add to totals
                subtotal += taxable_value
                total_cgst += cgst
                total_sgst += sgst
                total_igst += igst
            
            # Set bill totals
            bill.subtotal = subtotal
            bill.cgst_amount = total_cgst
            bill.sgst_amount = total_sgst
            bill.igst_amount = total_igst
            
            # Other charges and round-off
            other_charges = request.form.get('other_charges', '0')
            bill.other_charges = Decimal(other_charges) if other_charges else Decimal('0')
            
            # Calculate total
            calculated_total = subtotal + total_cgst + total_sgst + total_igst + bill.other_charges
            bill.total_amount = calculated_total.quantize(Decimal('0.01'))
            bill.balance_due = bill.total_amount
            
            # Set status
            bill.status = 'draft'
            bill.payment_status = 'unpaid'
            
            db.session.add(bill)
            db.session.commit()
            
            flash(f'✅ Purchase Bill {bill.bill_number} created successfully!', 'success')
            return redirect(url_for('purchase_bills.view_bill', bill_id=bill.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'❌ Error creating bill: {str(e)}', 'error')
            print(f"Error creating purchase bill: {str(e)}")
    
    # GET request - show form
    vendors = Vendor.query.filter_by(tenant_id=tenant_id, is_active=True).order_by(Vendor.name).all()
    sites = Site.query.filter_by(tenant_id=tenant_id).all()
    items = Item.query.filter_by(tenant_id=tenant_id).all()
    
    # Convert items to JSON-serializable format
    items_json = [
        {
            'id': item.id,
            'name': item.name,
            'item_code': item.sku or '',  # Use sku field
            'purchase_price': float(item.cost_price or 0),  # Use cost_price field
            'sale_price': float(item.selling_price or 0),  # Use selling_price field
            'hsn_code': item.hsn_code or '',
            'unit': item.unit or 'nos',  # Default is 'nos' in model
            'tax_preference': item.tax_preference or 'GST@18%'
        }
        for item in items
    ]
    
    # Convert vendors to JSON-serializable format
    vendors_json = [
        {
            'id': vendor.id,
            'name': vendor.name,
            'company_name': vendor.company_name or '',
            'phone': vendor.phone or '',
            'email': vendor.email or '',
            'gstin': vendor.gstin or '',
            'address': vendor.address or '',
            'state': vendor.state or 'Maharashtra'
        }
        for vendor in vendors
    ]
    
    return render_template('admin/purchase_bills/create.html',
                         tenant=g.tenant,
                         vendors=vendors_json,
                         sites=sites,
                         items=items_json,
                         today=date.today().strftime('%Y-%m-%d'))

@purchase_bills_bp.route('/<int:bill_id>')
@check_license
def view_bill(bill_id):
    """View purchase bill details"""
    tenant_id = get_current_tenant_id()
    
    try:
        bill = PurchaseBill.query.filter_by(id=bill_id, tenant_id=tenant_id).first_or_404()
        
        # Safely check if payment_allocations relationship exists
        try:
            _ = bill.payment_allocations
        except Exception as rel_error:
            print(f"⚠️ Warning: payment_allocations relationship error: {str(rel_error)}")
            # Create a temporary empty list if relationship doesn't work
            bill.payment_allocations = []
        
        return render_template('admin/purchase_bills/view.html',
                             tenant=g.tenant,
                             bill=bill)
    except Exception as e:
        print(f"❌ Error viewing purchase bill {bill_id}: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'❌ Error loading bill: {str(e)}', 'error')
        return redirect(url_for('purchase_bills.list_bills'))

@purchase_bills_bp.route('/<int:bill_id>/edit', methods=['GET', 'POST'])
@check_license
def edit_bill(bill_id):
    """Edit purchase bill (only if draft)"""
    tenant_id = get_current_tenant_id()
    bill = PurchaseBill.query.filter_by(id=bill_id, tenant_id=tenant_id).first_or_404()
    
    if bill.status != 'draft':
        flash('⚠️ Only draft bills can be edited', 'warning')
        return redirect(url_for('purchase_bills.view_bill', bill_id=bill.id))
    
    if request.method == 'POST':
        try:
            # Update bill details
            bill.bill_date = datetime.strptime(request.form.get('bill_date'), '%Y-%m-%d').date()
            if request.form.get('due_date'):
                bill.due_date = datetime.strptime(request.form.get('due_date'), '%Y-%m-%d').date()
            
            # Update vendor details
            vendor_id = request.form.get('vendor_id')
            if vendor_id:
                vendor = Vendor.query.get(vendor_id)
                if vendor:
                    bill.vendor_id = vendor.id
                    bill.vendor_name = vendor.name
                    bill.vendor_phone = vendor.phone
                    bill.vendor_email = vendor.email
                    bill.vendor_gstin = vendor.gstin
                    bill.vendor_address = vendor.address
                    bill.vendor_state = vendor.state or 'Maharashtra'
            
            bill.reference_number = request.form.get('reference_number', '')
            bill.payment_terms = request.form.get('payment_terms', '')
            bill.notes = request.form.get('notes', '')
            bill.terms_conditions = request.form.get('terms_conditions', '')
            
            # Delete existing items
            PurchaseBillItem.query.filter_by(purchase_bill_id=bill.id).delete()
            
            # Re-add items (same logic as create)
            item_names = request.form.getlist('item_name[]')
            quantities = request.form.getlist('quantity[]')
            rates = request.form.getlist('rate[]')
            gst_rates = request.form.getlist('gst_rate[]')
            
            subtotal = Decimal('0')
            total_cgst = Decimal('0')
            total_sgst = Decimal('0')
            total_igst = Decimal('0')
            
            use_igst = (bill.vendor_state and bill.vendor_state.lower() != 'maharashtra')
            
            for i in range(len(item_names)):
                if not item_names[i].strip():
                    continue
                
                qty = Decimal(quantities[i]) if quantities[i] else Decimal('0')
                rate = Decimal(rates[i]) if rates[i] else Decimal('0')
                gst_rate = Decimal(gst_rates[i]) if gst_rates[i] else Decimal('0')
                
                taxable_value = qty * rate
                
                if use_igst:
                    igst = (taxable_value * gst_rate / Decimal('100')).quantize(Decimal('0.01'))
                    cgst = Decimal('0')
                    sgst = Decimal('0')
                else:
                    cgst = (taxable_value * gst_rate / Decimal('200')).quantize(Decimal('0.01'))
                    sgst = cgst
                    igst = Decimal('0')
                
                item_total = taxable_value + cgst + sgst + igst
                
                line_item = PurchaseBillItem()
                line_item.tenant_id = tenant_id
                line_item.purchase_bill_id = bill.id
                line_item.item_name = item_names[i].strip()
                line_item.quantity = qty
                line_item.rate = rate
                line_item.taxable_value = taxable_value
                line_item.gst_rate = gst_rate
                line_item.cgst_amount = cgst
                line_item.sgst_amount = sgst
                line_item.igst_amount = igst
                line_item.total_amount = item_total
                
                db.session.add(line_item)
                
                subtotal += taxable_value
                total_cgst += cgst
                total_sgst += sgst
                total_igst += igst
            
            bill.subtotal = subtotal
            bill.cgst_amount = total_cgst
            bill.sgst_amount = total_sgst
            bill.igst_amount = total_igst
            
            other_charges = request.form.get('other_charges', '0')
            bill.other_charges = Decimal(other_charges) if other_charges else Decimal('0')
            
            bill.total_amount = (subtotal + total_cgst + total_sgst + total_igst + bill.other_charges).quantize(Decimal('0.01'))
            bill.balance_due = bill.total_amount - bill.paid_amount
            
            db.session.commit()
            
            flash(f'✅ Purchase Bill {bill.bill_number} updated successfully!', 'success')
            return redirect(url_for('purchase_bills.view_bill', bill_id=bill.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'❌ Error updating bill: {str(e)}', 'error')
    
    vendors = Vendor.query.filter_by(tenant_id=tenant_id, is_active=True).order_by(Vendor.name).all()
    
    return render_template('admin/purchase_bills/edit.html',
                         tenant=g.tenant,
                         bill=bill,
                         vendors=vendors)

@purchase_bills_bp.route('/<int:bill_id>/approve', methods=['POST'])
@check_license
def approve_bill(bill_id):
    """Approve purchase bill and update inventory using Weighted Average Cost"""
    from models import Item, ItemStock, ItemStockMovement, Site
    
    tenant_id = get_current_tenant_id()
    bill = PurchaseBill.query.filter_by(id=bill_id, tenant_id=tenant_id).first_or_404()
    
    if bill.status != 'draft':
        flash('⚠️ Bill is already approved or cancelled', 'warning')
        return redirect(url_for('purchase_bills.view_bill', bill_id=bill.id))
    
    try:
        # Get first active site for inventory updates
        default_site = Site.query.filter_by(tenant_id=tenant_id, active=True).first()
        
        if not default_site:
            # If no active site, try any site
            default_site = Site.query.filter_by(tenant_id=tenant_id).first()
        
        if not default_site:
            flash('⚠️ No site found for inventory updates. Please create a site first.', 'warning')
            return redirect(url_for('purchase_bills.view_bill', bill_id=bill.id))
        
        inventory_updates = []
        
        # Process each line item and update inventory
        for line_item in bill.items:
            if not line_item.item_id:
                # Skip items not linked to inventory
                print(f"⏭️ Skipping '{line_item.item_name}' - not linked to inventory item")
                continue
            
            item = Item.query.get(line_item.item_id)
            if not item:
                print(f"⚠️ Item ID {line_item.item_id} not found, skipping")
                continue
            
            # Get or create stock record for this item at default site
            stock = ItemStock.query.filter_by(
                tenant_id=tenant_id,
                item_id=item.id,
                site_id=default_site.id
            ).first()
            
            if not stock:
                # Create new stock record
                stock = ItemStock(
                    tenant_id=tenant_id,
                    item_id=item.id,
                    site_id=default_site.id,
                    quantity_available=0.0,
                    stock_value=0.0,
                    valuation_method='WAC'  # Weighted Average Cost
                )
                db.session.add(stock)
                db.session.flush()  # Get stock ID
                print(f"📦 Created new stock record for {item.name}")
            
            # Calculate Weighted Average Cost
            old_qty = float(stock.quantity_available)
            old_cost = float(item.cost_price or 0)
            old_value = old_qty * old_cost
            
            new_qty = float(line_item.quantity)
            new_cost = float(line_item.rate)  # Rate from purchase bill
            new_value = new_qty * new_cost
            
            total_qty = old_qty + new_qty
            total_value = old_value + new_value
            
            # Calculate weighted average cost
            if total_qty > 0:
                weighted_avg_cost = total_value / total_qty
            else:
                weighted_avg_cost = new_cost
            
            # Update item cost price (Weighted Average)
            old_item_cost = item.cost_price
            item.cost_price = weighted_avg_cost
            
            # Update stock quantity and value
            stock.quantity_available = total_qty
            stock.stock_value = total_value
            stock.last_stock_date = datetime.now(pytz.timezone('Asia/Kolkata'))
            
            # Create stock movement record for audit trail
            movement = ItemStockMovement(
                tenant_id=tenant_id,
                item_id=item.id,
                site_id=default_site.id,
                movement_type='stock_in',
                quantity=new_qty,
                unit_cost=new_cost,
                total_value=new_value,
                reference_number=bill.bill_number,
                reference_type='purchase_bill',
                reference_id=bill.id,
                reason='Purchase Bill Approval',
                notes=f"Vendor: {bill.vendor_name or 'Unknown'}",
                created_by=session.get('username', 'Admin')
            )
            db.session.add(movement)
            
            inventory_updates.append({
                'item': item.name,
                'old_qty': old_qty,
                'new_qty': new_qty,
                'total_qty': total_qty,
                'old_cost': old_item_cost,
                'new_cost': new_cost,
                'weighted_avg': weighted_avg_cost
            })
            
            print(f"✅ Updated inventory for {item.name}:")
            print(f"   Quantity: {old_qty} + {new_qty} = {total_qty}")
            print(f"   Cost: ₹{old_item_cost:.2f} → ₹{weighted_avg_cost:.2f} (WAC)")
            print(f"   Calculation: (₹{old_value:.2f} + ₹{new_value:.2f}) / {total_qty} = ₹{weighted_avg_cost:.2f}")
        
        # Update bill status
        bill.status = 'approved'
        bill.approved_at = datetime.now(pytz.timezone('Asia/Kolkata'))
        bill.approved_by = session.get('tenant_admin_id')
        
        db.session.commit()
        
        # Build success message
        if inventory_updates:
            update_summary = f"Updated inventory for {len(inventory_updates)} items using Weighted Average Cost"
            flash(f'✅ Purchase Bill {bill.bill_number} approved! {update_summary}', 'success')
        else:
            flash(f'✅ Purchase Bill {bill.bill_number} approved! (No inventory items to update)', 'success')
            
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"❌ Error approving bill: {str(e)}")
        print(traceback.format_exc())
        flash(f'❌ Error approving bill: {str(e)}', 'error')
    
    return redirect(url_for('purchase_bills.view_bill', bill_id=bill.id))

@purchase_bills_bp.route('/<int:bill_id>/print')
@check_license
def print_bill(bill_id):
    """Print-friendly view of purchase bill"""
    tenant_id = get_current_tenant_id()
    
    try:
        bill = PurchaseBill.query.filter_by(id=bill_id, tenant_id=tenant_id).first_or_404()
        
        return render_template('admin/purchase_bills/print.html',
                             tenant=g.tenant,
                             bill=bill)
    except Exception as e:
        print(f"❌ Error loading purchase bill for print {bill_id}: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'❌ Error loading bill: {str(e)}', 'error')
        return redirect(url_for('purchase_bills.list_bills'))

@purchase_bills_bp.route('/<int:bill_id>/delete', methods=['POST'])
@check_license
def delete_bill(bill_id):
    """Delete purchase bill (only if draft)"""
    tenant_id = get_current_tenant_id()
    bill = PurchaseBill.query.filter_by(id=bill_id, tenant_id=tenant_id).first_or_404()
    
    if bill.status != 'draft':
        flash('⚠️ Only draft bills can be deleted', 'warning')
        return redirect(url_for('purchase_bills.view_bill', bill_id=bill.id))
    
    try:
        db.session.delete(bill)
        db.session.commit()
        
        flash(f'✅ Purchase Bill {bill.bill_number} deleted successfully!', 'success')
        return redirect(url_for('purchase_bills.list_bills'))
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error deleting bill: {str(e)}', 'error')
        return redirect(url_for('purchase_bills.view_bill', bill_id=bill.id))

@purchase_bills_bp.route('/<int:bill_id>/record-payment', methods=['GET', 'POST'])
@check_license
def record_payment_form(bill_id):
    """Record payment for purchase bill"""
    tenant_id = get_current_tenant_id()
    bill = PurchaseBill.query.filter_by(id=bill_id, tenant_id=tenant_id).first_or_404()
    
    if bill.status != 'approved':
        flash('⚠️ Only approved bills can have payments recorded', 'warning')
        return redirect(url_for('purchase_bills.view_bill', bill_id=bill.id))
    
    if request.method == 'POST':
        try:
            # Get payment details
            payment_date_str = request.form.get('payment_date')
            payment_date = datetime.strptime(payment_date_str, '%Y-%m-%d').date()
            amount = Decimal(request.form.get('amount', '0'))
            payment_method = request.form.get('payment_method', 'cash')
            reference_number = request.form.get('reference_number', '')
            bank_account = request.form.get('bank_account', '')
            notes = request.form.get('notes', '')
            
            # Validate amount
            if amount <= 0:
                flash('⚠️ Payment amount must be greater than 0', 'warning')
                return redirect(url_for('purchase_bills.record_payment_form', bill_id=bill.id))
            
            if amount > bill.balance_due:
                flash('⚠️ Payment amount cannot exceed balance due', 'warning')
                return redirect(url_for('purchase_bills.record_payment_form', bill_id=bill.id))
            
            # AUTO-CREATE VENDOR if bill doesn't have vendor_id (manual entry)
            if not bill.vendor_id and bill.vendor_name:
                # Check if vendor already exists (by GSTIN or name)
                existing_vendor = None
                
                if bill.vendor_gstin:
                    existing_vendor = Vendor.query.filter_by(
                        tenant_id=tenant_id,
                        gstin=bill.vendor_gstin
                    ).first()
                
                if not existing_vendor:
                    # Try by name (case-insensitive)
                    existing_vendor = Vendor.query.filter(
                        Vendor.tenant_id == tenant_id,
                        db.func.lower(Vendor.name) == db.func.lower(bill.vendor_name)
                    ).first()
                
                if existing_vendor:
                    # Use existing vendor
                    bill.vendor_id = existing_vendor.id
                    print(f"✅ Linked bill to existing vendor: {existing_vendor.name} (ID: {existing_vendor.id})")
                else:
                    # Create new vendor from bill details
                    new_vendor = Vendor()
                    new_vendor.tenant_id = tenant_id
                    new_vendor.name = bill.vendor_name
                    new_vendor.company_name = bill.vendor_name
                    new_vendor.phone = bill.vendor_phone or ''
                    new_vendor.email = bill.vendor_email or ''
                    new_vendor.gstin = bill.vendor_gstin or ''
                    new_vendor.address = bill.vendor_address or ''
                    new_vendor.state = bill.vendor_state or 'Maharashtra'
                    new_vendor.is_active = True
                    
                    # Generate vendor code
                    last_vendor = Vendor.query.filter_by(tenant_id=tenant_id).order_by(Vendor.id.desc()).first()
                    if last_vendor and last_vendor.vendor_code:
                        try:
                            last_num = int(last_vendor.vendor_code.replace('VEN-', ''))
                            new_vendor.vendor_code = f'VEN-{last_num + 1:04d}'
                        except:
                            new_vendor.vendor_code = 'VEN-0001'
                    else:
                        new_vendor.vendor_code = 'VEN-0001'
                    
                    db.session.add(new_vendor)
                    db.session.flush()  # Get vendor ID
                    
                    # Link bill to new vendor
                    bill.vendor_id = new_vendor.id
                    
                    print(f"✅ Auto-created vendor: {new_vendor.name} (ID: {new_vendor.id}, Code: {new_vendor.vendor_code})")
            
            # Create payment
            payment = VendorPayment()
            payment.tenant_id = tenant_id
            payment.payment_number = payment.generate_payment_number()
            payment.payment_date = payment_date
            payment.vendor_id = bill.vendor_id  # Now guaranteed to have a value
            payment.vendor_name = bill.vendor_name
            payment.amount = amount
            payment.payment_method = payment_method
            payment.reference_number = reference_number
            payment.bank_account = bank_account
            payment.notes = notes
            payment.created_by = session.get('tenant_name', 'Admin')
            
            db.session.add(payment)
            db.session.flush()  # Get payment ID
            
            # Create payment allocation
            allocation = PaymentAllocation()
            allocation.payment_id = payment.id
            allocation.purchase_bill_id = bill.id
            allocation.amount_allocated = amount
            
            db.session.add(allocation)
            
            # Update bill payment status
            bill.paid_amount = (bill.paid_amount or Decimal('0')) + amount
            bill.balance_due = bill.total_amount - bill.paid_amount
            
            if bill.balance_due <= 0:
                bill.payment_status = 'paid'
                bill.balance_due = Decimal('0')
            elif bill.paid_amount > 0:
                bill.payment_status = 'partial'
            
            db.session.commit()
            
            flash(f'✅ Payment of ₹{amount:,.2f} recorded successfully!', 'success')
            return redirect(url_for('purchase_bills.view_bill', bill_id=bill.id))
            
        except ValueError as e:
            flash('⚠️ Invalid payment amount', 'error')
            return redirect(url_for('purchase_bills.record_payment_form', bill_id=bill.id))
        except Exception as e:
            db.session.rollback()
            flash(f'❌ Error recording payment: {str(e)}', 'error')
            print(f"Error recording payment: {str(e)}")
            import traceback
            traceback.print_exc()
    
    # GET - show form
    return render_template('admin/purchase_bills/record_payment.html',
                         tenant=g.tenant,
                         bill=bill,
                         today=date.today().strftime('%Y-%m-%d'))

# API Endpoints

@purchase_bills_bp.route('/api/search-items')
def api_search_items():
    """API endpoint to search items"""
    tenant_id = get_current_tenant_id()
    query = request.args.get('q', '').strip()
    
    if len(query) < 2:
        return jsonify([])
    
    items = Item.query.filter_by(tenant_id=tenant_id).filter(
        db.or_(
            Item.name.like(f'%{query}%'),
            Item.item_code.like(f'%{query}%')
        )
    ).limit(20).all()
    
    results = []
    for item in items:
        results.append({
            'id': item.id,
            'name': item.name,
            'item_code': item.item_code,
            'unit': item.unit or 'pcs',
            'purchase_price': float(item.cost_price or 0),
            'sale_price': float(item.selling_price or 0),
            'hsn_code': item.hsn_code or '',
            'gst_rate': float(item.gst_rate) if item.gst_rate is not None else 18.0
        })
    
    return jsonify(results)

@purchase_bills_bp.route('/gstr2')
@check_license
def gstr2_report():
    """GSTR-2 Report - Input Tax Credit Report"""
    from datetime import datetime, timedelta
    from decimal import Decimal
    from collections import defaultdict
    
    tenant_id = get_current_tenant_id()
    
    # Get date range (default to current month)
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')
    
    if not start_date_str:
        # Default to start of current month
        today = datetime.now()
        start_date = datetime(today.year, today.month, 1).date()
        start_date_str = start_date.strftime('%Y-%m-%d')
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    
    if not end_date_str:
        # Default to today
        end_date = datetime.now().date()
        end_date_str = end_date.strftime('%Y-%m-%d')
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    # Get all approved purchase bills with GST in date range
    # Only include bills that have GST (either CGST+SGST or IGST > 0)
    bills = PurchaseBill.query.filter(
        PurchaseBill.tenant_id == tenant_id,
        PurchaseBill.status == 'approved',
        PurchaseBill.bill_date >= start_date,
        PurchaseBill.bill_date <= end_date,
        db.or_(
            PurchaseBill.cgst_amount > 0,
            PurchaseBill.sgst_amount > 0,
            PurchaseBill.igst_amount > 0
        )
    ).order_by(PurchaseBill.bill_date.asc()).all()
    
    # Calculate summaries
    total_taxable_value = Decimal('0')
    total_cgst = Decimal('0')
    total_sgst = Decimal('0')
    total_igst = Decimal('0')
    total_amount = Decimal('0')
    
    # HSN-wise summary
    hsn_summary = defaultdict(lambda: {
        'hsn_code': '',
        'description': '',
        'uqc': '',
        'total_quantity': Decimal('0'),
        'total_value': Decimal('0'),
        'taxable_value': Decimal('0'),
        'cgst': Decimal('0'),
        'sgst': Decimal('0'),
        'igst': Decimal('0'),
        'total_tax': Decimal('0')
    })
    
    # Vendor-wise summary
    vendor_summary = defaultdict(lambda: {
        'vendor_name': '',
        'vendor_gstin': '',
        'bill_count': 0,
        'total_taxable_value': Decimal('0'),
        'total_cgst': Decimal('0'),
        'total_sgst': Decimal('0'),
        'total_igst': Decimal('0'),
        'total_amount': Decimal('0')
    })
    
    # Process each bill
    for bill in bills:
        # Bill totals
        total_taxable_value += bill.subtotal
        total_cgst += bill.cgst_amount
        total_sgst += bill.sgst_amount
        total_igst += bill.igst_amount
        total_amount += bill.total_amount
        
        # Vendor summary
        vendor_key = bill.vendor_id or bill.vendor_name
        vendor_summary[vendor_key]['vendor_name'] = bill.vendor_name
        vendor_summary[vendor_key]['vendor_gstin'] = bill.vendor_gstin or 'N/A'
        vendor_summary[vendor_key]['bill_count'] += 1
        vendor_summary[vendor_key]['total_taxable_value'] += bill.subtotal
        vendor_summary[vendor_key]['total_cgst'] += bill.cgst_amount
        vendor_summary[vendor_key]['total_sgst'] += bill.sgst_amount
        vendor_summary[vendor_key]['total_igst'] += bill.igst_amount
        vendor_summary[vendor_key]['total_amount'] += bill.total_amount
        
        # HSN summary - process line items
        for item in bill.items:
            hsn_code = item.hsn_code or 'N/A'
            
            hsn_summary[hsn_code]['hsn_code'] = hsn_code
            hsn_summary[hsn_code]['description'] = item.item_name if not hsn_summary[hsn_code]['description'] else hsn_summary[hsn_code]['description']
            hsn_summary[hsn_code]['uqc'] = item.unit
            hsn_summary[hsn_code]['total_quantity'] += item.quantity
            hsn_summary[hsn_code]['taxable_value'] += item.taxable_value
            hsn_summary[hsn_code]['cgst'] += item.cgst_amount
            hsn_summary[hsn_code]['sgst'] += item.sgst_amount
            hsn_summary[hsn_code]['igst'] += item.igst_amount
            hsn_summary[hsn_code]['total_tax'] += (item.cgst_amount + item.sgst_amount + item.igst_amount)
            hsn_summary[hsn_code]['total_value'] += item.total_amount
    
    # Convert defaultdicts to lists
    hsn_list = sorted(hsn_summary.values(), key=lambda x: x['hsn_code'])
    vendor_list = sorted(vendor_summary.values(), key=lambda x: x['vendor_name'])
    
    # ITC Summary
    itc_summary = {
        'total_taxable_value': float(total_taxable_value),
        'total_cgst': float(total_cgst),
        'total_sgst': float(total_sgst),
        'total_igst': float(total_igst),
        'total_gst': float(total_cgst + total_sgst + total_igst),
        'total_amount': float(total_amount),
        'bill_count': len(bills)
    }
    
    return render_template('admin/purchase_bills/gstr2.html',
                         tenant=g.tenant,
                         bills=bills,
                         start_date=start_date_str,
                         end_date=end_date_str,
                         itc_summary=itc_summary,
                         hsn_summary=hsn_list,
                         vendor_summary=vendor_list)


@purchase_bills_bp.route('/gstr2/export-excel')
@login_required
def gstr2_export_excel():
    """Export GSTR-2 Report to Excel (Real .xlsx format)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime, timedelta
    from decimal import Decimal
    from collections import defaultdict
    from io import BytesIO
    from flask import send_file
    
    tenant_id = get_current_tenant_id()
    
    # Get date range (same logic as gstr2_report)
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')
    
    if not start_date_str:
        today = datetime.now()
        start_date = datetime(today.year, today.month, 1).date()
        start_date_str = start_date.strftime('%Y-%m-%d')
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    
    if not end_date_str:
        end_date = datetime.now().date()
        end_date_str = end_date.strftime('%Y-%m-%d')
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    # Get all approved purchase bills with GST
    bills = PurchaseBill.query.filter(
        PurchaseBill.tenant_id == tenant_id,
        PurchaseBill.status == 'approved',
        PurchaseBill.bill_date >= start_date,
        PurchaseBill.bill_date <= end_date,
        db.or_(
            PurchaseBill.cgst_amount > 0,
            PurchaseBill.sgst_amount > 0,
            PurchaseBill.igst_amount > 0
        )
    ).order_by(PurchaseBill.bill_date.asc()).all()
    
    # Calculate summaries (same logic)
    total_taxable_value = Decimal('0')
    total_cgst = Decimal('0')
    total_sgst = Decimal('0')
    total_igst = Decimal('0')
    total_amount = Decimal('0')
    
    hsn_summary = defaultdict(lambda: {
        'hsn_code': '',
        'description': '',
        'uqc': '',
        'total_quantity': Decimal('0'),
        'total_value': Decimal('0'),
        'taxable_value': Decimal('0'),
        'cgst': Decimal('0'),
        'sgst': Decimal('0'),
        'igst': Decimal('0'),
        'total_tax': Decimal('0')
    })
    
    vendor_summary = defaultdict(lambda: {
        'vendor_name': '',
        'vendor_gstin': '',
        'bill_count': 0,
        'total_taxable_value': Decimal('0'),
        'total_cgst': Decimal('0'),
        'total_sgst': Decimal('0'),
        'total_igst': Decimal('0'),
        'total_amount': Decimal('0')
    })
    
    for bill in bills:
        total_taxable_value += bill.subtotal
        total_cgst += bill.cgst_amount
        total_sgst += bill.sgst_amount
        total_igst += bill.igst_amount
        total_amount += bill.total_amount
        
        vendor_key = bill.vendor_id or bill.vendor_name
        vendor_summary[vendor_key]['vendor_name'] = bill.vendor_name
        vendor_summary[vendor_key]['vendor_gstin'] = bill.vendor_gstin or 'N/A'
        vendor_summary[vendor_key]['bill_count'] += 1
        vendor_summary[vendor_key]['total_taxable_value'] += bill.subtotal
        vendor_summary[vendor_key]['total_cgst'] += bill.cgst_amount
        vendor_summary[vendor_key]['total_sgst'] += bill.sgst_amount
        vendor_summary[vendor_key]['total_igst'] += bill.igst_amount
        vendor_summary[vendor_key]['total_amount'] += bill.total_amount
        
        for item in bill.items:
            hsn_code = item.hsn_code or 'N/A'
            hsn_summary[hsn_code]['hsn_code'] = hsn_code
            hsn_summary[hsn_code]['description'] = item.item_name if not hsn_summary[hsn_code]['description'] else hsn_summary[hsn_code]['description']
            hsn_summary[hsn_code]['uqc'] = item.unit
            hsn_summary[hsn_code]['total_quantity'] += item.quantity
            hsn_summary[hsn_code]['taxable_value'] += item.taxable_value
            hsn_summary[hsn_code]['cgst'] += item.cgst_amount
            hsn_summary[hsn_code]['sgst'] += item.sgst_amount
            hsn_summary[hsn_code]['igst'] += item.igst_amount
            hsn_summary[hsn_code]['total_tax'] += (item.cgst_amount + item.sgst_amount + item.igst_amount)
            hsn_summary[hsn_code]['total_value'] += item.total_amount
    
    hsn_list = sorted(hsn_summary.values(), key=lambda x: x['hsn_code'])
    vendor_list = sorted(vendor_summary.values(), key=lambda x: x['vendor_name'])
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "GSTR-2 Report"
    
    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = g.tenant.company_name
    ws['A1'].font = title_font
    ws['A2'] = f'GSTR-2 Report: {start_date_str} to {end_date_str}'
    ws['A2'].font = Font(bold=True)
    
    row = 4
    
    # ITC Summary
    ws[f'A{row}'] = 'ITC Summary'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    headers = ['Taxable Value', 'CGST', 'SGST', 'IGST', 'Total ITC', 'Bill Count']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row += 1
    ws.cell(row, 1, float(total_taxable_value)).number_format = '#,##0.00'
    ws.cell(row, 2, float(total_cgst)).number_format = '#,##0.00'
    ws.cell(row, 3, float(total_sgst)).number_format = '#,##0.00'
    ws.cell(row, 4, float(total_igst)).number_format = '#,##0.00'
    ws.cell(row, 5, float(total_cgst + total_sgst + total_igst)).number_format = '#,##0.00'
    ws.cell(row, 6, len(bills))
    
    for col in range(1, 7):
        ws.cell(row, col).border = border
        ws.cell(row, col).alignment = Alignment(horizontal='right')
    
    row += 3
    
    # Purchase Register
    ws[f'A{row}'] = 'Purchase Register'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    headers = ['Date', 'Bill No.', 'Vendor Name', 'GSTIN', 'State', 'Taxable Value', 'CGST', 'SGST', 'IGST', 'Total Amount']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row += 1
    for bill in bills:
        ws.cell(row, 1, bill.bill_date.strftime('%d-%m-%Y'))
        ws.cell(row, 2, bill.bill_number)
        ws.cell(row, 3, bill.vendor_name)
        ws.cell(row, 4, bill.vendor_gstin or 'N/A')
        ws.cell(row, 5, bill.vendor_state or 'Maharashtra')
        ws.cell(row, 6, float(bill.subtotal)).number_format = '#,##0.00'
        ws.cell(row, 7, float(bill.cgst_amount)).number_format = '#,##0.00'
        ws.cell(row, 8, float(bill.sgst_amount)).number_format = '#,##0.00'
        ws.cell(row, 9, float(bill.igst_amount)).number_format = '#,##0.00'
        ws.cell(row, 10, float(bill.total_amount)).number_format = '#,##0.00'
        
        for col in range(1, 11):
            ws.cell(row, col).border = border
        
        row += 1
    
    # Total row
    ws.cell(row, 1, 'TOTAL').font = Font(bold=True)
    ws.cell(row, 6, float(total_taxable_value)).number_format = '#,##0.00'
    ws.cell(row, 6).font = Font(bold=True)
    ws.cell(row, 7, float(total_cgst)).number_format = '#,##0.00'
    ws.cell(row, 7).font = Font(bold=True)
    ws.cell(row, 8, float(total_sgst)).number_format = '#,##0.00'
    ws.cell(row, 8).font = Font(bold=True)
    ws.cell(row, 9, float(total_igst)).number_format = '#,##0.00'
    ws.cell(row, 9).font = Font(bold=True)
    ws.cell(row, 10, float(total_amount)).number_format = '#,##0.00'
    ws.cell(row, 10).font = Font(bold=True)
    
    for col in range(1, 11):
        ws.cell(row, col).border = border
        ws.cell(row, col).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    row += 3
    
    # HSN Summary
    ws[f'A{row}'] = 'HSN-wise Summary'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    headers = ['HSN Code', 'Description', 'UQC', 'Quantity', 'Taxable Value', 'CGST', 'SGST', 'IGST', 'Total']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row += 1
    for hsn in hsn_list:
        ws.cell(row, 1, hsn['hsn_code'])
        ws.cell(row, 2, hsn['description'])
        ws.cell(row, 3, hsn['uqc'])
        ws.cell(row, 4, float(hsn['total_quantity'])).number_format = '#,##0.00'
        ws.cell(row, 5, float(hsn['taxable_value'])).number_format = '#,##0.00'
        ws.cell(row, 6, float(hsn['cgst'])).number_format = '#,##0.00'
        ws.cell(row, 7, float(hsn['sgst'])).number_format = '#,##0.00'
        ws.cell(row, 8, float(hsn['igst'])).number_format = '#,##0.00'
        ws.cell(row, 9, float(hsn['total_value'])).number_format = '#,##0.00'
        
        for col in range(1, 10):
            ws.cell(row, col).border = border
        
        row += 1
    
    row += 2
    
    # Vendor Summary
    ws[f'A{row}'] = 'Vendor-wise Summary'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    headers = ['Vendor Name', 'GSTIN', 'Bills', 'Taxable Value', 'CGST', 'SGST', 'IGST', 'Total']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row += 1
    for vendor in vendor_list:
        ws.cell(row, 1, vendor['vendor_name'])
        ws.cell(row, 2, vendor['vendor_gstin'])
        ws.cell(row, 3, vendor['bill_count'])
        ws.cell(row, 4, float(vendor['total_taxable_value'])).number_format = '#,##0.00'
        ws.cell(row, 5, float(vendor['total_cgst'])).number_format = '#,##0.00'
        ws.cell(row, 6, float(vendor['total_sgst'])).number_format = '#,##0.00'
        ws.cell(row, 7, float(vendor['total_igst'])).number_format = '#,##0.00'
        ws.cell(row, 8, float(vendor['total_amount'])).number_format = '#,##0.00'
        
        for col in range(1, 9):
            ws.cell(row, col).border = border
        
        row += 1
    
    # Auto-adjust column widths
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = f'GSTR2_Report_{start_date_str}_to_{end_date_str}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@purchase_bills_bp.route('/api/vendors/search')
def api_search_vendors():
    """API endpoint to search vendors"""
    tenant_id = get_current_tenant_id()
    query = request.args.get('q', '').strip()
    
    if len(query) < 2:
        return jsonify([])
    
    vendors = Vendor.query.filter_by(tenant_id=tenant_id, is_active=True).filter(
        db.or_(
            Vendor.name.like(f'%{query}%'),
            Vendor.vendor_code.like(f'%{query}%'),
            Vendor.company_name.like(f'%{query}%')
        )
    ).limit(20).all()
    
    results = []
    for vendor in vendors:
        results.append({
            'id': vendor.id,
            'name': vendor.name,
            'company_name': vendor.company_name or '',
            'phone': vendor.phone or '',
            'email': vendor.email or '',
            'gstin': vendor.gstin or '',
            'address': vendor.address or '',
            'state': vendor.state or 'Maharashtra',
            'payment_terms_days': vendor.payment_terms_days or 30
        })
    
    return jsonify(results)

