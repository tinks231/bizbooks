"""
Items management routes (Professional inventory - Zoho style)
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, g, jsonify, send_file
from models import db, Item, ItemCategory, ItemGroup, ItemStock, ItemStockMovement, Site, InventoryAdjustment, InventoryAdjustmentLine
from utils.tenant_middleware import require_tenant, get_current_tenant_id
from utils.license_check import check_license
from functools import wraps
from datetime import datetime
from sqlalchemy import func
import pytz
import os

# Create blueprint
items_bp = Blueprint('items', __name__, url_prefix='/admin/items')

def login_required(f):
    """Decorator to require admin login (also checks license)"""
    @wraps(f)
    @check_license  # Check license/trial before allowing access
    def decorated_function(*args, **kwargs):
        from flask import session
        if 'tenant_admin_id' not in session:
            flash('Please login first', 'error')
            return redirect(url_for('admin.login'))
        # Verify session tenant matches current tenant
        if session.get('tenant_admin_id') != get_current_tenant_id():
            session.clear()
            flash('Session mismatch. Please login again.', 'error')
            return redirect(url_for('admin.login'))
        return f(*args, **kwargs)
    return decorated_function


def generate_sku(tenant_id):
    """Auto-generate SKU for new items - finds highest ITEM-#### number for THIS tenant"""
    # SKU must be unique PER TENANT (not globally)
    # Database has UNIQUE(tenant_id, sku) constraint
    # Each tenant has their own sequence: Tenant A (ITEM-0001, 0002...), Tenant B (ITEM-0001, 0002...)
    
    # Get all items with ITEM-#### pattern SKUs for THIS TENANT ONLY
    items = Item.query.filter_by(tenant_id=tenant_id).filter(
        Item.sku.like('ITEM-%')
    ).all()
    
    # Find the highest number in use FOR THIS TENANT
    max_number = 0
    for item in items:
        try:
            # Extract number from SKU (e.g., ITEM-0012 -> 12)
            if item.sku and '-' in item.sku:
                number_part = item.sku.split('-')[1]
                number = int(number_part)
                if number > max_number:
                    max_number = number
        except (ValueError, IndexError):
            # Skip malformed SKUs
            continue
    
    # Generate next number
    new_number = max_number + 1
    
    # Format as ITEM-0001, ITEM-0002, etc.
    new_sku = f"ITEM-{new_number:04d}"
    
    # Double-check it doesn't exist for THIS TENANT (safety check for race conditions)
    while Item.query.filter_by(tenant_id=tenant_id, sku=new_sku).first():
        new_number += 1
        new_sku = f"ITEM-{new_number:04d}"
    
    return new_sku


# ===== API: ITEM SEARCH =====
@items_bp.route('/api/search', methods=['GET'])
@require_tenant
@login_required
def api_search_items():
    """Fast API endpoint for searching items (for invoice/bill creation)"""
    from sqlalchemy import or_
    
    tenant_id = get_current_tenant_id()
    query_text = request.args.get('q', '').strip()
    limit = request.args.get('limit', 20, type=int)
    
    if not query_text:
        return jsonify([])
    
    # Search by name, SKU, or barcode (fast indexed query)
    items = Item.query.filter(
        Item.tenant_id == tenant_id,
        Item.is_active == True,
        or_(
            Item.name.ilike(f'%{query_text}%'),
            Item.sku.ilike(f'%{query_text}%'),
            Item.barcode.ilike(f'%{query_text}%')
        )
    ).limit(limit).all()
    
    # Build JSON response
    results = []
    for item in items:
        # Get total stock (optimized - only for returned items)
        total_stock = item.get_total_stock() if item.track_inventory else None
        
        results.append({
            'id': item.id,
            'name': item.name,
            'sku': item.sku,
            'barcode': item.barcode,
            'mrp': float(item.mrp) if item.mrp else None,
            'discount_percent': float(item.discount_percent) if item.discount_percent else 0,
            'selling_price': float(item.selling_price) if item.selling_price else 0,
            'cost_price': float(item.cost_price) if item.cost_price else 0,
            'gst_rate': float(item.gst_rate) if item.gst_rate else 18,
            'hsn_code': item.hsn_code or '',
            'unit': item.unit or 'nos',
            'track_inventory': item.track_inventory,
            'stock': total_stock if total_stock is not None else 'N/A',
            'is_low_stock': item.is_low_stock() if item.track_inventory else False,
            'category': item.category.name if item.category else None,
            'brand': item.brand,
            'manufacturer': item.manufacturer
        })
    
    return jsonify(results)


# ===== ITEMS LISTING =====
@items_bp.route('/', strict_slashes=False)
@require_tenant
@login_required
def index():
    """List all items with OPTIMIZED QUERIES"""
    from sqlalchemy.orm import joinedload
    from sqlalchemy import func
    
    tenant_id = get_current_tenant_id()
    
    # Get filter parameters
    category_id = request.args.get('category', type=int)
    group_id = request.args.get('group', type=int)
    search = request.args.get('search', '')
    status = request.args.get('status', 'active')  # 'active', 'inactive', 'all'
    low_stock_filter = request.args.get('low_stock', type=int)  # 1 = show only low stock
    page = request.args.get('page', 1, type=int)
    per_page = 50  # Show 50 items per page
    
    # Build query with EAGER LOADING (fixes N+1 problem!)
    query = Item.query.options(joinedload(Item.stocks)).filter_by(tenant_id=tenant_id)
    
    # Apply filters
    if category_id:
        query = query.filter_by(category_id=category_id)
    if group_id:
        query = query.filter_by(item_group_id=group_id)
    if search:
        query = query.filter(Item.name.ilike(f'%{search}%') | Item.sku.ilike(f'%{search}%'))
    if status == 'active':
        query = query.filter_by(is_active=True)
    elif status == 'inactive':
        query = query.filter_by(is_active=False)
    
    # Order by created date
    query = query.order_by(Item.created_at.desc())
    
    # Apply pagination
    if low_stock_filter == 1:
        # For low stock filter, we need to get all items first then filter
        # (Can't filter in SQL as it requires aggregation across sites)
        all_items = query.all()
        items_filtered = [item for item in all_items if item.track_inventory and item.reorder_point and item.get_total_stock() < item.reorder_point]
        
        # Manual pagination for filtered items
        total_items = len(items_filtered)
        start = (page - 1) * per_page
        end = start + per_page
        items = items_filtered[start:end]
        total_pages = (total_items + per_page - 1) // per_page
    else:
        # Regular pagination via SQLAlchemy
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        items = pagination.items
        total_pages = pagination.pages
        total_items = pagination.total
    
    # Get categories and groups for filters
    categories = ItemCategory.query.filter_by(tenant_id=tenant_id).all()
    groups = ItemGroup.query.filter_by(tenant_id=tenant_id).all()
    
    # OPTIMIZED: Calculate low stock count using SQL aggregation (ONE query!)
    items_with_stock_query = db.session.query(
        Item.id,
        Item.reorder_point,
        func.sum(ItemStock.quantity_available).label('total_stock')
    ).join(
        ItemStock, Item.id == ItemStock.item_id, isouter=True
    ).filter(
        Item.tenant_id == tenant_id,
        Item.is_active == True,
        Item.track_inventory == True,
        Item.reorder_point.isnot(None)
    ).group_by(Item.id, Item.reorder_point).all()
    
    # Count items where total_stock < reorder_point
    low_stock_count = sum(1 for item in items_with_stock_query if (item.total_stock or 0) < item.reorder_point)
    
    return render_template('admin/items/list.html',
                         items=items,
                         categories=categories,
                         groups=groups,
                         low_stock_count=low_stock_count,
                         page=page,
                         total_pages=total_pages,
                         total_items=total_items if not low_stock_filter else len(items_filtered) if low_stock_filter == 1 else total_items,
                         tenant=g.tenant)


# ===== ADD NEW ITEM =====
@items_bp.route('/add', methods=['GET', 'POST'])
@require_tenant
@login_required
def add():
    """Add new item (like Zoho form)"""
    tenant_id = get_current_tenant_id()
    
    if request.method == 'POST':
        try:
            # Check if manual SKU provided, otherwise auto-generate
            manual_sku = request.form.get('sku', '').strip()
            if manual_sku:
                # Use manual SKU (tenant provided)
                sku = manual_sku
            else:
                # Auto-generate SKU
                sku = generate_sku(tenant_id)
            
            # Create new item
            item = Item(
                tenant_id=tenant_id,
                sku=sku,
                name=request.form.get('name'),
                type=request.form.get('type', 'goods'),
                unit=request.form.get('unit', 'nos'),
                
                # Category & Group
                category_id=request.form.get('category_id') or None,
                item_group_id=request.form.get('item_group_id') or None,
                
                # Dimensions
                dimensions_length=float(request.form.get('dimensions_length') or 0),
                dimensions_width=float(request.form.get('dimensions_width') or 0),
                dimensions_height=float(request.form.get('dimensions_height') or 0),
                dimensions_unit=request.form.get('dimensions_unit', 'cm'),
                
                # Weight
                weight=float(request.form.get('weight') or 0),
                weight_unit=request.form.get('weight_unit', 'kg'),
                
                # Product Identifiers
                barcode=request.form.get('barcode', '').strip() or None,
                manufacturer=request.form.get('manufacturer'),
                brand=request.form.get('brand'),
                upc=request.form.get('upc'),
                ean=request.form.get('ean'),
                mpn=request.form.get('mpn'),
                isbn=request.form.get('isbn'),
                hsn_code=request.form.get('hsn_code'),
                
                # Sales Information
                mrp=float(request.form.get('mrp') or 0) if request.form.get('mrp') else None,
                discount_percent=float(request.form.get('discount_percent') or 0),
                selling_price=float(request.form.get('selling_price') or 0),
                sales_description=request.form.get('sales_description'),
                sales_account=request.form.get('sales_account', 'Sales'),
                gst_rate=float(request.form.get('gst_rate') or 18),
                
                # Purchase Information
                cost_price=float(request.form.get('cost_price') or 0),
                purchase_description=request.form.get('purchase_description'),
                purchase_account=request.form.get('purchase_account', 'Cost of Goods Sold'),
                preferred_vendor=request.form.get('preferred_vendor'),
                
                # Inventory Tracking
                track_inventory='track_inventory' in request.form,
                opening_stock=float(request.form.get('opening_stock') or 0),
                opening_stock_value=float(request.form.get('opening_stock_value') or 0),
                reorder_point=float(request.form.get('reorder_point') or 0),
                
                # Flags
                is_returnable='is_returnable' in request.form,
                is_active=True,
                
                # Metadata
                created_by=request.form.get('created_by', 'Admin')
            )
            
            # Save dynamic attributes (Phase 3)
            from models.item_attribute import ItemAttribute, TenantAttributeConfig
            config = TenantAttributeConfig.query.filter_by(tenant_id=tenant_id).first()
            if config and config.is_enabled:
                attributes = ItemAttribute.query.filter_by(
                    tenant_id=tenant_id,
                    is_active=True
                ).all()
                
                attribute_data = {}
                for attr in attributes:
                    field_name = f'attr_{attr.id}'
                    value = request.form.get(field_name, '').strip()
                    if value:  # Only save non-empty values
                        attribute_data[attr.attribute_name] = value
                
                if attribute_data:
                    item.attribute_data = attribute_data
            
            db.session.add(item)
            db.session.flush()  # Get the item ID
            
            # If tracking inventory and has opening stock, create stock records for all sites
            if item.track_inventory and item.opening_stock > 0:
                from models.site import Site
                sites = Site.query.filter_by(tenant_id=tenant_id, active=True).all()
                
                # Get default site (marked as is_default=True)
                default_site = Site.query.filter_by(
                    tenant_id=tenant_id,
                    is_default=True,
                    active=True
                ).first()
                
                # Fallback to first active site if no default is set
                if not default_site and sites:
                    default_site = sites[0]
                
                for site in sites:
                    # Create stock record - put opening stock only in default site
                    is_default = (site.id == default_site.id) if default_site else (site == sites[0])
                    
                    stock = ItemStock(
                        tenant_id=tenant_id,
                        item_id=item.id,
                        site_id=site.id,
                        quantity_available=item.opening_stock if is_default else 0,
                        stock_value=item.opening_stock_value if is_default else 0
                    )
                    db.session.add(stock)
                    
                    # Create opening stock movement only for default site
                    if is_default:
                        movement = ItemStockMovement(
                            tenant_id=tenant_id,
                            item_id=item.id,
                            site_id=site.id,
                            movement_type='opening_stock',
                            quantity=item.opening_stock,
                            unit_cost=item.cost_price,
                            total_value=item.opening_stock_value,
                            reason='Opening stock',
                            created_by='System'
                        )
                        db.session.add(movement)
            
            db.session.commit()
            
            flash(f'✅ Item "{item.name}" added successfully! SKU: {item.sku}', 'success')
            return redirect(url_for('items.index'))
            
        except Exception as e:
            db.session.rollback()
            
            # Check if it's a duplicate SKU error
            error_str = str(e)
            if 'duplicate key' in error_str.lower() and 'sku' in error_str.lower():
                # Extract SKU from error if possible
                try:
                    sku_value = request.form.get('sku', '').strip() or sku
                except:
                    sku_value = 'unknown'
                
                flash(f'❌ Duplicate SKU Error: An item with SKU "{sku_value}" already exists!', 'error')
                flash('💡 Solution: Either use a different SKU or leave it blank to auto-generate', 'info')
            else:
                # Generic error
                flash(f'❌ Error adding item: {str(e)}', 'error')
    
    # GET request - show form
    categories = ItemCategory.query.filter_by(tenant_id=tenant_id).all()
    groups = ItemGroup.query.filter_by(tenant_id=tenant_id).all()
    
    # Get configured item attributes (Phase 3)
    from models.item_attribute import ItemAttribute, TenantAttributeConfig
    config = TenantAttributeConfig.query.filter_by(tenant_id=tenant_id).first()
    attributes = []
    if config and config.is_enabled:
        attr_objects = ItemAttribute.query.filter_by(
            tenant_id=tenant_id,
            is_active=True
        ).order_by(ItemAttribute.display_order).all()
        
        # Convert to dictionaries for JSON serialization
        attributes = [{
            'id': attr.id,
            'attribute_name': attr.attribute_name,
            'attribute_type': attr.attribute_type,
            'is_required': attr.is_required,
            'dropdown_options': attr.dropdown_options or [],
            'include_in_item_name': attr.include_in_item_name,
            'display_order': attr.display_order
        } for attr in attr_objects]
    
    return render_template('admin/items/add.html',
                         categories=categories,
                         groups=groups,
                         tenant=g.tenant,
                         attributes=attributes)


# ===== EDIT ITEM =====
@items_bp.route('/edit/<int:item_id>', methods=['GET', 'POST'])
@require_tenant
@login_required
def edit(item_id):
    """Edit existing item"""
    tenant_id = get_current_tenant_id()
    item = Item.query.filter_by(id=item_id, tenant_id=tenant_id).first_or_404()
    
    if request.method == 'POST':
        try:
            # Update item fields
            item.name = request.form.get('name')
            item.type = request.form.get('type', 'goods')
            item.unit = request.form.get('unit', 'nos')
            
            # Category & Group
            item.category_id = request.form.get('category_id') or None
            item.item_group_id = request.form.get('item_group_id') or None
            
            # Dimensions
            item.dimensions_length = float(request.form.get('dimensions_length') or 0)
            item.dimensions_width = float(request.form.get('dimensions_width') or 0)
            item.dimensions_height = float(request.form.get('dimensions_height') or 0)
            item.dimensions_unit = request.form.get('dimensions_unit', 'cm')
            
            # Weight
            item.weight = float(request.form.get('weight') or 0)
            item.weight_unit = request.form.get('weight_unit', 'kg')
            
            # Product Identifiers
            item.barcode = request.form.get('barcode', '').strip() or None
            item.manufacturer = request.form.get('manufacturer')
            item.brand = request.form.get('brand')
            item.upc = request.form.get('upc')
            item.ean = request.form.get('ean')
            item.mpn = request.form.get('mpn')
            item.isbn = request.form.get('isbn')
            item.hsn_code = request.form.get('hsn_code')
            
            # Sales Information
            item.mrp = float(request.form.get('mrp') or 0) if request.form.get('mrp') else None
            item.discount_percent = float(request.form.get('discount_percent') or 0)
            item.selling_price = float(request.form.get('selling_price') or 0)
            item.sales_description = request.form.get('sales_description')
            item.sales_account = request.form.get('sales_account', 'Sales')
            item.gst_rate = float(request.form.get('gst_rate') or 18)
            
            # Purchase Information
            item.cost_price = float(request.form.get('cost_price') or 0)
            item.purchase_description = request.form.get('purchase_description')
            item.purchase_account = request.form.get('purchase_account', 'Cost of Goods Sold')
            item.preferred_vendor = request.form.get('preferred_vendor')
            
            # Inventory Tracking
            item.track_inventory = 'track_inventory' in request.form
            item.reorder_point = float(request.form.get('reorder_point') or 0)
            
            # Flags
            item.is_returnable = 'is_returnable' in request.form
            item.is_active = 'is_active' in request.form
            
            db.session.commit()
            
            flash(f'✅ Item "{item.name}" updated successfully!', 'success')
            return redirect(url_for('items.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'❌ Error updating item: {str(e)}', 'error')
    
    # GET request - show form with current values
    categories = ItemCategory.query.filter_by(tenant_id=tenant_id).all()
    groups = ItemGroup.query.filter_by(tenant_id=tenant_id).all()
    
    return render_template('admin/items/edit.html',
                         item=item,
                         categories=categories,
                         groups=groups,
                         tenant=g.tenant)


# ===== DELETE ITEM =====
@items_bp.route('/delete/<int:item_id>', methods=['POST'])
@require_tenant
@login_required
def delete(item_id):
    """Delete item"""
    tenant_id = get_current_tenant_id()
    item = Item.query.filter_by(id=item_id, tenant_id=tenant_id).first_or_404()
    
    try:
        item_name = item.name
        db.session.delete(item)  # Cascade will delete related records
        db.session.commit()
        
        flash(f'✅ Item "{item_name}" deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error deleting item: {str(e)}', 'error')
    
    return redirect(url_for('items.index'))


# ===== ITEM CATEGORIES =====
@items_bp.route('/categories')
@require_tenant
@login_required
def categories():
    """Manage item categories"""
    tenant_id = get_current_tenant_id()
    categories = ItemCategory.query.filter_by(tenant_id=tenant_id).all()
    
    return render_template('admin/items/categories.html',
                         categories=categories,
                         tenant=g.tenant)


@items_bp.route('/categories/add', methods=['POST'])
@require_tenant
@login_required
def add_category():
    """Add new category"""
    tenant_id = get_current_tenant_id()
    
    try:
        category = ItemCategory(
            tenant_id=tenant_id,
            name=request.form.get('name'),
            description=request.form.get('description')
        )
        db.session.add(category)
        db.session.commit()
        
        flash(f'✅ Category "{category.name}" added successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error adding category: {str(e)}', 'error')
    
    return redirect(url_for('items.categories'))


@items_bp.route('/categories/delete/<int:category_id>', methods=['POST'])
@require_tenant
@login_required
def delete_category(category_id):
    """Delete category"""
    tenant_id = get_current_tenant_id()
    category = ItemCategory.query.filter_by(id=category_id, tenant_id=tenant_id).first_or_404()
    
    try:
        category_name = category.name
        db.session.delete(category)
        db.session.commit()
        
        flash(f'✅ Category "{category_name}" deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error deleting category: {str(e)}', 'error')
    
    return redirect(url_for('items.categories'))


# ===== ITEM GROUPS =====
@items_bp.route('/groups')
@require_tenant
@login_required
def groups():
    """Manage item groups"""
    tenant_id = get_current_tenant_id()
    groups = ItemGroup.query.filter_by(tenant_id=tenant_id).all()
    
    return render_template('admin/items/groups.html',
                         groups=groups,
                         tenant=g.tenant)


@items_bp.route('/groups/add', methods=['POST'])
@require_tenant
@login_required
def add_group():
    """Add new group"""
    tenant_id = get_current_tenant_id()
    
    try:
        group = ItemGroup(
            tenant_id=tenant_id,
            name=request.form.get('name'),
            description=request.form.get('description')
        )
        db.session.add(group)
        db.session.commit()
        
        flash(f'✅ Group "{group.name}" added successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error adding group: {str(e)}', 'error')
    
    return redirect(url_for('items.groups'))


@items_bp.route('/groups/delete/<int:group_id>', methods=['POST'])
@require_tenant
@login_required
def delete_group(group_id):
    """Delete group"""
    tenant_id = get_current_tenant_id()
    group = ItemGroup.query.filter_by(id=group_id, tenant_id=tenant_id).first_or_404()
    
    try:
        group_name = group.name
        db.session.delete(group)
        db.session.commit()
        
        flash(f'✅ Group "{group_name}" deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error deleting group: {str(e)}', 'error')
    
    return redirect(url_for('items.groups'))


# ===== JSON APIs for Modal Dropdowns =====
@items_bp.route('/categories/json', methods=['GET'])
@require_tenant
@login_required
def get_categories_json():
    """Get categories as JSON (for refreshing dropdown after modal add)"""
    tenant_id = get_current_tenant_id()
    categories = ItemCategory.query.filter_by(tenant_id=tenant_id).order_by(ItemCategory.name).all()
    
    return jsonify({
        'categories': [{'id': cat.id, 'name': cat.name} for cat in categories]
    })


@items_bp.route('/groups/json', methods=['GET'])
@require_tenant
@login_required
def get_groups_json():
    """Get groups as JSON (for refreshing dropdown after modal add)"""
    tenant_id = get_current_tenant_id()
    groups = ItemGroup.query.filter_by(tenant_id=tenant_id).order_by(ItemGroup.name).all()
    
    return jsonify({
        'groups': [{'id': grp.id, 'name': grp.name} for grp in groups]
    })


# ===== STOCK SUMMARY =====
@items_bp.route('/stock-summary')
@require_tenant
@login_required
def stock_summary():
    """Show current stock levels for all items across all sites - HYPER-OPTIMIZED"""
    from sqlalchemy.orm import joinedload
    
    tenant_id = get_current_tenant_id()
    
    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    # Get all sites (usually just 2-3 sites, so this is fast)
    sites = Site.query.filter_by(tenant_id=tenant_id).all()
    
    # HYPER-OPTIMIZATION: Paginate FIRST, then load stock for ONLY those items
    items_query = Item.query.filter_by(tenant_id=tenant_id, is_active=True).order_by(Item.name)
    pagination = items_query.paginate(page=page, per_page=per_page, error_out=False)
    items = pagination.items
    
    # Get item IDs for this page only
    item_ids = [item.id for item in items]
    
    # OPTIMIZATION: Load stock ONLY for paginated items (not all items!)
    # Before: 500 items × 2 sites = 1000 stock records
    # After: 50 items × 2 sites = 100 stock records (10x less!)
    stock_records = ItemStock.query.filter(
        ItemStock.tenant_id == tenant_id,
        ItemStock.item_id.in_(item_ids)
    ).all()
    
    # Build a lookup dictionary: {(item_id, site_id): quantity}
    stock_lookup = {}
    for stock in stock_records:
        stock_lookup[(stock.item_id, stock.site_id)] = stock.quantity_available
    
    # Build stock summary using lookup (no queries in loop!)
    stock_data = []
    for item in items:
        # Calculate total stock from lookup (no query!)
        total_stock = sum(
            stock_lookup.get((item.id, site.id), 0) for site in sites
        )
        
        # Get stock by site from lookup (no query!)
        stock_by_site = {
            site.id: stock_lookup.get((item.id, site.id), 0) for site in sites
        }
        
        stock_data.append({
            'item': item,
            'total_stock': total_stock,
            'stock_by_site': stock_by_site,
            'low_stock': total_stock < item.reorder_point if item.reorder_point else False
        })
    
    return render_template('admin/items/stock_summary.html',
                         stock_data=stock_data,
                         sites=sites,
                         page=page,
                         total_pages=pagination.pages,
                         total_items=pagination.total,
                         tenant=g.tenant)


# ===== ADJUSTMENTS =====
@items_bp.route('/adjustments')
@require_tenant
@login_required
def adjustments():
    """List all inventory adjustments"""
    tenant_id = get_current_tenant_id()
    
    # Get all adjustments
    adjustments = InventoryAdjustment.query.filter_by(
        tenant_id=tenant_id
    ).order_by(InventoryAdjustment.adjustment_date.desc()).all()
    
    # Get items and sites for the form
    items = Item.query.filter_by(tenant_id=tenant_id, is_active=True).order_by(Item.name).all()
    sites = Site.query.filter_by(tenant_id=tenant_id).all()
    
    return render_template('admin/items/adjustments.html',
                         adjustments=adjustments,
                         items=items,
                         sites=sites,
                         tenant=g.tenant)


@items_bp.route('/adjustments/add', methods=['POST'])
@require_tenant
@login_required
def add_adjustment():
    """Create new inventory adjustment"""
    tenant_id = get_current_tenant_id()
    ist = pytz.timezone('Asia/Kolkata')
    
    try:
        # Create adjustment record
        adjustment = InventoryAdjustment(
            tenant_id=tenant_id,
            adjustment_date=datetime.now(ist),
            reason=request.form.get('reason'),
            notes=request.form.get('notes')
        )
        db.session.add(adjustment)
        db.session.flush()  # Get adjustment ID
        
        # Process each line item
        item_ids = request.form.getlist('item_id[]')
        site_ids = request.form.getlist('site_id[]')
        quantities = request.form.getlist('quantity[]')
        adjustment_types = request.form.getlist('adjustment_type[]')
        
        for item_id, site_id, qty, adj_type in zip(item_ids, site_ids, quantities, adjustment_types):
            if not item_id or not site_id or not qty:
                continue
            
            qty = int(qty)
            if qty == 0:
                continue
            
            # Create adjustment line
            line = InventoryAdjustmentLine(
                adjustment_id=adjustment.id,
                item_id=int(item_id),
                site_id=int(site_id),
                quantity_change=qty if adj_type == 'add' else -qty,
                adjustment_type=adj_type
            )
            db.session.add(line)
            
            # Update item stock
            stock = ItemStock.query.filter_by(
                item_id=int(item_id),
                site_id=int(site_id)
            ).first()
            
            if stock:
                stock.quantity_available += qty if adj_type == 'add' else -qty
            else:
                # Create new stock record
                stock = ItemStock(
                    tenant_id=tenant_id,
                    item_id=int(item_id),
                    site_id=int(site_id),
                    quantity_available=qty if adj_type == 'add' else 0
                )
                db.session.add(stock)
            
            # Create stock movement record
            movement = ItemStockMovement(
                tenant_id=tenant_id,
                item_id=int(item_id),
                site_id=int(site_id),
                quantity=qty if adj_type == 'add' else -qty,
                movement_type='adjustment_in' if adj_type == 'add' else 'adjustment_out',
                reference_type='adjustment',
                reference_id=adjustment.id,
                notes=f"Adjustment: {request.form.get('reason')}"
            )
            db.session.add(movement)
        
        db.session.commit()
        flash(f'✅ Inventory adjustment recorded successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error recording adjustment: {str(e)}', 'error')
    
    return redirect(url_for('items.adjustments'))


# ===== TRANSFERS =====
@items_bp.route('/transfers')
@require_tenant
@login_required
def transfers():
    """List all stock transfers"""
    tenant_id = get_current_tenant_id()
    
    # Get all transfers (movements with type 'transfer')
    transfers = ItemStockMovement.query.join(Item).filter(
        Item.tenant_id == tenant_id,
        ItemStockMovement.movement_type.in_(['transfer_out', 'transfer_in'])
    ).order_by(ItemStockMovement.created_at.desc()).all()
    
    # Get items and sites for the form
    items = Item.query.filter_by(tenant_id=tenant_id, is_active=True).order_by(Item.name).all()
    sites = Site.query.filter_by(tenant_id=tenant_id).all()
    
    return render_template('admin/items/transfers.html',
                         transfers=transfers,
                         items=items,
                         sites=sites,
                         tenant=g.tenant)


@items_bp.route('/transfers/add', methods=['POST'])
@require_tenant
@login_required
def add_transfer():
    """Transfer stock between sites"""
    tenant_id = get_current_tenant_id()
    ist = pytz.timezone('Asia/Kolkata')
    
    try:
        item_id = int(request.form.get('item_id'))
        from_site_id = int(request.form.get('from_site'))
        to_site_id = int(request.form.get('to_site'))
        quantity = int(request.form.get('quantity'))
        notes = request.form.get('notes', '')
        
        # Validate
        if from_site_id == to_site_id:
            flash('❌ Cannot transfer to the same site!', 'error')
            return redirect(url_for('items.transfers'))
        
        # Check source stock
        source_stock = ItemStock.query.filter_by(
            item_id=item_id,
            site_id=from_site_id
        ).first()
        
        if not source_stock or source_stock.quantity_available < quantity:
            flash('❌ Insufficient stock at source site!', 'error')
            return redirect(url_for('items.transfers'))
        
        # Deduct from source
        source_stock.quantity_available -= quantity
        
        # Add to destination
        dest_stock = ItemStock.query.filter_by(
            item_id=item_id,
            site_id=to_site_id
        ).first()
        
        if dest_stock:
            dest_stock.quantity_available += quantity
        else:
            dest_stock = ItemStock(
                tenant_id=tenant_id,
                item_id=item_id,
                site_id=to_site_id,
                quantity_available=quantity
            )
            db.session.add(dest_stock)
        
        # Record movements
        transfer_out = ItemStockMovement(
            tenant_id=tenant_id,
            item_id=item_id,
            site_id=from_site_id,
            quantity=-quantity,
            movement_type='transfer_out',
            reference_type='transfer',
            from_site_id=from_site_id,
            to_site_id=to_site_id,
            notes=f"Transfer to site {to_site_id}: {notes}"
        )
        db.session.add(transfer_out)
        
        transfer_in = ItemStockMovement(
            tenant_id=tenant_id,
            item_id=item_id,
            site_id=to_site_id,
            quantity=quantity,
            movement_type='transfer_in',
            reference_type='transfer',
            from_site_id=from_site_id,
            to_site_id=to_site_id,
            notes=f"Transfer from site {from_site_id}: {notes}"
        )
        db.session.add(transfer_in)
        
        db.session.commit()
        flash(f'✅ Stock transferred successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error transferring stock: {str(e)}', 'error')
    
    return redirect(url_for('items.transfers'))


# ===== REPORTS =====
@items_bp.route('/reports')
@require_tenant
@login_required
def reports():
    """Inventory reports - Stock valuation and movement summary"""
    tenant_id = get_current_tenant_id()
    
    # Get all items with stock
    items = Item.query.filter_by(tenant_id=tenant_id, is_active=True, track_inventory=True).all()
    
    # Calculate stock valuation
    total_stock_value = 0
    total_items = 0
    low_stock_items = 0
    out_of_stock_items = 0
    
    stock_details = []
    
    for item in items:
        # Get total stock across all sites
        total_qty = db.session.query(func.sum(ItemStock.quantity_available))\
            .filter_by(tenant_id=tenant_id, item_id=item.id)\
            .scalar() or 0
        
        # Calculate value (quantity * cost price)
        item_value = total_qty * (item.cost_price or 0)
        total_stock_value += item_value
        
        if total_qty > 0:
            total_items += 1
            
            # Check if low stock
            if item.reorder_point and total_qty < item.reorder_point:
                low_stock_items += 1
            
            stock_details.append({
                'item': item,
                'quantity': total_qty,
                'value': item_value,
                'low_stock': item.reorder_point and total_qty < item.reorder_point
            })
        else:
            out_of_stock_items += 1
    
    # Get recent stock movements (last 30 days)
    from datetime import timedelta
    thirty_days_ago = datetime.now(pytz.timezone('Asia/Kolkata')) - timedelta(days=30)
    
    recent_movements = ItemStockMovement.query.join(Item)\
        .filter(Item.tenant_id == tenant_id)\
        .filter(ItemStockMovement.created_at >= thirty_days_ago)\
        .order_by(ItemStockMovement.created_at.desc())\
        .limit(100)\
        .all()
    
    # Stock movements summary
    movements_in = db.session.query(func.sum(ItemStockMovement.quantity))\
        .join(Item)\
        .filter(Item.tenant_id == tenant_id)\
        .filter(ItemStockMovement.quantity > 0)\
        .filter(ItemStockMovement.created_at >= thirty_days_ago)\
        .scalar() or 0
    
    movements_out = abs(db.session.query(func.sum(ItemStockMovement.quantity))\
        .join(Item)\
        .filter(Item.tenant_id == tenant_id)\
        .filter(ItemStockMovement.quantity < 0)\
        .filter(ItemStockMovement.created_at >= thirty_days_ago)\
        .scalar() or 0)
    
    return render_template('admin/items/reports.html',
                         total_stock_value=total_stock_value,
                         total_items=total_items,
                         low_stock_items=low_stock_items,
                         out_of_stock_items=out_of_stock_items,
                         stock_details=stock_details,
                         recent_movements=recent_movements,
                         movements_in=movements_in,
                         movements_out=movements_out,
                         tenant=g.tenant)


# ===== EXPORT ITEMS TO EXCEL =====
@items_bp.route('/export-excel')
@require_tenant
@login_required
def export_excel():
    """Export all items to Excel with barcode column"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from flask import send_file
    
    tenant_id = get_current_tenant_id()
    
    # Get all active items
    items = Item.query.filter_by(tenant_id=tenant_id, is_active=True).order_by(Item.name).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Items Export"
    
    # Headers
    headers = ['Item Name', 'SKU', 'Barcode', 'Category', 'Group', 'Unit', 'Stock Quantity', 
               'MRP', 'Discount %', 'Selling Price', 'Cost Price', 'GST Rate (%)', 'HSN Code', 'Description']
    
    # Style headers
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data rows
    for row_num, item in enumerate(items, start=2):
        total_stock = item.get_total_stock() if item.track_inventory else 0
        
        row_data = [
            item.name,
            item.sku,
            item.barcode or '',  # Barcode (empty if not set)
            item.category.name if item.category else '',
            item.item_group.name if item.item_group else '',
            item.unit or 'nos',
            total_stock,
            item.mrp or 0,
            item.discount_percent or 0,
            item.selling_price or 0,
            item.cost_price or 0,
            item.gst_rate or 18,
            item.hsn_code or '',
            item.sales_description or ''
        ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30  # Item Name
    ws.column_dimensions['B'].width = 15  # SKU
    ws.column_dimensions['C'].width = 18  # Barcode
    ws.column_dimensions['D'].width = 20  # Category
    ws.column_dimensions['E'].width = 20  # Group
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with timestamp
    from datetime import datetime
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'items_export_{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ===== BULK GENERATE BARCODES =====
@items_bp.route('/bulk-generate-barcodes', methods=['POST'])
@require_tenant
@login_required
def bulk_generate_barcodes():
    """Auto-generate barcodes for all items that don't have one"""
    tenant_id = get_current_tenant_id()
    
    try:
        # Get all items without barcodes
        items_without_barcode = Item.query.filter_by(
            tenant_id=tenant_id,
            is_active=True
        ).filter(
            (Item.barcode == None) | (Item.barcode == '')
        ).all()
        
        if not items_without_barcode:
            flash('✅ All items already have barcodes!', 'success')
            return redirect(url_for('items.index'))
        
        # Helper function to calculate EAN-13 check digit
        def calculate_ean13_check_digit(barcode_base):
            if len(barcode_base) != 12:
                return 0
            odd_sum = sum(int(barcode_base[i]) for i in range(0, 12, 2))
            even_sum = sum(int(barcode_base[i]) * 3 for i in range(1, 12, 2))
            total = odd_sum + even_sum
            check_digit = (10 - (total % 10)) % 10
            return check_digit
        
        # Get the highest existing barcode number for this tenant to avoid duplicates
        last_barcode = db.session.query(Item.barcode).filter(
            Item.tenant_id == tenant_id,
            Item.barcode != None,
            Item.barcode.like(f'890{tenant_id:04d}%')
        ).order_by(Item.barcode.desc()).first()
        
        # Start from next sequential number
        if last_barcode and last_barcode[0]:
            try:
                # Extract the sequential part (digits 8-12) and increment
                last_num = int(last_barcode[0][7:12])
                next_num = last_num + 1
            except:
                next_num = 1
        else:
            next_num = 1
        
        count = 0
        for item in items_without_barcode:
            # Generate EAN-13 barcode: 890 (India) + tenant_id (4 digits) + sequential (5 digits) + check digit
            base = f"890{tenant_id:04d}{next_num:05d}"
            check_digit = calculate_ean13_check_digit(base)
            barcode = base + str(check_digit)
            
            # Double-check for duplicates before saving
            existing = Item.query.filter_by(tenant_id=tenant_id, barcode=barcode).first()
            while existing:
                next_num += 1
                base = f"890{tenant_id:04d}{next_num:05d}"
                check_digit = calculate_ean13_check_digit(base)
                barcode = base + str(check_digit)
                existing = Item.query.filter_by(tenant_id=tenant_id, barcode=barcode).first()
            
            item.barcode = barcode
            next_num += 1  # Increment for next item
            count += 1
        
        db.session.commit()
        
        flash(f'✅ Generated {count} barcodes successfully!', 'success')
        return redirect(url_for('items.index'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error generating barcodes: {str(e)}', 'error')
        return redirect(url_for('items.index'))


# ===== PRINT BARCODE LABELS =====
@items_bp.route('/print-labels', methods=['GET', 'POST'])
@require_tenant
@login_required
def print_labels():
    """Generate PDF with barcode labels for printing"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from io import BytesIO
    import barcode
    from barcode.writer import ImageWriter
    from PIL import Image
    
    tenant_id = get_current_tenant_id()
    
    if request.method == 'POST':
        # Get selected items or all items with barcodes
        selected_item_ids = request.form.getlist('item_ids')
        
        if selected_item_ids:
            items = Item.query.filter(
                Item.id.in_(selected_item_ids),
                Item.tenant_id == tenant_id
            ).all()
        else:
            # Get all items with barcodes
            items = Item.query.filter_by(
                tenant_id=tenant_id,
                is_active=True
            ).filter(
                Item.barcode != None,
                Item.barcode != ''
            ).all()
        
        if not items:
            flash('⚠️ No items with barcodes found!', 'warning')
            return redirect(url_for('items.index'))
        
        # Get quantity for each item (default 1)
        labels_to_print = []
        for item in items:
            qty = int(request.form.get(f'qty_{item.id}', 1))
            for _ in range(qty):
                labels_to_print.append(item)
        
        # Generate PDF
        pdf_buffer = BytesIO()
        c = canvas.Canvas(pdf_buffer, pagesize=A4)
        width, height = A4
        
        # Label dimensions for 30 labels per A4 (3 columns x 10 rows)
        # Standard Avery A4-J8160 compatible
        label_width = 70 * mm
        label_height = 29.7 * mm
        margin_left = 5 * mm
        margin_top = 10 * mm
        gap_x = 2.5 * mm
        gap_y = 0 * mm
        
        cols = 3
        rows = 10
        labels_per_page = cols * rows
        
        page_num = 0
        
        for idx, item in enumerate(labels_to_print):
            # Calculate position
            col = idx % cols
            row = (idx // cols) % rows
            
            # New page if needed
            if idx > 0 and idx % labels_per_page == 0:
                c.showPage()
                page_num += 1
            
            # Calculate x, y position (top-left origin)
            x = margin_left + col * (label_width + gap_x)
            y = height - margin_top - (row + 1) * (label_height + gap_y)
            
            # Draw label border (optional - for cutting guide)
            # c.setStrokeColor(colors.lightgrey)
            # c.rect(x, y, label_width, label_height)
            
            # Item name (top)
            c.setFont("Helvetica-Bold", 9)
            name_text = item.name[:35]  # Truncate if too long
            c.drawString(x + 3*mm, y + label_height - 6*mm, name_text)
            
            # Price info (below name)
            c.setFont("Helvetica", 7)
            price_text = f"MRP: ₹{item.mrp or item.selling_price:.0f}  |  Price: ₹{item.selling_price:.0f}"
            c.drawString(x + 3*mm, y + label_height - 11*mm, price_text)
            
            # SKU (small)
            c.setFont("Helvetica", 6)
            c.drawString(x + 3*mm, y + label_height - 15*mm, f"SKU: {item.sku}")
            
            # Generate barcode image
            if item.barcode:
                try:
                    # Determine barcode type based on length
                    if len(item.barcode) == 13:
                        barcode_class = barcode.get_barcode_class('ean13')
                    elif len(item.barcode) == 8:
                        barcode_class = barcode.get_barcode_class('ean8')
                    else:
                        barcode_class = barcode.get_barcode_class('code128')
                    
                    # Generate barcode
                    barcode_instance = barcode_class(item.barcode, writer=ImageWriter())
                    
                    # Save to BytesIO
                    barcode_buffer = BytesIO()
                    barcode_instance.write(barcode_buffer, options={
                        'module_height': 10,
                        'module_width': 0.25,
                        'quiet_zone': 2,
                        'font_size': 8,
                        'text_distance': 2,
                        'write_text': True
                    })
                    barcode_buffer.seek(0)
                    
                    # Draw barcode on PDF
                    barcode_img = Image.open(barcode_buffer)
                    
                    # Calculate barcode position and size
                    barcode_x = x + 3*mm
                    barcode_y = y + 3*mm
                    barcode_width = label_width - 6*mm
                    barcode_height = 12*mm
                    
                    # Draw image
                    c.drawInlineImage(
                        barcode_img,
                        barcode_x,
                        barcode_y,
                        width=barcode_width,
                        height=barcode_height,
                        preserveAspectRatio=True
                    )
                    
                except Exception as e:
                    # Fallback: just print barcode text
                    c.setFont("Courier", 8)
                    c.drawString(x + 3*mm, y + 10*mm, item.barcode)
        
        # Save PDF
        c.save()
        pdf_buffer.seek(0)
        
        # Generate filename
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'barcode_labels_{timestamp}.pdf'
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    
    # GET request - show form to select items
    items_with_barcodes = Item.query.filter_by(
        tenant_id=tenant_id,
        is_active=True
    ).filter(
        Item.barcode != None,
        Item.barcode != ''
    ).order_by(Item.name).all()
    
    return render_template('admin/items/print_labels.html',
                         items=items_with_barcodes,
                         tenant=g.tenant)

