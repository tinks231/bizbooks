"""
Excel Import Utilities for Bulk Data Import
Handles validation, parsing, and importing from Excel/CSV files
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime, timedelta
from models import db, Employee, Item, Customer, Site, ItemCategory, ItemGroup
from models.subscription import SubscriptionPlan, CustomerSubscription, SubscriptionDelivery

def create_employee_template():
    """
    Create Excel template for employee bulk import
    Returns: BytesIO object with Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Import"
    
    # Headers - match actual Employee model fields (including payroll)
    headers = ['Name*', 'PIN*', 'Phone', 'Email', 'Site Name', 'Monthly Salary', 'Designation', 'Date of Joining']
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add sample data (row 2)
    sample_data = [
        'John Doe',
        '1234',
        '9876543210',
        'john@example.com',
        'Main Office',
        '25000',
        'Manager',
        '2025-01-01'
    ]
    
    for col_num, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col_num, value=value)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Add instructions
    ws.cell(row=4, column=1, value="INSTRUCTIONS:")
    ws.cell(row=5, column=1, value="1. Fields marked with * are required (Name, PIN)")
    ws.cell(row=6, column=1, value="2. PIN must be 4 digits")
    ws.cell(row=7, column=1, value="3. Phone should be 10 digits (optional)")
    ws.cell(row=8, column=1, value="4. Email is optional (for purchase request notifications)")
    ws.cell(row=9, column=1, value="5. If Site Name doesn't exist, it will be created")
    ws.cell(row=10, column=1, value="6. Monthly Salary: Enter numbers only (e.g., 25000)")
    ws.cell(row=11, column=1, value="7. Designation: Job title (e.g., Manager, Helper, Sales)")
    ws.cell(row=12, column=1, value="8. Date of Joining: Format YYYY-MM-DD (e.g., 2025-01-15)")
    ws.cell(row=13, column=1, value="9. Delete row 2 (sample data) before uploading")
    ws.cell(row=11, column=1, value="7. You can add as many rows as you need")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def create_inventory_template(tenant_id=None):
    """
    Create Excel template for inventory bulk import
    Now includes dynamic attribute columns based on tenant configuration
    Returns: BytesIO object with Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Import"
    
    # Base headers
    headers = ['Item Name*', 'SKU', 'Barcode']
    
    # Add dynamic attribute columns if tenant has configured attributes
    attribute_columns = []
    if tenant_id:
        from models.item_attribute import ItemAttribute, TenantAttributeConfig
        config = TenantAttributeConfig.query.filter_by(tenant_id=tenant_id).first()
        if config and config.is_enabled:
            attributes = ItemAttribute.query.filter_by(
                tenant_id=tenant_id,
                is_active=True
            ).order_by(ItemAttribute.display_order).all()
            
            for attr in attributes:
                col_name = f"{attr.attribute_name}{'*' if attr.is_required else ''}"
                headers.append(col_name)
                attribute_columns.append({
                    'name': attr.attribute_name,
                    'required': attr.is_required,
                    'type': attr.attribute_type,
                    'options': attr.dropdown_options
                })
    
    # Continue with standard columns
    headers.extend(['Category*', 'Group*', 'Unit*', 'Stock Quantity*', 
                   'Reorder Point', 'Cost Price*', 'MRP', 'Discount %', 'Selling Price*', 
                   'Tax Rate (%)', 'HSN Code', 'Description'])
    
    # Style headers
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    attr_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        # Use purple for attribute columns
        is_attr_col = any(attr['name'] in header for attr in attribute_columns)
        cell.fill = attr_fill if is_attr_col else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add sample data (row 2) - Updated with attributes if applicable
    base_sample = [
        None,  # Item Name - Will be auto-generated by formula!
        'TSH-MEN-WHT-001',              # SKU
        '8901234567001',                 # Barcode
    ]
    
    # Track attribute column positions for formula
    attr_start_col = 4  # Column D (after Item Name, SKU, Barcode)
    attr_col_letters = []
    
    # Add sample attribute values
    for idx, attr in enumerate(attribute_columns):
        col_letter = chr(ord('A') + attr_start_col - 1 + idx)
        attr_col_letters.append(col_letter)
        
        if attr['options']:  # Dropdown
            base_sample.append(attr['options'][0] if attr['options'] else '')
        else:  # Text/Number/Date
            base_sample.append('Sample Value')
    
    # Continue with standard columns
    base_sample.extend([
        "Men's Wear",                    # Category
        'T-Shirts',                      # Group
        'Pcs',                          # Unit
        50,                             # Stock Quantity
        10,                             # Reorder Point
        450,                            # Cost Price
        699,                            # MRP
        15,                             # Discount %
        None,                           # Selling Price (formula)
        12,                             # Tax Rate
        '6109',                         # HSN Code
        'Premium cotton t-shirt'        # Description
    ])
    
    for col_num, value in enumerate(base_sample, 1):
        cell = ws.cell(row=2, column=col_num, value=value)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # ✨ AUTO-GENERATE ITEM NAME FROM ATTRIBUTES (Excel Formula)
    if attribute_columns and attr_col_letters:
        # Build formula: =TRIM(D2&" "&E2&" "&F2&...)
        # Apply to rows 2-1000 so it auto-copies
        separator = ' & " " & '
        
        for row_num in range(2, 101):  # Apply to first 100 rows (sample + data)
            formula_parts = [f"{col}{row_num}" for col in attr_col_letters]
            formula = f'=TRIM({separator.join(formula_parts)})'
            
            item_name_cell = ws.cell(row=row_num, column=1)
            item_name_cell.value = formula
            if row_num == 2:  # Highlight sample row
                item_name_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                item_name_cell.font = Font(italic=True, color="666666")
    
    # ✨ ADD DATA VALIDATION (Dropdowns)
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # 1. Unit dropdown validation (ALL units from system)
    unit_col_idx = 3 + len(attribute_columns) + 3  # After attrs, then Category, Group, Unit
    unit_col_letter = chr(ord('A') + unit_col_idx - 1)
    
    # Complete list of units from item form
    unit_options = '"Nos,Pcs,Dozen,Kg,Qtl,Ton,Gram,Liter,Ml,Meter,Cm,Ft,Sqm,Cum,Cft,Box,Bag,Packet,Bundle,Pair,Set"'
    unit_dv = DataValidation(type="list", formula1=unit_options, allow_blank=False)
    unit_dv.error = 'Invalid Unit! Choose from dropdown or type: Nos, Pcs, Dozen, Kg, Qtl, Ton, Gram, Liter, Ml, Meter, Cm, Ft, Sqm, Cum, Cft, Box, Bag, Packet, Bundle, Pair, Set'
    unit_dv.errorTitle = 'Invalid Unit'
    unit_dv.promptTitle = 'Select Unit'
    unit_dv.prompt = 'Choose a unit from the dropdown'
    ws.add_data_validation(unit_dv)
    unit_dv.add(f'{unit_col_letter}2:{unit_col_letter}1000')
    
    # 2. Attribute dropdowns (SOFT - only for attributes with options)
    for idx, attr in enumerate(attribute_columns):
        if attr['options']:  # Only add dropdown if options are configured
            col_letter = attr_col_letters[idx]
            options_str = ','.join(attr['options'])
            
            # SOFT DROPDOWN: Shows warning but allows free text
            attr_dv = DataValidation(
                type="list",
                formula1=f'"{options_str}"',
                allow_blank=True,  # Allow empty
                showErrorMessage=False  # ✨ SOFT: Show warning, but allow anyway
            )
            attr_dv.promptTitle = f'Select {attr["name"]}'
            attr_dv.prompt = f'Suggested: {", ".join(attr["options"][:5])}{"..." if len(attr["options"]) > 5 else ""}\n\nYou can also type a new value.'
            ws.add_data_validation(attr_dv)
            attr_dv.add(f'{col_letter}2:{col_letter}1000')
    
    # Add instructions with attribute info
    instructions_start = 4
    ws.cell(row=instructions_start, column=1, value="INSTRUCTIONS:")
    row_idx = instructions_start + 1
    
    if attribute_columns:
        ws.cell(row=row_idx, column=1, value=f"🎨 {len(attribute_columns)} ATTRIBUTE COLUMNS CONFIGURED:")
        ws.cell(row=row_idx, column=1).font = Font(bold=True, color="667eea")
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="   ✨ Item Name will auto-generate from these attributes!")
        ws.cell(row=row_idx, column=1).font = Font(italic=True, color="667eea")
        row_idx += 1
        for attr in attribute_columns:
            required_mark = " (Required)" if attr['required'] else " (Optional)"
            if attr['options']:
                options_str = ", ".join(attr['options'][:5])
                if len(attr['options']) > 5:
                    options_str += "..."
                ws.cell(row=row_idx, column=1, value=f"   - {attr['name']}{required_mark}: Options: {options_str}")
            else:
                ws.cell(row=row_idx, column=1, value=f"   - {attr['name']}{required_mark}: {attr['type'].title()}")
            row_idx += 1
        row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="1. Item Name: AUTO-FILLED from attributes (DO NOT EDIT)")
    ws.cell(row=row_idx, column=1).font = Font(bold=True)
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="2. Unit: Choose from dropdown (Pcs, Kg, Liter, etc.)")
    ws.cell(row=row_idx, column=1).font = Font(bold=True)
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="3. Attribute columns have dropdowns - select from list")
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="4. If SKU is blank, it will be auto-generated")
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="5. If Category/Group doesn't exist, it will be created")
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="6. Fill Selling Price (required)")
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="7. Delete row 2 (sample data) before uploading")
    
    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        # Set reasonable default width based on header length
        ws.column_dimensions[col_letter].width = max(12, len(str(header)) + 2)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def create_customer_template():
    """
    Create Excel template for customer bulk import
    Returns: BytesIO object with Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Import"
    
    # Headers - match Customer model fields (including loyalty program fields)
    headers = ['Customer Name*', 'Phone*', 'Email', 'GSTIN', 'Address', 'State', 'Credit Limit', 
               'Payment Terms (Days)', 'Opening Balance', 'Date of Birth', 'Anniversary Date', 'Notes']
    
    # Style headers
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add sample data (row 2)
    sample_data = [
        'ABC Corporation',
        '9876543210',
        'contact@abc.com',
        '27AABCU9603R1ZM',
        '123 Business Park, Mumbai, 400001',
        'Maharashtra',
        '50000',  # Credit Limit
        '30',  # Payment Terms (Days)
        '0',  # Opening Balance
        '1990-05-15',  # Date of Birth (YYYY-MM-DD format)
        '2020-01-10',  # Anniversary Date (YYYY-MM-DD format)
        'VIP Customer'  # Notes
    ]
    
    for col_num, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col_num, value=value)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Add instructions
    ws.cell(row=4, column=1, value="INSTRUCTIONS:")
    ws.cell(row=5, column=1, value="1. Fields marked with * are required (Name, Phone)")
    ws.cell(row=6, column=1, value="2. Phone must be 10 digits")
    ws.cell(row=7, column=1, value="3. GSTIN format: 15 characters (optional)")
    ws.cell(row=8, column=1, value="4. Customer code will be auto-generated (CUST-0001, CUST-0002, etc.)")
    ws.cell(row=9, column=1, value="5. Credit Limit, Payment Terms, Opening Balance are optional (default: 0, 30, 0)")
    ws.cell(row=10, column=1, value="6. Payment Terms is in days (e.g., 30 = payment due in 30 days)")
    ws.cell(row=11, column=1, value="7. Date of Birth & Anniversary Date: Format YYYY-MM-DD (e.g., 1990-05-15) - Optional, for loyalty bonuses")
    ws.cell(row=12, column=1, value="8. Delete row 2 (sample data) before uploading")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25  # Customer Name
    ws.column_dimensions['B'].width = 15  # Phone
    ws.column_dimensions['C'].width = 25  # Email
    ws.column_dimensions['D'].width = 20  # GSTIN
    ws.column_dimensions['E'].width = 40  # Address
    ws.column_dimensions['F'].width = 15  # State
    ws.column_dimensions['G'].width = 15  # Credit Limit
    ws.column_dimensions['H'].width = 20  # Payment Terms
    ws.column_dimensions['I'].width = 18  # Opening Balance
    ws.column_dimensions['J'].width = 16  # Date of Birth
    ws.column_dimensions['K'].width = 16  # Anniversary Date
    ws.column_dimensions['L'].width = 25  # Notes
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def validate_employee_row(row_data, row_num):
    """
    Validate a single employee row
    Returns: (is_valid, error_message)
    """
    name, pin, phone, email, site_name = row_data
    
    # Required fields
    if not name or str(name).strip() == '':
        return False, f"Row {row_num}: Name is required"
    
    if not pin:
        return False, f"Row {row_num}: PIN is required"
    
    # Handle PIN (might be float from Excel, e.g., 1111.0)
    try:
        if isinstance(pin, float):
            pin = int(pin)  # Convert 1111.0 → 1111
        pin_str = str(pin).strip()
    except:
        return False, f"Row {row_num}: PIN must be a valid number"
    
    if len(pin_str) != 4 or not pin_str.isdigit():
        return False, f"Row {row_num}: PIN must be exactly 4 digits"
    
    # Phone is optional, but if provided, validate it
    if phone and str(phone).strip():
        try:
            if isinstance(phone, float):
                phone = int(phone)  # Convert 9876543210.0 → 9876543210
            phone_str = str(phone).strip()
        except:
            return False, f"Row {row_num}: Phone must be a valid number"
        
        if len(phone_str) != 10 or not phone_str.isdigit():
            return False, f"Row {row_num}: Phone must be 10 digits (or leave blank)"
    
    return True, None


def validate_inventory_row(row_data, row_num):
    """
    Validate a single inventory row
    Updated to handle new format: Item Name, SKU, Barcode, Category, Group, Unit, Stock, Reorder Point, Cost, MRP, Discount %, Selling, Tax, HSN, Description
    Returns: (is_valid, error_message)
    """
    # Unpack all 15 columns (added Reorder Point and Discount %)
    if len(row_data) < 15:
        row_data = list(row_data) + [None] * (15 - len(row_data))
    
    item_name, sku, barcode, category, group, unit, stock, reorder_point, cost_price, mrp, discount_percent, selling_price, tax_rate, hsn, description = row_data[:15]
    
    # Required fields
    if not item_name or str(item_name).strip() == '':
        return False, f"Row {row_num}: Item Name is required"
    
    if not category or str(category).strip() == '':
        return False, f"Row {row_num}: Category is required"
    
    if not group or str(group).strip() == '':
        return False, f"Row {row_num}: Group is required"
    
    if not unit or str(unit).strip() == '':
        return False, f"Row {row_num}: Unit is required"
    
    # Validate Stock Quantity
    if stock is None or (isinstance(stock, str) and stock.strip() == ''):
        return False, f"Row {row_num}: Stock Quantity is required"
    
    try:
        stock_val = float(stock)
        if stock_val < 0:
            return False, f"Row {row_num}: Stock Quantity cannot be negative"
    except (ValueError, TypeError):
        return False, f"Row {row_num}: Stock Quantity must be a number (found: {stock})"
    
    # Validate Reorder Point (optional)
    if reorder_point is not None and str(reorder_point).strip() != '':
        try:
            reorder_val = float(reorder_point)
            if reorder_val < 0:
                return False, f"Row {row_num}: Reorder Point cannot be negative"
        except (ValueError, TypeError):
            return False, f"Row {row_num}: Reorder Point must be a number"
    
    # Validate Cost Price
    if cost_price is None or (isinstance(cost_price, str) and cost_price.strip() == ''):
        return False, f"Row {row_num}: Cost Price is required"
    
    try:
        cost_val = float(cost_price)
        if cost_val < 0:
            return False, f"Row {row_num}: Cost Price cannot be negative"
    except (ValueError, TypeError):
        return False, f"Row {row_num}: Cost Price must be a number"
    
    # NEW LOGIC: At least one of (Discount % OR Selling Price) must be provided
    has_discount = discount_percent is not None and str(discount_percent).strip() != ''
    has_selling = selling_price is not None and str(selling_price).strip() != ''
    
    # Check if at least one is provided
    if not has_discount and not has_selling:
        return False, f"Row {row_num}: Either 'Discount %' or 'Selling Price' must be provided"
    
    # Validate Discount % (if provided)
    if has_discount:
        try:
            discount_val = float(discount_percent)
            if discount_val < 0:
                return False, f"Row {row_num}: Discount % cannot be negative"
            if discount_val > 100:
                return False, f"Row {row_num}: Discount % cannot exceed 100"
        except (ValueError, TypeError):
            return False, f"Row {row_num}: Discount % must be a number"
    
    # Validate Selling Price (if provided)
    if has_selling:
        try:
            selling_val = float(selling_price)
            if selling_val < 0:
                return False, f"Row {row_num}: Selling Price cannot be negative"
        except (ValueError, TypeError):
            return False, f"Row {row_num}: Selling Price must be a number"
    
    # Validate MRP (optional, but required for discount calculation)
    if mrp is not None and str(mrp).strip() != '':
        try:
            mrp_val = float(mrp)
            if mrp_val < 0:
                return False, f"Row {row_num}: MRP cannot be negative"
            
            # If both MRP and Selling Price are provided, validate Selling ≤ MRP
            if has_selling:
                selling_val = float(selling_price)
                if selling_val > mrp_val:
                    return False, f"Row {row_num}: Selling Price (₹{selling_val:.2f}) cannot exceed MRP (₹{mrp_val:.2f})"
        except (ValueError, TypeError):
            return False, f"Row {row_num}: MRP must be a number"
    
    # If discount % is provided, MRP is required for calculation
    if has_discount and not has_selling:
        if mrp is None or str(mrp).strip() == '':
            return False, f"Row {row_num}: MRP is required when providing Discount % (for calculating Selling Price)"
    
    return True, None


def validate_customer_row(row_data, row_num):
    """
    Validate a single customer row (now with DOB and Anniversary)
    Returns: (is_valid, error_message)
    """
    # Pad row_data to 12 fields if needed
    if len(row_data) < 12:
        row_data = list(row_data) + [None] * (12 - len(row_data))
    
    name, phone, email, gstin, address, state, credit_limit, payment_terms, opening_balance, date_of_birth, anniversary_date, notes = row_data[:12]
    
    # Required fields
    if not name or str(name).strip() == '':
        return False, f"Row {row_num}: Customer Name is required"
    
    if not phone:
        return False, f"Row {row_num}: Phone is required"
    
    # Handle phone as float (Excel issue)
    try:
        if isinstance(phone, float):
            phone = int(phone)
        phone_str = str(phone).strip()
    except:
        return False, f"Row {row_num}: Phone must be a valid number"
    
    if len(phone_str) != 10 or not phone_str.isdigit():
        return False, f"Row {row_num}: Phone must be 10 digits"
    
    # Validate GSTIN if provided
    if gstin and str(gstin).strip():
        gstin_str = str(gstin).strip()
        if len(gstin_str) != 15:
            return False, f"Row {row_num}: GSTIN must be 15 characters"
    
    # Validate numeric fields if provided
    if credit_limit and credit_limit != '':
        try:
            float(credit_limit)
        except:
            return False, f"Row {row_num}: Credit Limit must be a number"
    
    if payment_terms and payment_terms != '':
        try:
            int(payment_terms)
        except:
            return False, f"Row {row_num}: Payment Terms must be a number (days)"
    
    if opening_balance and opening_balance != '':
        try:
            float(opening_balance)
        except:
            return False, f"Row {row_num}: Opening Balance must be a number"
    
    return True, None


def import_employees_from_excel(file, tenant_id):
    """
    Import employees from Excel file
    Returns: (success_count, errors_list)
    """
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        success_count = 0
        errors = []
        
        # Skip header row, start from row 2
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            # Extract data (pad with None if row is shorter)
            row_data = list(row) + [None] * (8 - len(row))
            name, pin, phone, email, site_name, monthly_salary, designation, date_of_joining = row_data[:8]
            
            # Validate
            is_valid, error_msg = validate_employee_row(row_data[:5], row_num)
            if not is_valid:
                errors.append(error_msg)
                continue
            
            try:
                # Check if employee already exists (PIN is unique per tenant)
                existing = Employee.query.filter_by(
                    tenant_id=tenant_id,
                    pin=str(pin).strip()
                ).first()
                
                if existing:
                    errors.append(f"Row {row_num}: Employee with PIN {pin} already exists")
                    continue
                
                # Get or create site
                site = None
                if site_name and str(site_name).strip():
                    site = Site.query.filter_by(
                        tenant_id=tenant_id,
                        name=str(site_name).strip()
                    ).first()
                    
                    if not site:
                        site = Site(
                            tenant_id=tenant_id,
                            name=str(site_name).strip(),
                            address='',
                            latitude=0.0,
                            longitude=0.0
                        )
                        db.session.add(site)
                        db.session.flush()
                
                # Create employee (only with fields that exist in model)
                # Convert PIN and phone from float if needed
                pin_final = str(int(pin) if isinstance(pin, float) else pin).strip()
                phone_final = str(int(phone) if isinstance(phone, float) else phone).strip() if phone else None
                
                # Parse payroll fields
                salary_final = None
                if monthly_salary:
                    try:
                        salary_final = float(monthly_salary)
                    except:
                        pass  # Ignore invalid salary
                
                designation_final = str(designation).strip() if designation else None
                
                doj_final = None
                if date_of_joining:
                    try:
                        from datetime import datetime
                        if isinstance(date_of_joining, datetime):
                            doj_final = date_of_joining.date()
                        elif isinstance(date_of_joining, str):
                            doj_final = datetime.strptime(date_of_joining.strip(), '%Y-%m-%d').date()
                    except:
                        pass  # Ignore invalid date
                
                employee = Employee(
                    tenant_id=tenant_id,
                    name=str(name).strip(),
                    pin=pin_final,
                    phone=phone_final,
                    email=str(email).strip() if email else None,
                    site_id=site.id if site else None,
                    monthly_salary=salary_final,
                    designation=designation_final,
                    date_of_joining=doj_final,
                    active=True
                )
                
                db.session.add(employee)
                success_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                continue
        
        if success_count > 0:
            db.session.commit()
        
        return success_count, errors
        
    except Exception as e:
        db.session.rollback()
        return 0, [f"File error: {str(e)}"]


def import_inventory_from_excel(file, tenant_id):
    """
    Import inventory items from Excel file
    Updated to support dynamic attributes (Phase 3)
    Returns: (success_count, errors_list)
    """
    try:
        wb = load_workbook(file, data_only=True)
        
        # Find the data sheet
        ws = None
        data_sheet_names = ['Inventory Import', 'Import Data', 'Data', 'Sheet1', 'Sheet']
        
        for sheet_name in data_sheet_names:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                break
        
        if ws is None:
            for sheet in wb.worksheets:
                if 'instruction' not in sheet.title.lower():
                    ws = sheet
                    break
        
        if ws is None:
            ws = wb.active
        
        success_count = 0
        errors = []
        
        # Get configured attributes for this tenant
        from models.item_attribute import ItemAttribute, TenantAttributeConfig
        config = TenantAttributeConfig.query.filter_by(tenant_id=tenant_id).first()
        configured_attributes = []
        if config and config.is_enabled:
            configured_attributes = ItemAttribute.query.filter_by(
                tenant_id=tenant_id,
                is_active=True
            ).order_by(ItemAttribute.display_order).all()
        
        # Read header row to find column positions
        headers = {}
        for col_idx, cell in enumerate(ws[1], start=0):
            if cell.value:
                header_clean = str(cell.value).strip().replace('*', '').replace('🔶', '').strip()
                headers[header_clean] = col_idx
        
        # Define required column mappings (flexible names)
        column_map = {
            'item_name': ['Item Name', 'Item Name (Auto)', 'Product Name'],
            'sku': ['SKU'],
            'barcode': ['Barcode'],
            'category': ['Category'],
            'group': ['Group'],
            'unit': ['Unit'],
            'stock': ['Stock Quantity', 'Stock'],
            'reorder_point': ['Reorder Point'],
            'cost_price': ['Cost Price', 'Cost'],
            'mrp': ['MRP'],
            'discount_percent': ['Discount %', 'Discount'],
            'selling_price': ['Selling Price', 'Selling'],
            'tax_rate': ['Tax Rate (%)', 'Tax Rate', 'Tax'],
            'hsn': ['HSN/SAC Code', 'HSN Code', 'HSN'],
            'description': ['Description']
        }
        
        # Find column positions for each field
        col_positions = {}
        for field, possible_names in column_map.items():
            for name in possible_names:
                if name in headers:
                    col_positions[field] = headers[name]
                    break
        
        # Check if required columns exist
        required_fields = ['item_name', 'category', 'group', 'unit', 'stock', 'cost_price', 'selling_price']
        missing_fields = [f for f in required_fields if f not in col_positions]
        if missing_fields:
            return 0, [f"Missing required columns: {', '.join(missing_fields)}. Please check your Excel template."]
        
        # Process data rows (skip header row)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip completely empty rows
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            # Extract data by header position (flexible!)
            item_name = row[col_positions['item_name']] if 'item_name' in col_positions and col_positions['item_name'] < len(row) else None
            sku = row[col_positions['sku']] if 'sku' in col_positions and col_positions['sku'] < len(row) else None
            barcode = row[col_positions['barcode']] if 'barcode' in col_positions and col_positions['barcode'] < len(row) else None
            category = row[col_positions['category']] if 'category' in col_positions and col_positions['category'] < len(row) else None
            group = row[col_positions['group']] if 'group' in col_positions and col_positions['group'] < len(row) else None
            unit = row[col_positions['unit']] if 'unit' in col_positions and col_positions['unit'] < len(row) else None
            stock = row[col_positions['stock']] if 'stock' in col_positions and col_positions['stock'] < len(row) else None
            reorder_point = row[col_positions['reorder_point']] if 'reorder_point' in col_positions and col_positions['reorder_point'] < len(row) else None
            cost_price = row[col_positions['cost_price']] if 'cost_price' in col_positions and col_positions['cost_price'] < len(row) else None
            mrp = row[col_positions['mrp']] if 'mrp' in col_positions and col_positions['mrp'] < len(row) else None
            discount_percent = row[col_positions['discount_percent']] if 'discount_percent' in col_positions and col_positions['discount_percent'] < len(row) else None
            selling_price = row[col_positions['selling_price']] if 'selling_price' in col_positions and col_positions['selling_price'] < len(row) else None
            tax_rate = row[col_positions['tax_rate']] if 'tax_rate' in col_positions and col_positions['tax_rate'] < len(row) else None
            hsn = row[col_positions['hsn']] if 'hsn' in col_positions and col_positions['hsn'] < len(row) else None
            description = row[col_positions['description']] if 'description' in col_positions and col_positions['description'] < len(row) else None
            
            # IMPROVED: Skip rows where ALL required fields are empty (handles templates with formulas in empty rows)
            required_values = [item_name, category, group, unit, stock, cost_price, selling_price]
            if all(val is None or str(val).strip() == '' for val in required_values):
                continue  # Skip this row - all required fields are empty
            
            # Package for validation
            row_data = [item_name, sku, barcode, category, group, unit, stock, reorder_point, cost_price, mrp, discount_percent, selling_price, tax_rate, hsn, description]
            
            # Validate
            is_valid, error_msg = validate_inventory_row(row_data[:15], row_num)
            if not is_valid:
                errors.append(error_msg)
                continue
            
            try:
                # Get or create group
                group_obj = ItemGroup.query.filter_by(
                    tenant_id=tenant_id,
                    name=str(group).strip()
                ).first()
                
                if not group_obj:
                    group_obj = ItemGroup(
                        tenant_id=tenant_id,
                        name=str(group).strip()
                    )
                    db.session.add(group_obj)
                    db.session.flush()
                
                # Get or create category
                category_obj = ItemCategory.query.filter_by(
                    tenant_id=tenant_id,
                    name=str(category).strip(),
                    group_id=group_obj.id
                ).first()
                
                if not category_obj:
                    category_obj = ItemCategory(
                        tenant_id=tenant_id,
                        name=str(category).strip(),
                        group_id=group_obj.id
                    )
                    db.session.add(category_obj)
                    db.session.flush()
                
                # Generate SKU if not provided
                if not sku or str(sku).strip() == '':
                    # Auto-generate SKU
                    prefix = str(item_name).strip()[:3].upper()
                    count = Item.query.filter_by(tenant_id=tenant_id).count()
                    sku = f"{prefix}-{count + 1:04d}"
                else:
                    sku = str(sku).strip()
                
                # Check if item already exists
                existing = Item.query.filter_by(
                    tenant_id=tenant_id,
                    sku=sku
                ).first()
                
                if existing:
                    errors.append(f"Row {row_num}: Item with SKU {sku} already exists")
                    continue
                
                # PRIORITY-BASED PRICING LOGIC
                mrp_val = float(mrp) if mrp else None
                has_discount = discount_percent is not None and str(discount_percent).strip() != ''
                has_selling = selling_price is not None and str(selling_price).strip() != ''
                
                final_selling_price = 0.0
                final_discount_percent = 0.0
                
                # Scenario 1: Both Discount % and Selling Price provided → SELLING PRICE WINS
                if has_discount and has_selling:
                    final_selling_price = float(selling_price)
                    # Recalculate discount % from MRP and Selling Price
                    if mrp_val and mrp_val > 0:
                        final_discount_percent = ((mrp_val - final_selling_price) / mrp_val) * 100
                        final_discount_percent = round(max(0.0, min(100.0, final_discount_percent)), 2)
                
                # Scenario 2: Only Discount % provided → Calculate Selling Price from MRP
                elif has_discount and not has_selling:
                    final_discount_percent = float(discount_percent)
                    if mrp_val and mrp_val > 0:
                        final_selling_price = mrp_val * (1 - final_discount_percent / 100)
                        final_selling_price = round(final_selling_price, 2)
                
                # Scenario 3: Only Selling Price provided → Calculate Discount % from MRP
                elif has_selling and not has_discount:
                    final_selling_price = float(selling_price)
                    if mrp_val and mrp_val > 0:
                        final_discount_percent = ((mrp_val - final_selling_price) / mrp_val) * 100
                        final_discount_percent = round(max(0.0, min(100.0, final_discount_percent)), 2)
                
                # Extract reorder point (optional)
                reorder_point_val = float(reorder_point) if reorder_point else 0.0
                
                # Clean barcode - handle Excel converting numbers to floats
                barcode_clean = None
                if barcode:
                    try:
                        if isinstance(barcode, float):
                            barcode_clean = str(int(barcode))
                        else:
                            barcode_clean = str(barcode).strip()
                    except:
                        barcode_clean = str(barcode).strip() if barcode else None
                
                # Extract attribute values (Phase 3)
                attribute_data = {}
                for attr in configured_attributes:
                    if attr.attribute_name in headers:
                        col_idx = headers[attr.attribute_name]
                        if col_idx < len(row):
                            attr_value = row[col_idx]
                            if attr_value is not None and str(attr_value).strip() != '':
                                # Clean up numeric values (32.0 → 32)
                                clean_value = str(attr_value).strip()
                                try:
                                    # If it's a number like 32.0, convert to 32
                                    float_val = float(clean_value)
                                    if float_val.is_integer():
                                        clean_value = str(int(float_val))
                                except (ValueError, AttributeError):
                                    # Not a number, keep as string
                                    pass
                                attribute_data[attr.attribute_name] = clean_value
                
                # Create item
                item = Item(
                    tenant_id=tenant_id,
                    name=str(item_name).strip(),
                    sku=sku,
                    barcode=barcode_clean,
                    category_id=category_obj.id,
                    item_group_id=group_obj.id,
                    unit=str(unit).strip(),
                    opening_stock=float(stock) if stock else 0.0,
                    reorder_point=reorder_point_val,
                    cost_price=float(cost_price) if cost_price else 0.0,
                    selling_price=final_selling_price,
                    mrp=mrp_val,
                    discount_percent=final_discount_percent,
                    hsn_code=str(hsn).strip() if hsn else None,
                    gst_rate=float(tax_rate) if tax_rate else 18.0,
                    tax_preference=f"GST {tax_rate}%" if tax_rate else "GST 18%",
                    sales_description=str(description).strip() if description else '',
                    purchase_description=str(description).strip() if description else '',
                    attribute_data=attribute_data if attribute_data else None  # Phase 3
                )
                
                db.session.add(item)
                db.session.flush()  # Get item.id
                
                # Create stock record for default site
                from models import Site, ItemStock
                # Get default site (marked as is_default=True)
                default_site = Site.query.filter_by(
                    tenant_id=tenant_id,
                    is_default=True,
                    active=True
                ).first()
                
                # Fallback to first active site if no default is set
                if not default_site:
                    default_site = Site.query.filter_by(tenant_id=tenant_id, active=True).first()
                
                if default_site and item.track_inventory:
                    item_stock = ItemStock(
                        tenant_id=tenant_id,
                        item_id=item.id,
                        site_id=default_site.id,
                        quantity_available=float(stock) if stock else 0.0,
                        stock_value=(float(stock) if stock else 0.0) * (float(cost_price) if cost_price else 0.0)
                    )
                    db.session.add(item_stock)
                
                success_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                continue
        
        if success_count > 0:
            # ✅ CRITICAL FIX: Create double-entry accounting entries for inventory import
            # This ensures trial balance stays balanced automatically!
            try:
                from models.bank_account import AccountTransaction
                from models import ItemStock
                from datetime import datetime
                import pytz
                
                ist = pytz.timezone('Asia/Kolkata')
                current_date = datetime.now(ist).date()
                
                # Calculate total inventory value imported
                total_inventory_value = 0.0
                inventory_stocks = ItemStock.query.filter_by(tenant_id=tenant_id).all()
                for stock in inventory_stocks:
                    total_inventory_value += stock.stock_value or 0.0
                
                # Check if opening equity entry already exists for today
                existing_equity = AccountTransaction.query.filter_by(
                    tenant_id=tenant_id,
                    transaction_type='opening_balance_inventory_equity',
                    transaction_date=current_date
                ).first()
                
                if not existing_equity and total_inventory_value > 0:
                    # Generate voucher number
                    voucher_number = f"OB-INV-{tenant_id}-{current_date.strftime('%Y%m%d')}"
                    
                    # DEBIT: Inventory (Asset) - increases asset
                    inventory_debit = AccountTransaction(
                        tenant_id=tenant_id,
                        account_id=None,  # Special system account
                        transaction_type='inventory_opening_debit',
                        transaction_date=current_date,
                        debit_amount=total_inventory_value,
                        credit_amount=0.0,
                        description=f"Bulk Import - Inventory Opening (₹{total_inventory_value:,.2f} worth of stock)",
                        voucher_number=voucher_number,
                        created_at=datetime.now(ist)
                    )
                    
                    # CREDIT: Owner's Capital - Inventory Opening (Equity) - increases equity
                    equity_credit = AccountTransaction(
                        tenant_id=tenant_id,
                        account_id=None,  # Special system account
                        transaction_type='opening_balance_inventory_equity',
                        transaction_date=current_date,
                        debit_amount=0.0,
                        credit_amount=total_inventory_value,
                        description=f"Opening Balance - Inventory Equity (₹{total_inventory_value:,.2f} worth of stock)",
                        voucher_number=voucher_number,
                        created_at=datetime.now(ist)
                    )
                    
                    db.session.add(inventory_debit)
                    db.session.add(equity_credit)
                    
                    print(f"✅ Created accounting entries for inventory import: ₹{total_inventory_value:,.2f}")
                
            except Exception as e:
                print(f"⚠️ Warning: Could not create accounting entries: {str(e)}")
                # Don't fail the import if accounting entries fail
                pass
            
            db.session.commit()
        
        return success_count, errors
        
    except Exception as e:
        db.session.rollback()
        return 0, [f"File error: {str(e)}"]


def import_customers_from_excel(file, tenant_id):
    """
    Import customers from Excel file
    Returns: (success_count, errors_list)
    """
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        success_count = 0
        errors = []
        
        # Skip header row, start from row 2
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            # Extract data - match new template format (now with DOB and Anniversary)
            row_data = list(row) + [None] * (12 - len(row))
            name, phone, email, gstin, address, state, credit_limit, payment_terms, opening_balance, date_of_birth, anniversary_date, notes = row_data[:12]
            
            # Validate
            is_valid, error_msg = validate_customer_row(row_data[:12], row_num)
            if not is_valid:
                errors.append(error_msg)
                continue
            
            try:
                # Check if customer already exists
                existing = Customer.query.filter_by(
                    tenant_id=tenant_id,
                    phone=str(phone).strip()
                ).first()
                
                if existing:
                    errors.append(f"Row {row_num}: Customer with phone {phone} already exists")
                    continue
                
                # Generate customer code
                last_customer = Customer.query.filter_by(tenant_id=tenant_id).order_by(Customer.id.desc()).first()
                if last_customer and last_customer.customer_code:
                    last_num = int(last_customer.customer_code.split('-')[-1])
                    next_num = last_num + 1
                else:
                    next_num = 1
                customer_code = f"CUST-{next_num:04d}"
                
                # Handle phone as float (Excel issue)
                phone_final = str(int(phone) if isinstance(phone, float) else phone).strip()
                
                # Parse dates (format: YYYY-MM-DD or DD-MM-YYYY or MM/DD/YYYY)
                from datetime import datetime
                
                dob_final = None
                if date_of_birth and str(date_of_birth).strip():
                    try:
                        dob_str = str(date_of_birth).strip()
                        # Try YYYY-MM-DD format first
                        if '-' in dob_str and len(dob_str.split('-')[0]) == 4:
                            dob_final = datetime.strptime(dob_str, '%Y-%m-%d').date()
                        # Try DD-MM-YYYY
                        elif '-' in dob_str:
                            dob_final = datetime.strptime(dob_str, '%d-%m-%Y').date()
                        # Try MM/DD/YYYY
                        elif '/' in dob_str:
                            dob_final = datetime.strptime(dob_str, '%m/%d/%Y').date()
                    except:
                        pass  # If parsing fails, leave as None
                
                anniversary_final = None
                if anniversary_date and str(anniversary_date).strip():
                    try:
                        ann_str = str(anniversary_date).strip()
                        if '-' in ann_str and len(ann_str.split('-')[0]) == 4:
                            anniversary_final = datetime.strptime(ann_str, '%Y-%m-%d').date()
                        elif '-' in ann_str:
                            anniversary_final = datetime.strptime(ann_str, '%d-%m-%Y').date()
                        elif '/' in ann_str:
                            anniversary_final = datetime.strptime(ann_str, '%m/%d/%Y').date()
                    except:
                        pass
                
                # Create customer with all fields
                customer = Customer(
                    tenant_id=tenant_id,
                    customer_code=customer_code,
                    name=str(name).strip(),
                    phone=phone_final,
                    email=str(email).strip() if email else '',
                    gstin=str(gstin).strip() if gstin else '',
                    address=str(address).strip() if address else '',
                    state=str(state).strip() if state else '',
                    credit_limit=float(credit_limit) if credit_limit else 0,
                    payment_terms_days=int(payment_terms) if payment_terms else 30,
                    opening_balance=float(opening_balance) if opening_balance else 0,
                    date_of_birth=dob_final,
                    anniversary_date=anniversary_final,
                    notes=str(notes).strip() if notes else ''
                )
                
                db.session.add(customer)
                success_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                continue
        
        if success_count > 0:
            db.session.commit()
        
        return success_count, errors
        
    except Exception as e:
        db.session.rollback()
        return 0, [f"File error: {str(e)}"]


# ============================================
# SUBSCRIPTION ENROLLMENT IMPORT
# ============================================

def create_subscription_enrollment_template(tenant_id):
    """
    Create Excel template for subscription enrollment bulk import
    This is DYNAMIC - includes tenant's plans and customers for reference
    
    Returns: BytesIO object with Excel file
    """
    wb = Workbook()
    
    # ==========================================
    # SHEET 1: Enrollment Data (User fills this)
    # ==========================================
    ws_data = wb.active
    ws_data.title = "Enrollment Data"
    
    # Headers
    headers = ['Customer Phone*', 'Customer Name (optional)', 'Plan ID*', 'Start Date*', 'Quantity', 'Auto Renew']
    
    # Style headers
    header_fill = PatternFill(start_color="7B68EE", end_color="7B68EE", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add sample data (row 2)
    sample_data = [
        '9876543210',
        '(auto-matched)',
        '1',
        datetime.now().strftime('%Y-%m-%d'),
        '2',
        'Yes'
    ]
    
    for col_num, value in enumerate(sample_data, 1):
        cell = ws_data.cell(row=2, column=col_num, value=value)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Instructions in Sheet 1
    ws_data.cell(row=4, column=1, value="INSTRUCTIONS:")
    ws_data.cell(row=5, column=1, value="1. Fields marked with * are required")
    ws_data.cell(row=6, column=1, value="2. Customer Phone: 10-digit phone number (must exist in system)")
    ws_data.cell(row=7, column=1, value="3. Customer Name: Optional - for your reference only")
    ws_data.cell(row=8, column=1, value="4. Plan ID: Look up from 'Your Plans' sheet (DO NOT use plan name)")
    ws_data.cell(row=9, column=1, value="5. Start Date: Format YYYY-MM-DD (e.g., 2024-12-01)")
    ws_data.cell(row=10, column=1, value="6. Quantity: For metered plans (liters/day). Leave blank for fixed plans.")
    ws_data.cell(row=11, column=1, value="7. Auto Renew: 'Yes' or 'No' (default: Yes)")
    ws_data.cell(row=12, column=1, value="8. Delete row 2 (sample data) before uploading")
    ws_data.cell(row=13, column=1, value="")
    ws_data.cell(row=14, column=1, value="DUPLICATE HANDLING:")
    ws_data.cell(row=15, column=1, value="   - Same customer + Same plan = SKIPPED (no duplicate created)")
    ws_data.cell(row=16, column=1, value="   - Same customer + Different plan = ENROLLED (customer can have multiple)")
    
    # Style instructions
    for row in range(4, 17):
        cell = ws_data.cell(row=row, column=1)
        if 'INSTRUCTIONS' in str(cell.value) or 'DUPLICATE' in str(cell.value):
            cell.font = Font(bold=True, color="7B68EE")
        else:
            cell.font = Font(color="666666")
    
    # Adjust column widths
    ws_data.column_dimensions['A'].width = 18
    ws_data.column_dimensions['B'].width = 25
    ws_data.column_dimensions['C'].width = 10
    ws_data.column_dimensions['D'].width = 15
    ws_data.column_dimensions['E'].width = 12
    ws_data.column_dimensions['F'].width = 12
    
    # ==========================================
    # SHEET 2: Your Plans (Reference)
    # ==========================================
    ws_plans = wb.create_sheet("Your Plans (Reference)")
    
    # Headers
    plan_headers = ['Plan ID', 'Plan Name', 'Type', 'Price/Rate', 'Unit', 'Duration', 'Delivery Pattern']
    plan_header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    
    for col_num, header in enumerate(plan_headers, 1):
        cell = ws_plans.cell(row=1, column=col_num, value=header)
        cell.fill = plan_header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Get tenant's plans
    plans = SubscriptionPlan.query.filter_by(tenant_id=tenant_id, is_active=True).order_by(SubscriptionPlan.id).all()
    
    for row_num, plan in enumerate(plans, 2):
        ws_plans.cell(row=row_num, column=1, value=plan.id)
        ws_plans.cell(row=row_num, column=2, value=plan.name)
        ws_plans.cell(row=row_num, column=3, value=plan.plan_type.upper())
        
        if plan.plan_type == 'metered':
            ws_plans.cell(row=row_num, column=4, value=f"Rs.{plan.unit_rate}/unit")
            ws_plans.cell(row=row_num, column=5, value=plan.unit_name or 'unit')
            ws_plans.cell(row=row_num, column=6, value="Monthly billing")
            ws_plans.cell(row=row_num, column=7, value=plan.delivery_pattern or 'daily')
        else:
            ws_plans.cell(row=row_num, column=4, value=f"Rs.{plan.price}")
            ws_plans.cell(row=row_num, column=5, value="-")
            ws_plans.cell(row=row_num, column=6, value=plan.duration_display)
            ws_plans.cell(row=row_num, column=7, value="-")
    
    # Add note at the bottom
    note_row = len(plans) + 3
    ws_plans.cell(row=note_row, column=1, value="NOTE: Use the Plan ID (first column) in your enrollment data, NOT the plan name.")
    ws_plans.cell(row=note_row, column=1).font = Font(bold=True, color="FF6600")
    ws_plans.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=7)
    
    # Adjust column widths
    ws_plans.column_dimensions['A'].width = 10
    ws_plans.column_dimensions['B'].width = 40
    ws_plans.column_dimensions['C'].width = 10
    ws_plans.column_dimensions['D'].width = 15
    ws_plans.column_dimensions['E'].width = 10
    ws_plans.column_dimensions['F'].width = 18
    ws_plans.column_dimensions['G'].width = 18
    
    # ==========================================
    # SHEET 3: Your Customers (Reference)
    # ==========================================
    ws_customers = wb.create_sheet("Your Customers (Reference)")
    
    # Headers
    customer_headers = ['Phone', 'Customer Name', 'Email', 'Existing Subscriptions']
    customer_header_fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
    
    for col_num, header in enumerate(customer_headers, 1):
        cell = ws_customers.cell(row=1, column=col_num, value=header)
        cell.fill = customer_header_fill
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Get tenant's customers
    customers = Customer.query.filter_by(tenant_id=tenant_id).order_by(Customer.name).all()
    
    for row_num, customer in enumerate(customers, 2):
        ws_customers.cell(row=row_num, column=1, value=customer.phone)
        ws_customers.cell(row=row_num, column=2, value=customer.name)
        ws_customers.cell(row=row_num, column=3, value=customer.email or '-')
        
        # Check existing subscriptions
        existing_subs = CustomerSubscription.query.filter_by(
            customer_id=customer.id,
            status='active'
        ).all()
        
        if existing_subs:
            sub_names = [sub.plan.name for sub in existing_subs]
            ws_customers.cell(row=row_num, column=4, value=', '.join(sub_names))
            ws_customers.cell(row=row_num, column=4).font = Font(color="4CAF50")
        else:
            ws_customers.cell(row=row_num, column=4, value="No active subscriptions")
            ws_customers.cell(row=row_num, column=4).font = Font(color="999999")
    
    # Add note
    note_row = len(customers) + 3
    ws_customers.cell(row=note_row, column=1, value="NOTE: Use the Phone number (first column) in your enrollment data to match customers.")
    ws_customers.cell(row=note_row, column=1).font = Font(bold=True, color="FF6600")
    ws_customers.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
    
    # Adjust column widths
    ws_customers.column_dimensions['A'].width = 15
    ws_customers.column_dimensions['B'].width = 30
    ws_customers.column_dimensions['C'].width = 30
    ws_customers.column_dimensions['D'].width = 40
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def validate_subscription_enrollment_row(row_data, row_num, tenant_id):
    """
    Validate a single subscription enrollment row
    Returns: (is_valid, error_message, parsed_data)
    """
    phone, customer_name, plan_id, start_date, quantity, auto_renew = row_data
    
    parsed_data = {}
    
    # Required: Phone
    if not phone:
        return False, f"Row {row_num}: Customer Phone is required", None
    
    # Handle phone as float (Excel issue)
    try:
        if isinstance(phone, float):
            phone = int(phone)
        phone_str = str(phone).strip()
    except:
        return False, f"Row {row_num}: Phone must be a valid number", None
    
    if len(phone_str) != 10 or not phone_str.isdigit():
        return False, f"Row {row_num}: Phone must be 10 digits", None
    
    # Find customer by phone
    customer = Customer.query.filter_by(tenant_id=tenant_id, phone=phone_str).first()
    if not customer:
        return False, f"Row {row_num}: Customer with phone {phone_str} not found", None
    
    parsed_data['customer_id'] = customer.id
    parsed_data['customer_name'] = customer.name
    
    # Required: Plan ID
    if not plan_id:
        return False, f"Row {row_num}: Plan ID is required", None
    
    try:
        if isinstance(plan_id, float):
            plan_id = int(plan_id)
        plan_id_int = int(plan_id)
    except:
        return False, f"Row {row_num}: Plan ID must be a number", None
    
    # Find plan
    plan = SubscriptionPlan.query.filter_by(id=plan_id_int, tenant_id=tenant_id, is_active=True).first()
    if not plan:
        return False, f"Row {row_num}: Plan with ID {plan_id_int} not found or inactive", None
    
    parsed_data['plan_id'] = plan.id
    parsed_data['plan'] = plan
    
    # Required: Start Date
    if not start_date:
        return False, f"Row {row_num}: Start Date is required", None
    
    try:
        if isinstance(start_date, datetime):
            parsed_data['start_date'] = start_date.date()
        elif isinstance(start_date, str):
            # Try multiple date formats
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y']:
                try:
                    parsed_data['start_date'] = datetime.strptime(start_date.strip(), fmt).date()
                    break
                except:
                    continue
            else:
                return False, f"Row {row_num}: Invalid date format. Use YYYY-MM-DD", None
        else:
            return False, f"Row {row_num}: Invalid start date", None
    except Exception as e:
        return False, f"Row {row_num}: Invalid start date - {str(e)}", None
    
    # Optional: Quantity (for metered plans)
    if plan.plan_type == 'metered':
        if quantity and str(quantity).strip():
            try:
                parsed_data['quantity'] = float(quantity)
            except:
                return False, f"Row {row_num}: Quantity must be a number", None
        else:
            parsed_data['quantity'] = 1.0  # Default quantity for metered
    else:
        parsed_data['quantity'] = None
    
    # Optional: Auto Renew
    if auto_renew:
        auto_renew_str = str(auto_renew).strip().lower()
        parsed_data['auto_renew'] = auto_renew_str in ['yes', 'y', 'true', '1']
    else:
        parsed_data['auto_renew'] = True  # Default to auto-renew
    
    return True, None, parsed_data


def import_subscription_enrollments_from_excel(file, tenant_id):
    """
    Import subscription enrollments from Excel file
    Returns: (success_count, skipped_count, errors_list)
    
    Duplicate handling:
    - Same customer + Same plan = SKIP (already enrolled)
    - Same customer + Different plan = CREATE (multiple subscriptions allowed)
    """
    # Import here to avoid circular imports
    from routes.subscriptions import should_deliver_on_date
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        success_count = 0
        skipped_count = 0
        errors = []
        
        # Skip header row, start from row 2
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            # Extract data (pad with None if row is shorter)
            row_data = list(row) + [None] * (6 - len(row))
            phone, customer_name, plan_id, start_date, quantity, auto_renew = row_data[:6]
            
            # Validate
            is_valid, error_msg, parsed_data = validate_subscription_enrollment_row(row_data[:6], row_num, tenant_id)
            if not is_valid:
                errors.append(error_msg)
                continue
            
            try:
                customer_id = parsed_data['customer_id']
                plan = parsed_data['plan']
                
                # Check for duplicate: same customer + same plan (active or pending)
                existing = CustomerSubscription.query.filter(
                    CustomerSubscription.customer_id == customer_id,
                    CustomerSubscription.plan_id == plan.id,
                    CustomerSubscription.status.in_(['active', 'pending_payment'])
                ).first()
                
                if existing:
                    skipped_count += 1
                    errors.append(f"Row {row_num}: {parsed_data['customer_name']} already enrolled in {plan.name} - SKIPPED")
                    continue
                
                # Calculate period end based on plan type
                start_date = parsed_data['start_date']
                
                if plan.plan_type == 'metered':
                    # METERED: End on last day of start month
                    if start_date.month == 12:
                        period_end = datetime(start_date.year, 12, 31).date()
                    else:
                        period_end = datetime(start_date.year, start_date.month + 1, 1).date() - timedelta(days=1)
                else:
                    # FIXED: Add duration days
                    period_end = start_date + timedelta(days=plan.duration_days)
                
                # Create subscription
                subscription = CustomerSubscription(
                    tenant_id=tenant_id,
                    customer_id=customer_id,
                    plan_id=plan.id,
                    default_quantity=parsed_data['quantity'],
                    start_date=start_date,
                    current_period_start=start_date,
                    current_period_end=period_end,
                    status='active',
                    auto_renew=parsed_data['auto_renew']
                )
                
                db.session.add(subscription)
                db.session.flush()  # Get subscription ID
                
                # AUTO-GENERATE DELIVERIES for metered plans
                if plan.plan_type == 'metered':
                    current_date = start_date
                    delivery_pattern = plan.delivery_pattern or 'daily'
                    custom_days = plan.custom_days
                    
                    while current_date <= period_end:
                        if should_deliver_on_date(current_date, delivery_pattern, custom_days, start_date):
                            # Get customer for delivery assignment
                            customer = Customer.query.get(customer_id)
                            
                            delivery = SubscriptionDelivery(
                                tenant_id=tenant_id,
                                subscription_id=subscription.id,
                                delivery_date=current_date,
                                quantity=parsed_data['quantity'],
                                rate=plan.unit_rate,
                                amount=parsed_data['quantity'] * float(plan.unit_rate),
                                status='delivered',
                                is_modified=False,
                                assigned_to=customer.default_delivery_employee if customer else None
                            )
                            db.session.add(delivery)
                        
                        current_date += timedelta(days=1)
                
                success_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                continue
        
        if success_count > 0:
            db.session.commit()
        
        return success_count, skipped_count, errors
        
    except Exception as e:
        db.session.rollback()
        return 0, 0, [f"File error: {str(e)}"]

