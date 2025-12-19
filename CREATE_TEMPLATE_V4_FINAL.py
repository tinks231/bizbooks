"""
🎯 Clothing Template v4 - FINAL VERSION
========================================

Changes from v3:
- Reduced formula/dropdown range: 10,000 → 100 rows
- Better for performance and import
- Users can copy formulas if they need more rows

Version: 4.0 (Production-ready!)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

def create_final_template():
    """
    Create template with reasonable row limit (100 rows)
    """
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Import"
    
    print("🎨 Creating FINAL template (v4)...")
    
    # Column structure
    columns = [
        ("Brand*", "Brand name: Levi's, Nike, Adidas", True, "input"),
        ("Product Name*", "Model/Style: 501, Dri-FIT, Air Max", True, "input"),
        ("Size*", "Dropdown: 28, 30, 32, S, M, L, XL", True, "input"),
        ("Color*", "Dropdown: Blue, Black, Red, White", True, "input"),
        ("Style", "Optional: Slim Fit, Regular Fit, Casual", False, "input"),
        ("🔶 Item Name*", "Auto: Brand + Category + Product + Size + Color + Style", True, "auto"),
        ("SKU", "Optional: Leave blank to auto-generate", False, "optional"),
        ("Barcode", "Optional: Leave blank to auto-generate", False, "optional"),
        ("Category*", "Product type: Jeans, T-Shirts, Shirts, Shoes", True, "input"),
        ("Group*", "Department: Men's Wear, Women's Wear, Kids Wear", True, "input"),
        ("Unit*", "Unit: pc (pieces), pair, set", True, "input"),
        ("Stock Quantity*", "Current stock in inventory", True, "input"),
        ("Reorder Point", "Optional: Low stock alert threshold", False, "optional"),
        ("Cost Price*", "Your purchase/cost price", True, "input"),
        ("MRP", "Optional: Maximum Retail Price", False, "optional"),
        ("Discount %", "Auto-calculated from MRP and Selling Price", False, "auto"),
        ("Selling Price*", "Final price to customer", True, "input"),
        ("Tax Rate (%)", "GST: 0, 5, 12, 18, 28", False, "optional"),
        ("HSN/SAC Code", "Optional: For GST compliance", False, "optional"),
        ("Description", "Optional: Additional details", False, "optional"),
    ]
    
    # Create headers
    for col_idx, (name, description, required, col_type) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = name
        
        if required and col_type == "input":
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
        elif col_type == "auto":
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.font = Font(color="000000", bold=True, size=11)
        else:
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
        
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.comment = Comment(description, "BizBooks")
    
    # Set column widths
    widths = [18, 20, 10, 12, 15, 45, 15, 15, 15, 15, 10, 15, 12, 12, 12, 12, 15, 10, 15, 25]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Add dropdowns (ONLY 100 rows this time!)
    brand_values = '"Levi\'s,Nike,Adidas,Puma,Reebok,H&M,Zara,Mango,Pepe Jeans,Wrangler,Lee,Allen Solly,Peter England,Van Heusen,Arrow,Other"'
    brand_dv = DataValidation(type="list", formula1=brand_values, allow_blank=False)
    ws.add_data_validation(brand_dv)
    brand_dv.add('A2:A101')  # Only 100 rows!
    
    size_values = '"28,30,32,34,36,38,40,42,44,46,XS,S,M,L,XL,XXL,XXXL,6,7,8,9,10,11,12,Free Size"'
    size_dv = DataValidation(type="list", formula1=size_values, allow_blank=False)
    ws.add_data_validation(size_dv)
    size_dv.add('C2:C101')
    
    color_values = '"Blue,Black,White,Red,Green,Yellow,Grey,Brown,Pink,Purple,Orange,Navy,Maroon,Beige,Cream,Olive,Khaki,Indigo"'
    color_dv = DataValidation(type="list", formula1=color_values, allow_blank=False)
    ws.add_data_validation(color_dv)
    color_dv.add('D2:D101')
    
    style_values = '"Slim Fit,Regular Fit,Loose Fit,Relaxed Fit,Athletic Fit,Straight Fit,Bootcut,Skinny,Casual,Formal,Party,Sports"'
    style_dv = DataValidation(type="list", formula1=style_values, allow_blank=True)
    ws.add_data_validation(style_dv)
    style_dv.add('E2:E101')
    
    category_values = '"Jeans,T-Shirts,Shirts,Trousers,Shorts,Jackets,Sweaters,Hoodies,Shoes,Sandals,Sneakers,Formal Shoes,Accessories,Belts,Wallets,Caps"'
    category_dv = DataValidation(type="list", formula1=category_values, allow_blank=False)
    ws.add_data_validation(category_dv)
    category_dv.add('I2:I101')
    
    group_values = '"Men\'s Wear,Women\'s Wear,Boys Wear,Girls Wear,Kids Wear,Unisex,Accessories"'
    group_dv = DataValidation(type="list", formula1=group_values, allow_blank=False)
    ws.add_data_validation(group_dv)
    group_dv.add('J2:J101')
    
    unit_values = '"pc,pair,set,kg,gm,ltr,ml,dozen"'
    unit_dv = DataValidation(type="list", formula1=unit_values, allow_blank=False)
    ws.add_data_validation(unit_dv)
    unit_dv.add('K2:K101')
    
    tax_values = '"0,5,12,18,28"'
    tax_dv = DataValidation(type="list", formula1=tax_values, allow_blank=True)
    ws.add_data_validation(tax_dv)
    tax_dv.add('R2:R101')
    
    print("✅ Dropdowns added (100 rows)")
    
    # Add formulas (ONLY 100 rows!)
    for row in range(2, 102):  # Rows 2-101 = 100 rows
        ws[f'F{row}'].value = f'=TRIM(A{row}&" "&I{row}&" "&B{row}&" "&C{row}&" "&D{row}&IF(E{row}<>"", " "&E{row}, ""))'
        ws[f'F{row}'].font = Font(color="FF6B00", bold=True)
        
        ws[f'P{row}'].value = f'=IF(AND(O{row}>0,Q{row}>0,Q{row}<=O{row}),ROUND((O{row}-Q{row})/O{row}*100,2),0)'
        ws[f'P{row}'].number_format = '0.00'
    
    print("✅ Formulas added (100 rows)")
    
    # Add example data
    examples = [
        ["Levi's", "501", "32", "Blue", "Slim Fit", "", "", "", "Jeans", "Men's Wear", "pc", 15, 5, 1200, 3999, "", 2499, 12, "62034090", "Classic jeans"],
        ["Levi's", "501", "34", "Blue", "Slim Fit", "", "", "", "Jeans", "Men's Wear", "pc", 12, 5, 1200, 3999, "", 2499, 12, "62034090", ""],
        ["Levi's", "501", "32", "Black", "Regular Fit", "", "", "", "Jeans", "Men's Wear", "pc", 8, 5, 1200, 3999, "", 2499, 12, "62034090", ""],
        ["Nike", "Dri-FIT", "M", "Red", "", "", "", "", "T-Shirts", "Unisex", "pc", 50, 10, 400, 1299, "", 799, 12, "61091000", ""],
        ["Nike", "Dri-FIT", "L", "Red", "", "", "", "", "T-Shirts", "Unisex", "pc", 30, 10, 400, 1299, "", 799, 12, "61091000", ""],
        ["Adidas", "Superstar", "10", "White", "Casual", "", "", "", "Shoes", "Men's Wear", "pair", 8, 3, 2500, 5999, "", 3999, 18, "64039900", ""],
    ]
    
    for row_idx, example in enumerate(examples, start=2):
        col_idx = 1
        for value in example:
            if col_idx not in [6, 16]:
                ws.cell(row=row_idx, column=col_idx, value=value)
            col_idx += 1
        
        for col in range(1, 21):
            ws.cell(row=row_idx, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    print("✅ Example data added")
    
    # Add instructions
    ws_inst = wb.create_sheet("📖 Instructions", 0)
    
    instructions = [
        ["🎯 CLOTHING RETAIL IMPORT TEMPLATE v4 (FINAL)", ""],
        ["", ""],
        ["✅ Ready to use with improved performance!", ""],
        ["", ""],
        ["📝 WHAT'S NEW IN v4", ""],
        ["", ""],
        ["✅ Optimized for 100 rows (was 10,000 before)", ""],
        ["✅ Faster Excel performance", ""],
        ["✅ Faster import (no 9999 empty row errors!)", ""],
        ["✅ If you need more rows, just copy formulas down", ""],
        ["", ""],
        ["💡 HOW TO ADD MORE ROWS", ""],
        ["", ""],
        ["If you have more than 100 items:", ""],
        ["  1. Fill rows 2-101 with your data", ""],
        ["  2. Select row 101 (entire row)", ""],
        ["  3. Copy it (Ctrl+C or Cmd+C)", ""],
        ["  4. Select rows 102-200", ""],
        ["  5. Paste (Ctrl+V or Cmd+V)", ""],
        ["  6. Formulas and dropdowns copy automatically! ✅", ""],
        ["", ""],
        ["📋 HOW TO USE", ""],
        ["", ""],
        ["STEP 1: Fill structured columns", ""],
        ["  A. Brand* → Levi's, Nike, Adidas", ""],
        ["  B. Product Name* → 501, Dri-FIT", ""],
        ["  C. Size* → 28, 30, S, M, L (dropdown)", ""],
        ["  D. Color* → Blue, Black, Red (dropdown)", ""],
        ["  E. Style → Slim Fit, Regular Fit (optional)", ""],
        ["", ""],
        ["STEP 2: Item Name Auto-Generates (Column F)", ""],
        ["  Example: \"Levi's Jeans 501 32 Blue Slim Fit\"", ""],
        ["", ""],
        ["STEP 3: Fill standard columns", ""],
        ["  I. Category* → Jeans, T-Shirts (dropdown)", ""],
        ["  J. Group* → Men's Wear, Women's Wear (dropdown)", ""],
        ["  K. Unit* → pc, pair (dropdown)", ""],
        ["  L. Stock Quantity* → Current stock", ""],
        ["  N. Cost Price* → Your cost", ""],
        ["  Q. Selling Price* → Customer price", ""],
        ["", ""],
        ["⚡ DRAG-FILL TRICK", ""],
        ["", ""],
        ["For 80 items:", ""],
        ["  1. A2: \"Levi's\" → Drag to A81 (2 seconds)", ""],
        ["  2. B2: \"501\" → Drag to B81 (2 seconds)", ""],
        ["  3. I2: \"Jeans\" → Drag to I81 (2 seconds)", ""],
        ["  4. Fill Size, Color, Stock individually", ""],
        ["  Total: 5-6 minutes! 🚀", ""],
        ["", ""],
        ["💾 SAVE AND UPLOAD", ""],
        ["", ""],
        ["  1. Delete example rows 2-7 (grey rows)", ""],
        ["  2. Fill your data", ""],
        ["  3. Save as .xlsx", ""],
        ["  4. Upload to BizBooks", ""],
        ["  5. Success! ✅", ""],
        ["", ""],
        ["📞 SUPPORT", ""],
        ["  📱 +91 8983121201", ""],
        ["  📧 bizbooks.notifications@gmail.com", ""],
    ]
    
    for row_idx, (text, value) in enumerate(instructions, start=1):
        ws_inst[f'A{row_idx}'].value = text
        if any(emoji in text for emoji in ["🎯", "📝", "💡", "📋", "⚡", "💾", "📞"]):
            ws_inst[f'A{row_idx}'].font = Font(bold=True, size=14, color="0000FF")
        elif "STEP" in text:
            ws_inst[f'A{row_idx}'].font = Font(bold=True, size=12, color="C00000")
        ws_inst[f'A{row_idx}'].alignment = Alignment(wrap_text=True, vertical='top')
    
    ws_inst.column_dimensions['A'].width = 90
    
    print("✅ Instructions created")
    
    filename = "BizBooks_Clothing_Template_v4_FINAL.xlsx"
    wb.save(filename)
    
    return filename


if __name__ == "__main__":
    print("=" * 70)
    print("🚀 CREATING FINAL TEMPLATE v4")
    print("=" * 70)
    print()
    
    output_file = create_final_template()
    
    print()
    print("=" * 70)
    print("✅ SUCCESS!")
    print("=" * 70)
    print()
    print(f"📁 File: {output_file}")
    print()
    print("🎯 Key Improvements:")
    print("  ✅ Only 100 rows with formulas/dropdowns (was 10,000)")
    print("  ✅ Faster Excel performance")
    print("  ✅ No more 9999 empty row errors!")
    print("  ✅ Can copy formulas if you need more rows")
    print()
    print("💡 Why This is Better:")
    print("  Before: 10,000 empty rows → 9999 import errors ❌")
    print("  After: 100 rows → Only checks actual data ✅")
    print()
    print("🚀 Ready to use!")
    print()
    print("=" * 70)

